from io import BytesIO
from pathlib import Path
from unittest.mock import patch
from uuid import uuid4

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from doubao2api import platform_editor as _pe
from doubao2api import server as _server_module
from doubao2api.config import RuntimeConfig, SettingsStore
from doubao2api.server import create_app


def make_client(tmp_path: Path) -> TestClient:
    return TestClient(
        create_app(
            store=SettingsStore(tmp_path),
            runtime=RuntimeConfig(open_admin_browser=False),
        ),
        client=("127.0.0.1", 50000),
    )


def _add_platform_entries_without_persist(rows: list[dict[str, str]]) -> dict[str, int]:
    """Mock add_entries to update in-memory rules without writing source file."""
    added = 0
    for raw in rows:
        domain = str(raw.get("url") or raw.get("域名") or "").strip()
        name = str(raw.get("平台名") or raw.get("平台名称") or "").strip()
        category = str(raw.get("平台类型") or raw.get("类型") or "").strip()
        if not domain or not name or not category:
            continue
        _pe.PLATFORM_ENTRIES.append({"domain": domain, "name": name, "category": category})
        added += 1
    _pe._refresh_derived_mappings()
    return {"added": added, "ignored": len(rows) - added}


def _remove_platform_entry(domain: str) -> None:
    _pe.PLATFORM_ENTRIES[:] = [e for e in _pe.PLATFORM_ENTRIES if e["domain"] != domain]
    _pe._refresh_derived_mappings()


TEST_DOMAIN = "test-source.local"
TEST_LINK = f"https://{TEST_DOMAIN}/report"


def keyword_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["关键词"])
    sheet.append(["新能源汽车"])
    sheet.append(["家庭储能"])
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def platform_workbook(domain: str = TEST_DOMAIN) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["域名", "平台名", "平台类型"])
    sheet.append([domain, "测试平台", "企业官网/品牌站"])
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _create_job_with_result(
    client: TestClient,
    *,
    domain: str = TEST_DOMAIN,
    link: str = TEST_LINK,
) -> tuple[str, str]:
    imported = client.post(
        "/admin/api/research/import",
        files={
            "file": (
                "keywords.xlsx",
                keyword_workbook(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert imported.status_code == 200

    created = client.post(
        "/admin/api/research/jobs",
        json={
            "name": "测试任务",
            "keywords": imported.json()["keywords"],
            "prompt_template": "请调研 {keyword}",
            "scheduled_at": "2099-01-01T00:00:00+08:00",
            "interval_seconds": 60,
            "account_cooldown_seconds": 120,
        },
    )
    assert created.status_code == 200
    job_id = created.json()["id"]

    store = client.app.state.research_store
    task = store.get_job(job_id)["tasks"][0]
    assert store.mark_task_running(task["id"], "default")
    store.complete_task(
        task["id"],
        answer="",
        account_id="default",
        links=[
            {
                "link": link,
                "platform": domain,
                "title": "报告",
                "platform_type": "企业官网/品牌站",
            }
        ],
    )
    return job_id, task["id"]


def test_import_create_and_export(tmp_path: Path) -> None:
    with make_client(tmp_path) as client:
        job_id, _ = _create_job_with_result(client)
        job_name = client.get(f"/admin/api/research/jobs/{job_id}").json()["name"]

        exported = client.get("/admin/api/research/results/export.xlsx")
        assert exported.status_code == 200
        workbook = load_workbook(BytesIO(exported.content), read_only=True)
        rows = list(workbook.active.values)
        workbook.close()
        assert rows[0] == (
            "任务",
            "日期",
            "提问关键词",
            "AI平台",
            "资料名称",
            "检索资料链接",
            "检索资料平台",
            "平台类型",
        )
        assert rows[1][4] == "报告"
        assert rows[1][:7] == (
            job_name,
            rows[1][1],
            "新能源汽车",
            "doubao",
            "报告",
            TEST_LINK,
            TEST_DOMAIN,
        )
        assert rows[1][7] is not None
        assert isinstance(rows[1][7], str)

        results = client.get("/admin/api/research/results")
        assert results.status_code == 200
        assert results.json()["total"] == 1
        assert results.json()["keywords"] == ["新能源汽车"]
        assert results.json()["jobs"]
        assert results.json()["accounts"] == ["default"]

        keywords_response = client.get("/admin/api/research/results/keywords")
        assert keywords_response.status_code == 200
        assert "新能源汽车" in keywords_response.json()["keywords"]

        jobs_response = client.get("/admin/api/research/results/jobs")
        assert jobs_response.status_code == 200
        assert any(j["id"] == job_id for j in jobs_response.json()["jobs"])


def test_job_rename_and_delete(tmp_path: Path) -> None:
    with make_client(tmp_path) as client:
        job_id, _ = _create_job_with_result(client)

        renamed = client.post(
            f"/admin/api/research/jobs/{job_id}/rename",
            json={"name": "Renamed Job"},
        )
        assert renamed.status_code == 200

        # rename_job 返回的是同一事务内旧连接的快照，重新 GET 确认已持久化
        refreshed = client.get(f"/admin/api/research/jobs/{job_id}")
        assert refreshed.status_code == 200
        assert refreshed.json()["name"] == "Renamed Job"

        deleted = client.delete(f"/admin/api/research/jobs/{job_id}")
        assert deleted.status_code == 200
        assert deleted.json()["deleted"] is True

        missing = client.get(f"/admin/api/research/jobs/{job_id}")
        assert missing.status_code == 404


def test_create_job_with_multiple_platforms(tmp_path: Path) -> None:
    with make_client(tmp_path) as client:
        created = client.post(
            "/admin/api/research/jobs",
            json={
                "name": "多平台任务",
                "keywords": ["关键词 A"],
                "ai_platforms": ["doubao", "deepseek"],
            },
        )
        assert created.status_code == 200
        payload = created.json()
        assert payload["count"] == 2
        assert {job["ai_platform"] for job in payload["jobs"]} == {"doubao", "deepseek"}


def test_create_job_with_single_platform_keeps_old_shape(tmp_path: Path) -> None:
    with make_client(tmp_path) as client:
        created = client.post(
            "/admin/api/research/jobs",
            json={
                "name": "单平台任务",
                "keywords": ["关键词 A"],
                "ai_platforms": ["deepseek"],
            },
        )
        assert created.status_code == 200
        payload = created.json()
        assert payload["ai_platform"] == "deepseek"
        assert payload["id"]


def test_create_job_legacy_ai_platform_field(tmp_path: Path) -> None:
    with make_client(tmp_path) as client:
        created = client.post(
            "/admin/api/research/jobs",
            json={
                "name": "旧客户端任务",
                "keywords": ["关键词 A"],
                "ai_platform": "deepseek",
            },
        )
        assert created.status_code == 200
        assert created.json()["ai_platform"] == "deepseek"

        defaulted = client.post(
            "/admin/api/research/jobs",
            json={"name": "默认平台任务", "keywords": ["关键词 A"]},
        )
        assert defaulted.status_code == 200
        assert defaulted.json()["ai_platform"] == "doubao"


def test_template_with_multiple_platforms(tmp_path: Path) -> None:
    with make_client(tmp_path) as client:
        created = client.post(
            "/admin/api/research/templates",
            json={
                "name": "多平台模板",
                "keywords": ["关键词 A"],
                "ai_platforms": ["doubao", "deepseek"],
            },
        )
        assert created.status_code == 200
        template = created.json()
        assert template["ai_platforms"] == ["doubao", "deepseek"]

        fetched = client.get(f"/admin/api/research/templates/{template['id']}")
        assert fetched.json()["ai_platforms"] == ["doubao", "deepseek"]

        updated = client.post(
            f"/admin/api/research/templates/{template['id']}",
            json={
                "name": "多平台模板",
                "keywords": ["关键词 A"],
                "ai_platforms": ["deepseek"],
            },
        )
        assert updated.status_code == 200
        assert updated.json()["ai_platforms"] == ["deepseek"]


def test_sync_platform_info(tmp_path: Path) -> None:
    unique_domain = f"test-{uuid4().hex}.local"
    unique_link = f"https://{unique_domain}/report"
    try:
        with (
            patch.object(_server_module, "add_entries", _add_platform_entries_without_persist),
            make_client(tmp_path) as client,
        ):
            job_id, task_id = _create_job_with_result(
                client, domain=unique_domain, link=unique_link
            )
            store = client.app.state.research_store
            # 模拟旧数据：platform_type 为空，等待同步回填
            with store._connect() as conn:
                conn.execute("UPDATE research_results SET platform_type = ''")
                conn.commit()

            platform_data = platform_workbook(domain=unique_domain)
            imported = client.post(
                "/admin/api/research/platforms/import",
                files={
                    "file": (
                        "platforms.xlsx",
                        platform_data,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
            )
            assert imported.status_code == 200
            assert imported.json()["added"] >= 1

            synced = client.post(
                "/admin/api/research/results/sync-platform-info",
                json={"batch_size": 1000},
            )
            assert synced.status_code == 200
            assert synced.json()["updated"] >= 1

            platforms_response = client.get("/admin/api/research/platforms")
            assert platforms_response.status_code == 200
            assert any(e["domain"] == unique_domain for e in platforms_response.json()["entries"])
    finally:
        _remove_platform_entry(unique_domain)


def test_results_keywords_filter(tmp_path: Path) -> None:
    with make_client(tmp_path) as client:
        _create_job_with_result(client)

        filtered = client.get(
            "/admin/api/research/results",
            params={"keywords": ["不存在的关键词"]},
        )
        assert filtered.status_code == 200
        assert filtered.json()["total"] == 0

        matched = client.get(
            "/admin/api/research/results",
            params={"keywords": ["新能源汽车"]},
        )
        assert matched.status_code == 200
        assert matched.json()["total"] == 1


def test_source_comparison(tmp_path: Path) -> None:
    with make_client(tmp_path) as client:
        job_id, _ = _create_job_with_result(client)

        response = client.post(
            "/admin/api/research/results/source-comparison",
            json={
                "job_ids_a": [job_id],
                "job_ids_b": ["nonexistent-job"],
            },
        )
        assert response.status_code == 200
        data = response.json()
        assert "summary" in data
        assert "rows" in data


def test_long_tail_analysis_and_export(tmp_path: Path) -> None:
    with make_client(tmp_path) as client:
        _create_job_with_result(client)

        response = client.post(
            "/admin/api/research/results/long-tail-analysis",
            json={
                "freq_threshold": 1,
                "breadth_threshold": 1,
                "density_threshold": 0.1,
                "noise_density_threshold": 100,
            },
        )
        assert response.status_code == 200
        data = response.json()
        assert "summary" in data
        assert "platforms" in data

        exported = client.post(
            "/admin/api/research/results/long-tail/export.xlsx",
            json={
                "freq_threshold": 1,
                "breadth_threshold": 1,
                "density_threshold": 0.1,
                "noise_density_threshold": 100,
            },
        )
        assert exported.status_code == 200
        assert exported.headers["content-type"] == (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )


def test_schedule_run_now(tmp_path: Path) -> None:
    with make_client(tmp_path) as client:
        # 避免在 CI 上触发真正的浏览器调度；本测试只验证 endpoint 创建 job 并返回正确名称
        client.app.state.research_scheduler.wake = lambda: None

        template = client.post(
            "/admin/api/research/templates",
            json={
                "name": "测试模板",
                "keywords": ["新能源"],
                "prompt_template": "请调研 {keyword}",
            },
        )
        assert template.status_code == 200
        template_id = template.json()["id"]

        schedule = client.post(
            "/admin/api/research/schedules",
            json={
                "name": "测试计划",
                "template_id": template_id,
                "schedule_type": "once",
                "schedule_value": "2099-01-01T00:00:00",
            },
        )
        assert schedule.status_code == 200
        schedule_id = schedule.json()["id"]

        run = client.post(f"/admin/api/research/schedules/{schedule_id}/run")
        assert run.status_code == 200
        assert run.json()["name"].startswith("测试计划")
