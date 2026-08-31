from __future__ import annotations

import asyncio
import ipaddress
import json
import platform
import sys
import time
import uuid
import webbrowser
from collections import deque
from collections.abc import AsyncIterator, Callable
from contextlib import asynccontextmanager
from io import BytesIO
from pathlib import Path
from typing import Annotated, Any

from fastapi import FastAPI, File, Header, HTTPException, Query, Request, Response, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import load_workbook

from . import __version__
from .account_manager import BrowserAccountPool, normalize_account_id
from .browser_client import BrowserUnavailableError, LoginRequiredError
from .config import RuntimeConfig, SettingsStore
from .models import (
    AccountBackgroundActionRequest,
    AccountBatchCategoryRequest,
    AccountCacheClearRequest,
    AccountCategoryRequest,
    AccountProvisionRequest,
    AccountRenameRequest,
    AccountTabHiddenRequest,
    AdminSettingsUpdateRequest,
    ChatCompletionRequest,
    ImageGenerationRequest,
    ManualCookieImportRequest,
    ResearchJobCreateRequest,
    ResearchJobRenameRequest,
    ResearchJobTemplateCreateRequest,
    ResearchJobTemplateUpdateRequest,
    ResearchResultsLongTailRequest,
    ResearchResultsSourceComparisonRequest,
    ResearchResultsSyncRequest,
    ResearchScheduleCreateRequest,
    ResearchScheduleToggleRequest,
    ResearchScheduleUpdateRequest,
    VideoGenerationRequest,
)
from .platform_editor import PLATFORM_CATEGORIES, add_entries, all_entries
from .platforms import list_platforms
from .research_export import build_long_tail_workbook, build_results_workbook
from .research_import import normalize_keywords, parse_keyword_file
from .research_scheduler import ResearchScheduler
from .research_store import ResearchStore

CHAT_MODELS = ["doubao", "doubao-pro", "doubao-think", "doubao-expert"]
IMAGE_MODELS = ["doubao-image"]
VIDEO_MODELS = ["doubao-video"]
AUDIO_MODELS = ["doubao-music"]


def _is_loopback_host(value: str) -> bool:
    if value.casefold() == "localhost":
        return True
    try:
        return ipaddress.ip_address(value).is_loopback
    except ValueError:
        return False


def _error(message: str, status: int, code: int | str | None = None) -> JSONResponse:
    return JSONResponse(
        status_code=status,
        content={
            "error": {
                "message": message,
                "type": "api_error",
                "code": code if code is not None else status,
            }
        },
    )


def _last_user_message(request: ChatCompletionRequest) -> str:
    for message in reversed(request.messages):
        if message.role != "user":
            continue
        if isinstance(message.content, str):
            return message.content
        if isinstance(message.content, list):
            values: list[str] = []
            for item in message.content:
                if isinstance(item, str):
                    values.append(item)
                elif isinstance(item, dict) and isinstance(item.get("text"), str):
                    values.append(item["text"])
            return "\n".join(values)
    return ""


def _openai_completion(
    *,
    model: str,
    text: str,
    completion_id: str,
) -> dict[str, Any]:
    created = int(time.time())
    return {
        "id": completion_id,
        "object": "chat.completion",
        "created": created,
        "model": model,
        "choices": [
            {
                "index": 0,
                "message": {"role": "assistant", "content": text},
                "finish_reason": "stop",
            }
        ],
        "usage": {
            "prompt_tokens": 0,
            "completion_tokens": 0,
            "total_tokens": 0,
        },
    }


async def _completion_stream(
    *,
    model: str,
    text: str,
    completion_id: str,
) -> AsyncIterator[str]:
    created = int(time.time())
    first = {
        "id": completion_id,
        "object": "chat.completion.chunk",
        "created": created,
        "model": model,
        "choices": [{"index": 0, "delta": {"role": "assistant"}, "finish_reason": None}],
    }
    yield f"data: {json.dumps(first, ensure_ascii=False)}\n\n"
    for start in range(0, len(text), 24):
        chunk = {
            "id": completion_id,
            "object": "chat.completion.chunk",
            "created": created,
            "model": model,
            "choices": [
                {
                    "index": 0,
                    "delta": {"content": text[start : start + 24]},
                    "finish_reason": None,
                }
            ],
        }
        yield f"data: {json.dumps(chunk, ensure_ascii=False)}\n\n"
        await asyncio.sleep(0)
    final = {
        "id": completion_id,
        "object": "chat.completion.chunk",
        "created": created,
        "model": model,
        "choices": [{"index": 0, "delta": {}, "finish_reason": "stop"}],
    }
    yield f"data: {json.dumps(final, ensure_ascii=False)}\n\n"
    yield "data: [DONE]\n\n"


def create_app(
    *,
    store: SettingsStore | None = None,
    runtime: RuntimeConfig | None = None,
    client_factory: Callable[[Path, str, RuntimeConfig, str], Any] | None = None,
) -> FastAPI:
    settings_store = store or SettingsStore()
    runtime_config = runtime or RuntimeConfig.from_env()
    account_pool = BrowserAccountPool(
        settings_store,
        runtime_config,
        client_factory=client_factory,
    )
    research_store = ResearchStore(settings_store.data_root / "research.sqlite3")
    research_scheduler = ResearchScheduler(research_store, account_pool)
    started_at = time.monotonic()
    request_logs: deque[dict[str, Any]] = deque(maxlen=1000)

    @asynccontextmanager
    async def lifespan(_: FastAPI) -> AsyncIterator[None]:
        await research_scheduler.start()
        if runtime_config.open_admin_browser:
            asyncio.get_running_loop().call_later(
                1.0,
                lambda: webbrowser.open(
                    f"http://{runtime_config.host}:{runtime_config.port}/admin"
                ),
            )
        try:
            yield
        finally:
            await research_scheduler.stop()
            await account_pool.stop_all()

    app = FastAPI(
        title="Doubao Account Manager",
        version=__version__,
        lifespan=lifespan,
    )
    app.state.store = settings_store
    app.state.runtime = runtime_config
    app.state.account_pool = account_pool
    app.state.research_store = research_store
    app.state.research_scheduler = research_scheduler
    app.state.request_shutdown = None
    app.state.desktop_activate = None

    static_dir = Path(__file__).with_name("static")
    app.mount("/assets", StaticFiles(directory=static_dir), name="assets")

    @app.middleware("http")
    async def request_log_middleware(request: Request, call_next: Any) -> Response:
        if request.url.path.startswith(("/admin/api/", "/auth/")):
            client_host = request.client.host if request.client else ""
            if not _is_loopback_host(client_host):
                return _error("Administrative endpoints are local-only", 403)
        begin = time.perf_counter()
        response: Response
        try:
            response = await call_next(request)
        except Exception:
            request_logs.append(
                {
                    "ts": time.time(),
                    "method": request.method,
                    "path": request.url.path,
                    "account_id": request.query_params.get("account_id", ""),
                    "model": "",
                    "params": request.url.query,
                    "params_truncated": False,
                    "status": 500,
                    "ms": int((time.perf_counter() - begin) * 1000),
                }
            )
            raise
        request_logs.append(
            {
                "ts": time.time(),
                "method": request.method,
                "path": request.url.path,
                "account_id": request.query_params.get("account_id", ""),
                "model": "",
                "params": request.url.query[:1000],
                "params_truncated": len(request.url.query) > 1000,
                "status": response.status_code,
                "ms": int((time.perf_counter() - begin) * 1000),
            }
        )
        return response

    async def check_api_key(authorization: str | None) -> None:
        if not runtime_config.api_key:
            return
        if authorization != f"Bearer {runtime_config.api_key}":
            raise HTTPException(status_code=401, detail="Invalid or missing API key")

    @app.get("/health")
    async def health() -> dict[str, Any]:
        return {
            "status": "ok",
            "version": __version__,
            "activation_required": False,
            "open_source": True,
        }

    @app.get("/v1/models")
    async def list_models(authorization: str | None = Header(default=None)) -> dict[str, Any]:
        await check_api_key(authorization)
        now = int(time.time())
        return {
            "object": "list",
            "data": [
                {"id": model, "object": "model", "created": now, "owned_by": "doubao"}
                for model in CHAT_MODELS + IMAGE_MODELS + VIDEO_MODELS + AUDIO_MODELS
            ],
        }

    @app.post("/v1/chat/completions")
    async def chat_completions(
        body: ChatCompletionRequest,
        authorization: str | None = Header(default=None),
    ) -> Response:
        await check_api_key(authorization)
        try:
            account = await account_pool.get_or_start(body.account_id)
            result = await account.client.chat([message.model_dump() for message in body.messages])
        except LoginRequiredError as exc:
            return _error(str(exc), 401)
        except BrowserUnavailableError as exc:
            return _error(str(exc), 503)
        except (ValueError, RuntimeError, TimeoutError) as exc:
            return _error(str(exc), 502)

        completion_id = f"chatcmpl-{uuid.uuid4().hex}"
        text = result["text"]
        if body.stream:
            return StreamingResponse(
                _completion_stream(
                    model=body.model,
                    text=text,
                    completion_id=completion_id,
                ),
                media_type="text/event-stream",
                headers={"Cache-Control": "no-cache"},
            )
        return JSONResponse(
            _openai_completion(
                model=body.model,
                text=text,
                completion_id=completion_id,
            )
        )

    @app.post("/v1/images/generations")
    @app.post("/v1/images/generations/no-watermark")
    async def image_generations(
        body: ImageGenerationRequest,
        authorization: str | None = Header(default=None),
    ) -> Response:
        await check_api_key(authorization)
        return _error(
            "Image generation is not yet restored in the open-source rebuild. "
            "Account management and text chat are available.",
            501,
            "not_restored",
        )

    @app.get("/v1/images/generations/no-watermark/{task_id}")
    async def get_raw_image_generation(
        task_id: str,
        wait_timeout: float = 15,
        authorization: str | None = Header(default=None),
    ) -> Response:
        await check_api_key(authorization)
        return _error(f"Unknown image task: {task_id}", 404)

    @app.post("/v1/video/generations")
    async def video_generations(
        body: VideoGenerationRequest,
        authorization: str | None = Header(default=None),
    ) -> Response:
        await check_api_key(authorization)
        return _error(
            "Video generation is not yet restored in the open-source rebuild.",
            501,
            "not_restored",
        )

    @app.get("/v1/video/generations/{task_id}")
    async def get_video_generation(
        task_id: str,
        wait_timeout: float = 15,
        authorization: str | None = Header(default=None),
    ) -> Response:
        await check_api_key(authorization)
        return _error(f"Unknown video task: {task_id}", 404)

    @app.post("/v1/audio/generations")
    async def audio_generations(
        authorization: str | None = Header(default=None),
    ) -> Response:
        await check_api_key(authorization)
        return _error(
            "Audio generation is not yet restored in the open-source rebuild.",
            501,
            "not_restored",
        )

    @app.post("/v1/files")
    @app.post("/v1/images/upload")
    @app.post("/v1/chat/completions/with-file")
    async def file_routes(
        authorization: str | None = Header(default=None),
    ) -> Response:
        await check_api_key(authorization)
        return _error(
            "File upload and file-assisted chat are not yet restored.",
            501,
            "not_restored",
        )

    @app.get("/v1/files/download")
    async def file_download(
        uri: str,
        expire: int = 3600,
        authorization: str | None = Header(default=None),
    ) -> Response:
        await check_api_key(authorization)
        return _error(f"File download is not restored for URI: {uri}", 501, "not_restored")

    @app.get("/admin", response_class=HTMLResponse)
    async def admin_dashboard() -> FileResponse:
        return FileResponse(static_dir / "index.html", media_type="text/html")

    @app.get("/")
    async def home() -> Response:
        return JSONResponse(
            {
                "name": "Doubao Account Manager",
                "version": __version__,
                "admin": "/admin",
                "docs": "/docs",
                "activation_required": False,
            }
        )

    @app.get("/admin/api/system")
    async def admin_system() -> dict[str, Any]:
        settings = settings_store.settings
        return {
            "python_version": sys.version,
            "platform": platform.platform(),
            "uptime_seconds": int(time.monotonic() - started_at),
            "host": runtime_config.host,
            "port": runtime_config.port,
            "open_source": True,
            "activation_required": False,
            "browser": {
                "backend": "playwright",
                "channel": runtime_config.browser_channel,
                "executable_path": runtime_config.browser_executable_path,
            },
            "models": {
                "chat": CHAT_MODELS,
                "image": IMAGE_MODELS,
                "video": VIDEO_MODELS,
                "audio": AUDIO_MODELS,
            },
            "ai_platforms": [
                {"key": p.key, "name": p.name, "models": p.chat_models} for p in list_platforms()
            ],
            "accounts": {
                "default_account_id": account_pool.default_account_id,
                "accounts_root": str(account_pool.accounts_root),
                "known": account_pool.discover_account_ids(),
                "started": list(account_pool._managed),
            },
            "settings": settings.public_dict(),
        }

    @app.get("/admin/api/status")
    @app.get("/auth/status")
    async def admin_status(account_id: str | None = None) -> dict[str, Any]:
        snapshot = await account_pool.snapshot(
            normalize_account_id(account_id, account_pool.default_account_id)
        )
        return {
            "account_id": snapshot["account_id"],
            "default_account_id": account_pool.default_account_id,
            "started": snapshot["started"],
            "logged_in": snapshot["logged_in"],
            "browser": "ready" if snapshot["started"] else "not_started",
            "has_ms_token": snapshot["has_ms_token"],
            "chat_ready": snapshot["chat_ready"],
            "needs_captcha": snapshot["needs_captcha"],
        }

    @app.get("/admin/api/settings")
    async def get_admin_settings() -> dict[str, Any]:
        return settings_store.settings.public_dict()

    @app.post("/admin/api/settings")
    async def update_admin_settings(body: AdminSettingsUpdateRequest) -> dict[str, Any]:
        values = body.model_dump(exclude_none=True)
        settings = settings_store.update(values)
        return {**settings.public_dict(), "message": "Settings updated"}

    @app.get("/admin/api/accounts")
    async def admin_accounts() -> dict[str, Any]:
        return {
            "default_account_id": account_pool.default_account_id,
            "accounts": await account_pool.snapshots(),
        }

    @app.post("/admin/api/research/import")
    async def admin_import_research_keywords(
        file: Annotated[UploadFile, File()],
    ) -> dict[str, Any]:
        filename = file.filename or ""
        data = await file.read()
        if len(data) > 20 * 1024 * 1024:
            raise HTTPException(status_code=413, detail="关键词文件不能超过 20 MB")
        try:
            keywords = parse_keyword_file(filename, data)
        except (ValueError, OSError) as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        if not keywords:
            raise HTTPException(status_code=400, detail="文件中没有识别到关键词")
        return {"filename": filename, "count": len(keywords), "keywords": keywords}

    @app.post("/admin/api/research/jobs")
    async def admin_create_research_job(
        body: ResearchJobCreateRequest,
    ) -> dict[str, Any]:
        keywords = normalize_keywords(body.keywords)
        if not keywords:
            raise HTTPException(status_code=400, detail="请至少填写一个关键词")
        if len(keywords) * body.repeat_count > 10000:
            raise HTTPException(
                status_code=400,
                detail="关键词数 × 采集次数不能超过 10000 个采集单元",
            )
        if "{keyword}" not in body.prompt_template:
            raise HTTPException(
                status_code=400,
                detail="提问模板必须包含 {keyword} 占位符",
            )
        account_ids: list[str] = []
        for value in body.account_ids:
            try:
                account_id = normalize_account_id(value, account_pool.default_account_id)
            except ValueError as exc:
                raise HTTPException(status_code=400, detail=str(exc)) from exc
            if account_id not in account_ids:
                account_ids.append(account_id)
        try:
            jobs = [
                research_store.create_job(
                    name=body.name,
                    keywords=keywords,
                    account_ids=account_ids,
                    prompt_template=body.prompt_template,
                    scheduled_at=body.scheduled_at,
                    interval_seconds=body.interval_seconds,
                    account_cooldown_seconds=body.account_cooldown_seconds,
                    max_attempts=body.max_attempts,
                    ai_platform=ai_platform,
                    repeat_count=body.repeat_count,
                    round_interval_seconds=body.round_interval_seconds,
                )
                for ai_platform in body.ai_platforms
            ]
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        if len(jobs) == 1:
            return jobs[0]
        return {"jobs": jobs, "count": len(jobs)}

    @app.get("/admin/api/research/jobs")
    async def admin_list_research_jobs(limit: int = 100) -> dict[str, Any]:
        return {
            "jobs": research_store.list_jobs(min(max(limit, 1), 500)),
            "scheduler": research_scheduler.snapshot(),
        }

    @app.get("/admin/api/research/jobs/{job_id}")
    async def admin_get_research_job(job_id: str) -> dict[str, Any]:
        try:
            return research_store.get_job(job_id)
        except KeyError as exc:
            raise HTTPException(status_code=404, detail="采集任务不存在") from exc

    @app.post("/admin/api/research/jobs/{job_id}/pause")
    async def admin_pause_research_job(job_id: str) -> dict[str, Any]:
        try:
            return research_store.set_job_status(job_id, "paused")
        except KeyError as exc:
            raise HTTPException(status_code=404, detail="采集任务不存在") from exc

    @app.post("/admin/api/research/jobs/{job_id}/resume")
    async def admin_resume_research_job(job_id: str) -> dict[str, Any]:
        try:
            return research_store.set_job_status(job_id, "running")
        except KeyError as exc:
            raise HTTPException(status_code=404, detail="采集任务不存在") from exc

    @app.post("/admin/api/research/jobs/{job_id}/cancel")
    async def admin_cancel_research_job(job_id: str) -> dict[str, Any]:
        try:
            result = research_store.set_job_status(job_id, "cancelled")
            research_scheduler.cancel_job(job_id)
            return result
        except KeyError as exc:
            raise HTTPException(status_code=404, detail="采集任务不存在") from exc

    @app.delete("/admin/api/research/jobs/{job_id}")
    async def admin_delete_research_job(job_id: str) -> dict[str, Any]:
        try:
            research_store.delete_job(job_id)
            return {"deleted": True, "job_id": job_id}
        except KeyError as exc:
            raise HTTPException(status_code=404, detail="采集任务不存在") from exc

    @app.post("/admin/api/research/jobs/{job_id}/rename")
    async def admin_rename_research_job(
        job_id: str,
        body: ResearchJobRenameRequest,
    ) -> dict[str, Any]:
        try:
            return research_store.rename_job(job_id, body.name)
        except KeyError as exc:
            raise HTTPException(status_code=404, detail="采集任务不存在") from exc
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc

    # ------------------------------------------------------------------
    # Research job templates
    # ------------------------------------------------------------------

    @app.post("/admin/api/research/templates")
    async def admin_create_research_job_template(
        body: ResearchJobTemplateCreateRequest,
    ) -> dict[str, Any]:
        keywords = normalize_keywords(body.keywords)
        if not keywords:
            raise HTTPException(status_code=400, detail="请至少填写一个关键词")
        if "{keyword}" not in body.prompt_template:
            raise HTTPException(status_code=400, detail="提问模板必须包含 {keyword} 占位符")
        try:
            return research_store.create_job_template(
                name=body.name,
                keywords=keywords,
                prompt_template=body.prompt_template,
                interval_seconds=body.interval_seconds,
                account_cooldown_seconds=body.account_cooldown_seconds,
                max_attempts=body.max_attempts,
                ai_platforms=body.ai_platforms,
                repeat_count=body.repeat_count,
                round_interval_seconds=body.round_interval_seconds,
            )
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc

    @app.get("/admin/api/research/templates")
    async def admin_list_research_job_templates() -> dict[str, Any]:
        return {"templates": research_store.list_job_templates()}

    @app.get("/admin/api/research/templates/{template_id}")
    async def admin_get_research_job_template(template_id: str) -> dict[str, Any]:
        try:
            return research_store.get_job_template(template_id)
        except KeyError as exc:
            raise HTTPException(status_code=404, detail="任务模板不存在") from exc

    @app.post("/admin/api/research/templates/{template_id}")
    async def admin_update_research_job_template(
        template_id: str,
        body: ResearchJobTemplateUpdateRequest,
    ) -> dict[str, Any]:
        keywords = normalize_keywords(body.keywords)
        if not keywords:
            raise HTTPException(status_code=400, detail="请至少填写一个关键词")
        if "{keyword}" not in body.prompt_template:
            raise HTTPException(status_code=400, detail="提问模板必须包含 {keyword} 占位符")
        try:
            return research_store.update_job_template(
                template_id,
                name=body.name,
                keywords=keywords,
                prompt_template=body.prompt_template,
                interval_seconds=body.interval_seconds,
                account_cooldown_seconds=body.account_cooldown_seconds,
                max_attempts=body.max_attempts,
                ai_platforms=body.ai_platforms,
                repeat_count=body.repeat_count,
                round_interval_seconds=body.round_interval_seconds,
            )
        except KeyError as exc:
            raise HTTPException(status_code=404, detail="任务模板不存在") from exc
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc

    @app.delete("/admin/api/research/templates/{template_id}")
    async def admin_delete_research_job_template(template_id: str) -> dict[str, Any]:
        try:
            research_store.delete_job_template(template_id)
            return {"deleted": True}
        except KeyError as exc:
            raise HTTPException(status_code=404, detail="任务模板不存在") from exc

    # ------------------------------------------------------------------
    # Research schedules
    # ------------------------------------------------------------------

    @app.post("/admin/api/research/schedules")
    async def admin_create_research_schedule(
        body: ResearchScheduleCreateRequest,
    ) -> dict[str, Any]:
        if body.schedule_type not in ("interval", "once", "daily"):
            raise HTTPException(status_code=400, detail="触发类型必须是 interval、once 或 daily")
        try:
            return research_store.create_schedule(
                name=body.name,
                template_id=body.template_id,
                schedule_type=body.schedule_type,
                schedule_value=body.schedule_value,
            )
        except KeyError as exc:
            raise HTTPException(status_code=404, detail="任务模板不存在") from exc
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc

    @app.get("/admin/api/research/schedules")
    async def admin_list_research_schedules() -> dict[str, Any]:
        return {
            "schedules": research_store.list_schedules(),
            "templates": research_store.list_job_templates(),
        }

    @app.get("/admin/api/research/schedules/{schedule_id}")
    async def admin_get_research_schedule(schedule_id: str) -> dict[str, Any]:
        try:
            return research_store.get_schedule(schedule_id)
        except KeyError as exc:
            raise HTTPException(status_code=404, detail="定时计划不存在") from exc

    @app.post("/admin/api/research/schedules/{schedule_id}")
    async def admin_update_research_schedule(
        schedule_id: str,
        body: ResearchScheduleUpdateRequest,
    ) -> dict[str, Any]:
        if body.schedule_type not in ("interval", "once", "daily"):
            raise HTTPException(status_code=400, detail="触发类型必须是 interval、once 或 daily")
        try:
            return research_store.update_schedule(
                schedule_id,
                name=body.name,
                template_id=body.template_id,
                schedule_type=body.schedule_type,
                schedule_value=body.schedule_value,
            )
        except KeyError as exc:
            raise HTTPException(status_code=404, detail="定时计划或任务模板不存在") from exc
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc

    @app.post("/admin/api/research/schedules/{schedule_id}/toggle")
    async def admin_toggle_research_schedule(
        schedule_id: str,
        body: ResearchScheduleToggleRequest,
    ) -> dict[str, Any]:
        try:
            return research_store.toggle_schedule(schedule_id, body.enabled)
        except KeyError as exc:
            raise HTTPException(status_code=404, detail="定时计划不存在") from exc

    @app.delete("/admin/api/research/schedules/{schedule_id}")
    async def admin_delete_research_schedule(schedule_id: str) -> dict[str, Any]:
        try:
            research_store.delete_schedule(schedule_id)
            return {"deleted": True}
        except KeyError as exc:
            raise HTTPException(status_code=404, detail="定时计划不存在") from exc

    @app.post("/admin/api/research/schedules/{schedule_id}/run")
    async def admin_run_research_schedule_now(schedule_id: str) -> dict[str, Any]:
        try:
            jobs = research_store.create_jobs_from_schedule(schedule_id)
            research_scheduler.wake()
        except KeyError as exc:
            raise HTTPException(status_code=404, detail="定时计划不存在") from exc
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        if len(jobs) == 1:
            return jobs[0]
        return {"jobs": jobs, "count": len(jobs)}

    @app.get("/admin/api/research/results")
    async def admin_research_results(
        job_id: str = "",
        keyword: str = "",
        keywords: Annotated[list[str] | None, Query()] = None,
        platform: str = "",
        account_id: str = "",
        date_from: str = "",
        date_to: str = "",
        limit: int = 500,
        offset: int = 0,
    ) -> dict[str, Any]:
        filter_keywords = keywords or []
        filters = {
            "job_id": job_id,
            "keyword": filter_keywords if filter_keywords else keyword,
            "platform": platform,
            "account_id": account_id,
            "date_from": date_from,
            "date_to": date_to,
        }
        dashboard = research_store.result_dashboard(**filters)
        return {
            "items": research_store.list_results(
                **filters,
                limit=limit,
                offset=offset,
            ),
            "total": dashboard["summary"]["total"],
            "platforms": research_store.platforms(),
            "accounts": research_store.result_accounts(),
            "keywords": research_store.result_keywords(),
            "jobs": research_store.result_jobs(),
            "dashboard": dashboard,
        }

    @app.get("/admin/api/research/results/export.xlsx")
    async def admin_export_research_results(
        job_id: str = "",
        keyword: str = "",
        keywords: Annotated[list[str] | None, Query()] = None,
        platform: str = "",
        account_id: str = "",
        date_from: str = "",
        date_to: str = "",
    ) -> Response:
        filter_keywords = keywords or []
        rows = research_store.list_results(
            job_id=job_id,
            keyword=filter_keywords if filter_keywords else keyword,
            platform=platform,
            account_id=account_id,
            date_from=date_from,
            date_to=date_to,
            limit=100000,
        )
        content = build_results_workbook(rows)
        return Response(
            content,
            media_type=("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
            headers={
                "Content-Disposition": ('attachment; filename="doubao-thinking-references.xlsx"')
            },
        )

    @app.post("/admin/api/research/results/sync-platform-info")
    async def admin_sync_platform_info(
        body: ResearchResultsSyncRequest | None = None,
    ) -> dict[str, Any]:
        updated = research_store.sync_platform_info(
            batch_size=(body.batch_size if body else 10000),
        )
        return {"updated": updated}

    @app.get("/admin/api/research/results/keywords")
    async def admin_result_keywords() -> dict[str, Any]:
        return {"keywords": research_store.result_keywords()}

    @app.get("/admin/api/research/results/jobs")
    async def admin_result_jobs() -> dict[str, Any]:
        return {"jobs": research_store.result_jobs()}

    @app.post("/admin/api/research/results/source-comparison")
    async def admin_source_comparison(
        body: ResearchResultsSourceComparisonRequest,
    ) -> dict[str, Any]:
        return research_store.source_comparison(
            job_ids_a=body.job_ids_a,
            job_ids_b=body.job_ids_b,
            keyword=body.keywords,
            platform=body.platform,
            account_id=body.account_id,
        )

    @app.post("/admin/api/research/results/long-tail-analysis")
    async def admin_long_tail_analysis(
        body: ResearchResultsLongTailRequest,
    ) -> dict[str, Any]:
        return research_store.long_tail_analysis(
            job_id=body.job_id,
            keyword=body.keywords,
            platform=body.platform,
            platforms=body.platforms,
            account_id=body.account_id,
            account_ids=body.account_ids,
            date_from=body.date_from,
            date_to=body.date_to,
            split_mode=body.split_mode,
            breadth_threshold=body.breadth_threshold,
            freq_threshold=body.freq_threshold,
            density_threshold=body.density_threshold,
            noise_density_threshold=body.noise_density_threshold,
        )

    @app.post("/admin/api/research/results/long-tail/export.xlsx")
    async def admin_export_long_tail(
        body: ResearchResultsLongTailRequest,
    ) -> Response:
        analysis = research_store.long_tail_analysis(
            job_id=body.job_id,
            keyword=body.keywords,
            platform=body.platform,
            platforms=body.platforms,
            account_id=body.account_id,
            account_ids=body.account_ids,
            date_from=body.date_from,
            date_to=body.date_to,
            split_mode=body.split_mode,
            breadth_threshold=body.breadth_threshold,
            freq_threshold=body.freq_threshold,
            density_threshold=body.density_threshold,
            noise_density_threshold=body.noise_density_threshold,
        )
        content = build_long_tail_workbook(
            analysis.get("target_long_tail", []),
            analysis.get("params", {}),
            analysis.get("summary", {}),
        )
        return Response(
            content,
            media_type=("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
            headers={"Content-Disposition": ('attachment; filename="long-tail-platforms.xlsx"')},
        )

    @app.get("/admin/api/research/platforms")
    async def admin_research_platforms() -> dict[str, Any]:
        return {
            "entries": all_entries(),
            "categories": list(PLATFORM_CATEGORIES),
        }

    @app.post("/admin/api/research/platforms/import")
    async def admin_import_research_platforms(
        file: Annotated[UploadFile, File()],
    ) -> dict[str, Any]:
        data = await file.read()
        if len(data) > 20 * 1024 * 1024:
            raise HTTPException(status_code=413, detail="平台规则文件不能超过 20 MB")
        try:
            workbook = load_workbook(filename=BytesIO(data), read_only=True)
        except (ValueError, OSError) as exc:
            raise HTTPException(status_code=400, detail=f"无法解析 Excel: {exc}") from exc

        sheet = workbook.active
        if sheet is None:
            raise HTTPException(status_code=400, detail="Excel 中没有工作表")

        headers = [
            str(cell.value or "").strip() for cell in next(sheet.iter_rows(min_row=1, max_row=1))
        ]
        url_col = next(
            (i for i, h in enumerate(headers) if h in {"url", "域名", "URL"}),
            None,
        )
        name_col = next(
            (i for i, h in enumerate(headers) if h in {"平台名", "平台名称", "name"}),
            None,
        )
        category_col = next(
            (i for i, h in enumerate(headers) if h in {"平台类型", "类型", "category"}),
            None,
        )
        if url_col is None or name_col is None or category_col is None:
            workbook.close()
            raise HTTPException(
                status_code=400,
                detail="Excel 表头必须包含 URL/域名、平台名/平台名称、平台类型/类型 列",
            )

        rows: list[dict[str, str]] = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            url = str(row[url_col] or "").strip()
            name = str(row[name_col] or "").strip()
            category = str(row[category_col] or "").strip()
            if url and name and category:
                rows.append({"url": url, "平台名": name, "平台类型": category})
        workbook.close()

        return add_entries(rows)

    @app.post("/admin/api/accounts")
    async def admin_start_account(body: AccountProvisionRequest) -> dict[str, Any]:
        account_id = normalize_account_id(body.account_id, account_pool.default_account_id)
        account_pool.ensure_account_environment(account_id)
        if body.ai_platform:
            account_pool.set_account_platform(account_id, body.ai_platform)
        if body.start_browser:
            if body.background:
                asyncio.create_task(account_pool.start_account(account_id))
            else:
                try:
                    await account_pool.start_account(account_id)
                except BrowserUnavailableError as exc:
                    raise HTTPException(status_code=503, detail=str(exc)) from exc
        return await account_pool.snapshot(account_id)

    @app.post("/admin/api/accounts/{account_id}/stop")
    async def admin_stop_account(
        account_id: str,
        body: AccountBackgroundActionRequest | None = None,
    ) -> dict[str, Any]:
        if body and body.background:
            asyncio.create_task(account_pool.stop_account(account_id))
        else:
            await account_pool.stop_account(account_id)
        return {**(await account_pool.snapshot(account_id)), "message": "Account stopped"}

    @app.post("/admin/api/accounts/{account_id}/popup")
    async def admin_popup_account(account_id: str) -> dict[str, Any]:
        managed = account_pool.get_if_started(account_id)
        if not managed:
            managed = await account_pool.start_account(account_id)
        await managed.client.bring_to_front()
        return {"account_id": account_id, "message": "Account window focused"}

    @app.post("/admin/api/accounts/{account_id}/rename")
    async def admin_rename_account(
        account_id: str,
        body: AccountRenameRequest,
    ) -> dict[str, Any]:
        new_id = await account_pool.rename_account(account_id, body.new_account_id)
        return {
            **(await account_pool.snapshot(new_id)),
            "previous_account_id": account_id,
            "message": "Account renamed",
        }

    @app.post("/admin/api/accounts/{account_id}/category")
    async def admin_update_account_category(
        account_id: str,
        body: AccountCategoryRequest,
    ) -> dict[str, Any]:
        account_pool.set_category(account_id, body.category)
        return {
            **(await account_pool.snapshot(account_id)),
            "message": "Account category updated",
        }

    @app.post("/admin/api/accounts/{account_id}/tab-hidden")
    async def admin_update_account_tab_hidden(
        account_id: str,
        body: AccountTabHiddenRequest,
    ) -> dict[str, Any]:
        account_pool.set_tab_hidden(account_id, body.hidden)
        return {
            **(await account_pool.snapshot(account_id)),
            "message": "Account tab visibility updated",
        }

    @app.post("/admin/api/accounts/category")
    async def admin_update_account_categories(
        body: AccountBatchCategoryRequest,
    ) -> dict[str, Any]:
        for account_id in body.account_ids:
            account_pool.set_category(account_id, body.category)
        return {
            "updated": len(body.account_ids),
            "category": (body.category or "").strip(),
            "message": "Account categories updated",
        }

    @app.post("/admin/api/accounts/cache/clear")
    async def admin_clear_account_caches(
        body: AccountCacheClearRequest | None = None,
    ) -> dict[str, Any]:
        account_ids = (
            body.account_ids if body and body.account_ids else account_pool.discover_account_ids()
        )
        results: list[dict[str, Any]] = []
        for account_id in account_ids:
            try:
                result = await account_pool.clear_account_cache(account_id)
                results.append(
                    {
                        **result,
                        "account_id": account_id,
                        "user_data_dir": str(account_pool.get_user_data_path(account_id)),
                        "success": True,
                    }
                )
            except Exception as exc:
                results.append(
                    {
                        "account_id": account_id,
                        "success": False,
                        "error": str(exc),
                    }
                )
        return {
            "cleared": sum(1 for item in results if item["success"]),
            "failed": sum(1 for item in results if not item["success"]),
            "deleted_path_count": sum(item.get("deleted_path_count", 0) for item in results),
            "results": results,
            "message": "Account caches cleared",
        }

    @app.post("/admin/api/accounts/{account_id}/environment/reset")
    async def admin_reset_account_environment(account_id: str) -> dict[str, Any]:
        await account_pool.reset_account_environment(account_id)
        return {
            **(await account_pool.snapshot(account_id)),
            "message": "Account environment reset",
        }

    @app.delete("/admin/api/accounts/{account_id}")
    async def admin_delete_account(account_id: str) -> dict[str, Any]:
        before = await account_pool.snapshot(account_id)
        await account_pool.delete_account(account_id)
        return {
            **before,
            "environment_exists": False,
            "deleted": True,
            "message": "Account deleted",
        }

    @app.get("/admin/api/cookies")
    async def admin_cookies(account_id: str | None = None) -> dict[str, Any]:
        managed = account_pool.get_if_started(account_id)
        cookies = await managed.client.cookies() if managed else []
        redacted = [
            {
                "name": item.get("name", ""),
                "domain": item.get("domain", ""),
                "path": item.get("path", ""),
                "length": len(item.get("value", "")),
                "httpOnly": item.get("httpOnly", False),
                "secure": item.get("secure", False),
            }
            for item in cookies
        ]
        return {"cookies": redacted, "count": len(redacted)}

    @app.post("/admin/api/browser-cookies/import")
    async def admin_import_browser_cookies(
        body: ManualCookieImportRequest,
    ) -> dict[str, Any]:
        account_id = normalize_account_id(body.account_id, account_pool.default_account_id)
        if body.reset_environment and account_id != account_pool.default_account_id:
            await account_pool.reset_account_environment(account_id)
        managed = await account_pool.start_account(account_id)
        count = await managed.client.import_cookies(body.cookie_text)
        return {
            "account_id": account_id,
            "imported": count,
            "message": "Cookies imported",
        }

    @app.post("/auth/login")
    async def auth_login(account_id: str | None = None) -> dict[str, Any]:
        managed = await account_pool.start_account(account_id)
        await managed.client.bring_to_front()
        return await admin_status(managed.account_id)

    @app.post("/v1/session/qr-login")
    async def session_qr_login_start(
        account_id: str | None = None,
        authorization: str | None = Header(default=None),
    ) -> dict[str, Any]:
        await check_api_key(authorization)
        managed = await account_pool.start_account(account_id)
        await managed.client.bring_to_front()
        state = await admin_status(managed.account_id)
        return {
            **state,
            "status": "waiting_scan" if not state["logged_in"] else "success",
            "message": "Complete login in the account browser window.",
        }

    @app.get("/v1/session/qr-login")
    async def session_qr_login_poll(
        account_id: str | None = None,
        authorization: str | None = Header(default=None),
    ) -> dict[str, Any]:
        await check_api_key(authorization)
        state = await admin_status(account_id)
        return {
            **state,
            "status": "success"
            if state["logged_in"]
            else ("waiting_scan" if state["started"] else "idle"),
        }

    @app.post("/auth/reset_captcha")
    async def auth_reset_captcha(account_id: str | None = None) -> dict[str, Any]:
        managed = account_pool.get_if_started(account_id)
        if managed:
            await managed.client.reset_captcha()
        return await admin_status(account_id)

    @app.get("/auth/screenshot")
    async def auth_screenshot(account_id: str | None = None) -> Response:
        managed = account_pool.get_if_started(account_id)
        if not managed:
            raise HTTPException(status_code=404, detail="Account browser is not started")
        return Response(content=await managed.client.screenshot(), media_type="image/png")

    @app.get("/admin/api/screenshots")
    async def admin_screenshots() -> dict[str, Any]:
        screenshots = [
            {
                "account_id": account_id,
                "url": f"/auth/screenshot?account_id={account_id}",
            }
            for account_id in account_pool._managed
        ]
        return {"screenshots": screenshots}

    @app.post("/admin/api/probe")
    async def admin_probe(account_id: str | None = None) -> dict[str, Any]:
        managed = account_pool.get_if_started(account_id)
        if not managed:
            return await admin_status(account_id)
        return await managed.client.inspect_session_state()

    @app.get("/admin/api/browser-sources")
    async def admin_browser_sources() -> dict[str, Any]:
        return {
            "sources": [],
            "browser": {
                "backend": "playwright",
                "channel": runtime_config.browser_channel,
                "executable_path": runtime_config.browser_executable_path,
            },
            "message": (
                "Profile copying is intentionally omitted; use the login window or cookie import."
            ),
        }

    @app.get("/admin/api/logs")
    async def admin_logs() -> list[dict[str, Any]]:
        return list(request_logs)

    @app.post("/admin/api/shutdown")
    async def admin_shutdown() -> dict[str, Any]:
        callback = app.state.request_shutdown
        if callback is None:
            return {
                "status": "unavailable",
                "message": "Shutdown is only available when started by the bundled CLI.",
            }
        asyncio.get_running_loop().call_later(0.2, callback)
        return {"status": "shutting_down", "message": "Shutdown request accepted."}

    @app.post("/admin/api/desktop/activate")
    async def admin_activate_desktop() -> dict[str, Any]:
        callback = app.state.desktop_activate
        if callback is None:
            return {"status": "unavailable"}
        callback()
        return {"status": "activated"}

    return app


def run_server(runtime: RuntimeConfig | None = None) -> None:
    import uvicorn

    runtime_config = runtime or RuntimeConfig.from_env()
    if not _is_loopback_host(runtime_config.host) and not runtime_config.api_key:
        raise RuntimeError("Refusing to listen on a non-local address without DOUBAO_API_KEY.")
    app = create_app(runtime=runtime_config)
    windowed = sys.stdout is None or sys.stderr is None
    config = uvicorn.Config(
        app,
        host=runtime_config.host,
        port=runtime_config.port,
        log_level="info",
        log_config=None if windowed else uvicorn.config.LOGGING_CONFIG,
        access_log=not windowed,
    )
    server = uvicorn.Server(config)
    app.state.request_shutdown = lambda: setattr(server, "should_exit", True)
    server.run()
