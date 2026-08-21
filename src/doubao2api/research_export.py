from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .research_platforms import category_for_url, platform_category


def _platform_type_for_result(item: dict[str, Any]) -> str:
    platform_type = item.get("platform_type", "")
    if platform_type:
        return platform_type
    # 如果记录插入时平台类型为空，按当前最新的平台规则再查一次
    return category_for_url(item.get("link", "")) or platform_category(item.get("platform", ""))


def build_results_workbook(rows: list[dict[str, Any]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "思考过程资料链接"
    headers = [
        "任务",
        "日期",
        "提问关键词",
        "AI平台",
        "资料名称",
        "检索资料链接",
        "检索资料平台",
        "平台类型",
    ]
    sheet.append(headers)
    for item in rows:
        sheet.append(
            [
                item.get("job_name", ""),
                item["collected_date"],
                item["keyword"],
                item.get("ai_platform", "doubao"),
                item.get("title", ""),
                item["link"],
                item["platform"],
                _platform_type_for_result(item),
            ]
        )

    header_fill = PatternFill("solid", fgColor="183153")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 14
    sheet.column_dimensions["C"].width = 32
    sheet.column_dimensions["D"].width = 16
    sheet.column_dimensions["E"].width = 52
    sheet.column_dimensions["F"].width = 80
    sheet.column_dimensions["G"].width = 22
    sheet.column_dimensions["H"].width = 20
    for row in sheet.iter_rows(min_row=2):
        row[5].hyperlink = row[5].value
        row[5].style = "Hyperlink"
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def build_long_tail_workbook(
    targets: list[dict[str, Any]],
    params: dict[str, Any],
    summary: dict[str, Any],
) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "长尾信源推荐"
    sheet.append(["平台", "频次", "广度", "密度", "平台类型", "代表链接", "关键词示例", "象限分类"])
    for item in targets:
        sheet.append(
            [
                item.get("platform", ""),
                item.get("freq", 0),
                item.get("breadth", 0),
                round(item.get("density", 0.0), 2),
                item.get("type", ""),
                item.get("representative_link", ""),
                ", ".join(item.get("keywords_sample", [])[:5]),
                item.get("quadrant", ""),
            ]
        )

    header_fill = PatternFill("solid", fgColor="183153")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 10
    sheet.column_dimensions["C"].width = 10
    sheet.column_dimensions["D"].width = 10
    sheet.column_dimensions["E"].width = 20
    sheet.column_dimensions["F"].width = 60
    sheet.column_dimensions["G"].width = 40
    sheet.column_dimensions["H"].width = 24
    for row in sheet.iter_rows(min_row=2):
        if row[5].value:
            row[5].hyperlink = row[5].value
            row[5].style = "Hyperlink"
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()
