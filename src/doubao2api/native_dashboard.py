from __future__ import annotations

import asyncio
import concurrent.futures
import math
import threading
import time
from collections.abc import Callable, Coroutine
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import matplotlib

matplotlib.use("QtAgg")
matplotlib.rcParams["font.sans-serif"] = [
    "Microsoft YaHei",
    "SimHei",
    "SimSun",
    "sans-serif",
]
matplotlib.rcParams["axes.unicode_minus"] = False

from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas  # noqa: E402
from matplotlib.figure import Figure  # noqa: E402
from openpyxl import Workbook, load_workbook  # noqa: E402
from PySide6.QtCore import QDate, QEvent, QRectF, Qt, QTimer, QUrl, Signal  # noqa: E402
from PySide6.QtGui import QColor, QDesktopServices, QKeySequence, QPainter, QShortcut  # noqa: E402
from PySide6.QtWidgets import (  # noqa: E402
    QAbstractItemView,
    QApplication,
    QCalendarWidget,
    QCheckBox,
    QComboBox,
    QDateEdit,
    QDialog,
    QDoubleSpinBox,
    QFileDialog,
    QFrame,
    QGridLayout,
    QGroupBox,
    QHBoxLayout,
    QHeaderView,
    QInputDialog,
    QLabel,
    QLineEdit,
    QListWidget,
    QListWidgetItem,
    QMenu,
    QMessageBox,
    QProgressBar,
    QPushButton,
    QScrollArea,
    QSizePolicy,
    QSpinBox,
    QTableView,
    QTableWidget,
    QTableWidgetItem,
    QTabWidget,
    QTextEdit,
    QToolButton,
    QVBoxLayout,
    QWidget,
    QWidgetAction,
)

from . import __version__  # noqa: E402
from .account_manager import BrowserAccountPool, normalize_account_id  # noqa: E402
from .config import RuntimeConfig, SettingsStore  # noqa: E402
from .embedded_browser_client import EmbeddedBrowserClient  # noqa: E402
from .platform_editor import add_entries, all_entries  # noqa: E402
from .research_export import build_results_workbook  # noqa: E402
from .research_import import normalize_keywords, parse_keyword_file  # noqa: E402
from .research_scheduler import ResearchScheduler  # noqa: E402
from .research_store import ResearchStore  # noqa: E402
from .update_checker import (  # noqa: E402
    DownloadResult,
    UpdateChecker,
    UpdateInfo,
    _detect_variant,
    _normalize_version,
    load_cached_update_info,
    save_cached_update_info,
)

DEFAULT_PROMPT = "{keyword}"
STATUS_TEXT = {
    "running": "进行中",
    "paused": "已暂停",
    "completed": "已完成",
    "failed": "失败",
    "cancelled": "已取消",
    "pending": "等待中",
}


@dataclass(slots=True)
class PendingOperation:
    future: concurrent.futures.Future[Any]
    callback: Callable[[Any], None] | None
    error_callback: Callable[[BaseException], None] | None
    deadline: float
    silent: bool
    label: str


class MultiSelectFilter(QWidget):
    def __init__(self, default_text: str, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.default_text = default_text
        self._updating = False
        layout = QHBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)

        self.button = QToolButton()
        self.button.setObjectName("filterButton")
        self.button.setText(default_text)
        self.button.setPopupMode(QToolButton.ToolButtonPopupMode.InstantPopup)
        self.button.setMinimumWidth(190)
        self.button.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        layout.addWidget(self.button)

        self.menu = QMenu(self)
        self.menu.setObjectName("filterMenu")
        self.menu.setMinimumWidth(310)
        self.button.setMenu(self.menu)

        search_action = QWidgetAction(self.menu)
        self.search = QLineEdit()
        self.search.setPlaceholderText("搜索关键词")
        self.search.setClearButtonEnabled(True)
        search_action.setDefaultWidget(self.search)
        self.menu.addAction(search_action)

        controls_action = QWidgetAction(self.menu)
        controls = QWidget()
        controls_layout = QHBoxLayout(controls)
        controls_layout.setContentsMargins(0, 4, 0, 4)
        select_all = QPushButton("全选")
        clear = QPushButton("清空")
        select_all.setObjectName("linkButton")
        clear.setObjectName("linkButton")
        controls_layout.addWidget(select_all)
        controls_layout.addWidget(clear)
        controls_layout.addStretch()
        controls_action.setDefaultWidget(controls)
        self.menu.addAction(controls_action)

        list_action = QWidgetAction(self.menu)
        self.list_widget = QListWidget()
        self.list_widget.setMinimumHeight(190)
        self.list_widget.setMaximumHeight(300)
        self.list_widget.setSelectionMode(QAbstractItemView.SelectionMode.NoSelection)
        list_action.setDefaultWidget(self.list_widget)
        self.menu.addAction(list_action)

        self.search.textChanged.connect(self._filter_items)
        self.list_widget.itemChanged.connect(self._sync_label)
        select_all.clicked.connect(self.select_all)
        clear.clicked.connect(self.clear_selection)

    def set_options(self, values: list[str] | list[tuple[str, str]]) -> None:
        normalized: list[tuple[str, str]] = []
        for value in values:
            if isinstance(value, tuple):
                normalized.append(value)
            else:
                normalized.append((value, value))
        current = [
            (
                self.list_widget.item(index).text(),
                str(self.list_widget.item(index).data(Qt.ItemDataRole.UserRole)),
            )
            for index in range(self.list_widget.count())
        ]
        if current == normalized:
            return
        selected = set(self.selected_values())
        self._updating = True
        self.list_widget.clear()
        for display, value in normalized:
            item = QListWidgetItem(display)
            item.setData(Qt.ItemDataRole.UserRole, value)
            item.setFlags(item.flags() | Qt.ItemFlag.ItemIsUserCheckable)
            item.setCheckState(
                Qt.CheckState.Checked if value in selected else Qt.CheckState.Unchecked
            )
            self.list_widget.addItem(item)
        self._updating = False
        self._filter_items(self.search.text())
        self._sync_label()

    def selected_values(self) -> list[str]:
        return [
            str(item.data(Qt.ItemDataRole.UserRole))
            for index in range(self.list_widget.count())
            if (item := self.list_widget.item(index)).checkState() == Qt.CheckState.Checked
        ]

    def select_all(self) -> None:
        self._set_visible_items(Qt.CheckState.Checked)

    def clear_selection(self) -> None:
        self._updating = True
        for index in range(self.list_widget.count()):
            self.list_widget.item(index).setCheckState(Qt.CheckState.Unchecked)
        self._updating = False
        self._sync_label()

    def _set_visible_items(self, state: Qt.CheckState) -> None:
        self._updating = True
        for index in range(self.list_widget.count()):
            item = self.list_widget.item(index)
            if not item.isHidden():
                item.setCheckState(state)
        self._updating = False
        self._sync_label()

    def _filter_items(self, text: str) -> None:
        needle = text.strip().casefold()
        for index in range(self.list_widget.count()):
            item = self.list_widget.item(index)
            item.setHidden(bool(needle) and needle not in item.text().casefold())

    def _sync_label(self, *_: Any) -> None:
        if self._updating:
            return
        selected = self.selected_values()
        if not selected:
            text = self.default_text
        elif len(selected) == 1:
            text = selected[0]
        else:
            text = f"已选 {len(selected)} 个关键词"
        self.button.setText(text)
        self.button.setToolTip("\n".join(selected))


class DateRangeCalendar(QCalendarWidget):
    range_selected = Signal(QDate, QDate)

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        today = QDate.currentDate()
        self.range_start = today
        self.range_end = today
        self._press_date: QDate | None = None
        self._pending_anchor: QDate | None = None
        self._dragging = False
        self.setGridVisible(False)
        self.setMinimumSize(350, 285)
        self._calendar_view = self.findChild(QTableView, "qt_calendar_calendarview")
        if self._calendar_view is not None:
            self._calendar_view.viewport().installEventFilter(self)

    def set_range(self, start: QDate, end: QDate) -> None:
        self.range_start, self.range_end = sorted((start, end))
        self._pending_anchor = None
        self.setSelectedDate(self.range_end)
        self.setCurrentPage(self.range_end.year(), self.range_end.month())
        self.updateCells()

    def eventFilter(self, watched: Any, event: Any) -> bool:
        if (
            self._calendar_view is None
            or watched is not self._calendar_view.viewport()
            or event.type()
            not in {
                QEvent.Type.MouseButtonPress,
                QEvent.Type.MouseMove,
                QEvent.Type.MouseButtonRelease,
            }
        ):
            return super().eventFilter(watched, event)
        date = self._date_at(event.position().toPoint())
        if date is None:
            return super().eventFilter(watched, event)
        if event.type() == QEvent.Type.MouseButtonPress:
            if event.button() != Qt.MouseButton.LeftButton:
                return super().eventFilter(watched, event)
            self._press_date = date
            self._dragging = False
            return True
        if event.type() == QEvent.Type.MouseMove:
            if self._press_date is None or not event.buttons() & Qt.MouseButton.LeftButton:
                return super().eventFilter(watched, event)
            if date != self._press_date:
                self._dragging = True
                self.range_start, self.range_end = sorted((self._press_date, date))
                self.updateCells()
            return True
        if event.type() == QEvent.Type.MouseButtonRelease:
            if self._press_date is None or event.button() != Qt.MouseButton.LeftButton:
                return super().eventFilter(watched, event)
            pressed = self._press_date
            self._press_date = None
            if self._dragging:
                self.range_start, self.range_end = sorted((pressed, date))
                self._pending_anchor = None
                self.range_selected.emit(self.range_start, self.range_end)
            elif self._pending_anchor is None:
                self._pending_anchor = date
                self.range_start = date
                self.range_end = date
            else:
                self.range_start, self.range_end = sorted((self._pending_anchor, date))
                self._pending_anchor = None
                self.range_selected.emit(self.range_start, self.range_end)
            self.updateCells()
            return True
        return super().eventFilter(watched, event)

    def _date_at(self, point: Any) -> QDate | None:
        if self._calendar_view is None:
            return None
        index = self._calendar_view.indexAt(point)
        if not index.isValid() or index.row() < 1 or index.column() < 1:
            return None
        first = QDate(self.yearShown(), self.monthShown(), 1)
        offset = (first.dayOfWeek() - self.firstDayOfWeek().value) % 7
        return first.addDays((index.row() - 1) * 7 + index.column() - 1 - offset)

    def paintCell(self, painter: QPainter, rect: Any, date: QDate) -> None:
        super().paintCell(painter, rect, date)
        if not (self.range_start <= date <= self.range_end):
            return
        painter.save()
        endpoint = date in {self.range_start, self.range_end}
        painter.setPen(Qt.PenStyle.NoPen)
        painter.setBrush(QColor("#5b5ce2" if endpoint else "#e7e8ff"))
        painter.drawRoundedRect(rect.adjusted(2, 2, -2, -2), 5, 5)
        painter.setPen(QColor("#ffffff" if endpoint else "#34365f"))
        painter.drawText(rect, Qt.AlignmentFlag.AlignCenter, str(date.day()))
        painter.restore()


class DateRangePicker(QPushButton):
    def __init__(
        self,
        start: QDate,
        end: QDate,
        parent: QWidget | None = None,
    ) -> None:
        super().__init__(parent)
        self.start_date, self.end_date = sorted((start, end))
        self.setObjectName("rangePicker")
        self.setMinimumWidth(285)

        self.popup = QDialog(self, Qt.WindowType.Popup)
        self.popup.setObjectName("rangePopup")
        self.popup.setMinimumWidth(370)
        content_layout = QVBoxLayout(self.popup)
        content_layout.setContentsMargins(8, 6, 8, 8)
        hint = QLabel("按住鼠标拖过日期；也可依次点击开始日和结束日")
        hint.setObjectName("muted")
        content_layout.addWidget(hint)
        self.calendar = DateRangeCalendar()
        self.calendar.set_range(self.start_date, self.end_date)
        self.calendar.range_selected.connect(self._apply_range)
        content_layout.addWidget(self.calendar)
        self.clicked.connect(self._show_popup)
        self._sync_text()

    def _show_popup(self) -> None:
        self.popup.adjustSize()
        self.popup.move(self.mapToGlobal(self.rect().bottomLeft()))
        self.popup.show()

    def _apply_range(self, start: QDate, end: QDate) -> None:
        self.start_date, self.end_date = start, end
        self._sync_text()
        self.popup.hide()

    def _sync_text(self) -> None:
        self.setText(
            f"{self.start_date.toString('yyyy-MM-dd')}  —  {self.end_date.toString('yyyy-MM-dd')}"
        )

    def date_from(self) -> str:
        return self.start_date.toString("yyyy-MM-dd")

    def date_to(self) -> str:
        return self.end_date.toString("yyyy-MM-dd")


DISTRIBUTION_COLORS = (
    "#5b5ce2",
    "#22a06b",
    "#e69a2d",
    "#3b82c4",
    "#b15ac7",
    "#de5267",
)


class DonutChart(QWidget):
    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.rows: list[dict[str, Any]] = []
        self.total = 0
        self.setFixedWidth(220)
        self.setMinimumHeight(200)
        self.setSizePolicy(QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Expanding)

    def set_rows(self, rows: list[dict[str, Any]], total: int) -> None:
        self.rows = rows
        self.total = total
        self.update()

    def paintEvent(self, _: Any) -> None:
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        if not self.rows or self.total <= 0:
            painter.setPen(QColor("#738096"))
            painter.drawText(self.rect(), Qt.AlignmentFlag.AlignCenter, "当前筛选条件下暂无信源")
            return

        donut_size = min(170, self.height() - 40)
        donut = QRectF(
            (self.width() - donut_size) / 2,
            (self.height() - donut_size) / 2,
            donut_size,
            donut_size,
        )
        start_angle = 90 * 16
        for index, row in enumerate(self.rows):
            share = int(row["count"]) / self.total
            span = round(share * 360 * 16)
            painter.setPen(Qt.PenStyle.NoPen)
            painter.setBrush(QColor(DISTRIBUTION_COLORS[index % len(DISTRIBUTION_COLORS)]))
            painter.drawPie(donut, start_angle, -span)
            if share >= 0.06:
                middle_angle = math.radians(start_angle / 16 - share * 180)
                radius = donut_size * 0.38
                center = donut.center()
                x = center.x() + math.cos(middle_angle) * radius
                y = center.y() - math.sin(middle_angle) * radius
                label_rect = QRectF(x - 24, y - 10, 48, 20)
                percent_font = painter.font()
                percent_font.setPointSize(8)
                percent_font.setBold(True)
                painter.setFont(percent_font)
                painter.setPen(QColor("#ffffff"))
                painter.drawText(
                    label_rect,
                    Qt.AlignmentFlag.AlignCenter,
                    f"{share * 100:.0f}%",
                )
            start_angle -= span
        inset = donut_size * 0.28
        inner = donut.adjusted(inset, inset, -inset, -inset)
        painter.setBrush(QColor("#ffffff"))
        painter.drawEllipse(inner)
        font = painter.font()
        font.setBold(True)
        font.setPointSize(13)
        painter.setFont(font)
        painter.setPen(QColor("#252b47"))
        painter.drawText(inner, Qt.AlignmentFlag.AlignCenter, "100%")


class PlatformDistributionRow(QWidget):
    def __init__(
        self,
        row: dict[str, Any],
        total: int,
        color: str,
        parent: QWidget | None = None,
    ) -> None:
        super().__init__(parent)
        count = int(row["count"])
        share = count / total * 100 if total else 0.0
        layout = QHBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(8)

        swatch = QLabel()
        swatch.setFixedSize(14, 14)
        swatch.setStyleSheet(f"background: {color}; border-radius: 4px;")
        layout.addWidget(swatch)

        name = QLabel(str(row.get("platform") or "未知平台"))
        name.setObjectName("sourceName")
        name.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)
        layout.addWidget(name, 1)

        count_label = QLabel(f"{count}")
        count_label.setObjectName("muted")
        count_label.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        count_label.setFixedWidth(46)
        layout.addWidget(count_label)

        share_label = QLabel(f"{share:.1f}%")
        share_label.setObjectName("muted")
        share_label.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        share_label.setFixedWidth(52)
        layout.addWidget(share_label)

        bar = QProgressBar()
        bar.setTextVisible(False)
        bar.setFixedHeight(6)
        bar.setFixedWidth(80)
        bar.setRange(0, 100)
        bar.setValue(int(share))
        bar.setStyleSheet(
            f"""
            QProgressBar {{
                background: #edf0f6;
                border-radius: 3px;
            }}
            QProgressBar::chunk {{
                background: {color};
                border-radius: 3px;
            }}
            """
        )
        layout.addWidget(bar)
        self.setFixedHeight(28)


class PlatformDistributionList(QScrollArea):
    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.setWidgetResizable(True)
        self.setFrameShape(QFrame.Shape.NoFrame)
        self.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        container = QWidget()
        self._layout = QVBoxLayout(container)
        self._layout.setContentsMargins(4, 6, 4, 6)
        self._layout.setSpacing(4)
        self._layout.addStretch(1)
        self.setWidget(container)

    def set_rows(self, rows: list[dict[str, Any]], total: int) -> None:
        container = self.widget()
        assert container is not None
        while self._layout.count() > 1:
            item = self._layout.takeAt(0)
            widget = item.widget()
            if widget is not None:
                widget.deleteLater()
        for index, row in enumerate(rows):
            color = DISTRIBUTION_COLORS[index % len(DISTRIBUTION_COLORS)]
            self._layout.insertWidget(
                self._layout.count() - 1,
                PlatformDistributionRow(row, total, color, container),
            )


class SourceDistributionChart(QWidget):
    TOP_N = 20

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.setMinimumHeight(260)
        self.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        layout = QHBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(16)
        self.donut = DonutChart(self)
        self.list = PlatformDistributionList(self)
        layout.addWidget(self.donut)
        layout.addWidget(self.list, 1)

    def set_rows(self, rows: list[dict[str, Any]], total: int, detail: bool = False) -> None:
        if detail:
            visible = list(rows)
        else:
            visible = [dict(row) for row in rows[: self.TOP_N]]
            other_count = sum(int(row["count"]) for row in rows[self.TOP_N :])
            if other_count:
                visible.append({"platform": "其他", "count": other_count})
        self.donut.set_rows(visible, total)
        self.list.set_rows(visible, total)


class LongTailChart(FigureCanvas):
    QUADRANT_COLORS = {
        "垂直长尾宝藏": "#27ae60",
        "虚假长尾(噪声)": "#e74c3c",
        "头部主流媒体": "#3498db",
        "特定品类垂直站": "#f39c12",
        "普通垂直信源": "#9b59b6",
        "一次性/僵尸信源": "#95a5a6",
    }

    def __init__(self, parent: QWidget | None = None) -> None:
        self._figure = Figure(figsize=(7, 5.5), dpi=100)
        super().__init__(self._figure)
        self.setParent(parent)
        self.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self.setMinimumHeight(380)
        self._axes = self._figure.add_subplot(111)
        self._data: dict[str, Any] | None = None
        self._plot_points: list[tuple[str, int, int, float, float, float, str]] = []
        self._hover_annotation = self._axes.annotate(
            "",
            xy=(0, 0),
            xytext=(10, 10),
            textcoords="offset points",
            bbox={
                "boxstyle": "round,pad=0.3",
                "fc": "white",
                "ec": "#738096",
                "alpha": 0.95,
            },
            fontsize=9,
            color="#252b47",
            visible=False,
        )
        self.mpl_connect("motion_notify_event", self._on_hover)
        self._draw_empty()

    def set_data(
        self, data: dict[str, Any], log_scale: bool = False, x_log_scale: bool = False
    ) -> None:
        self._data = data
        self._draw(log_scale, x_log_scale)

    def _draw_empty(self) -> None:
        self._axes.clear()
        self._axes.text(
            0.5,
            0.5,
            "点击“分析长尾信源”查看四象限图",
            horizontalalignment="center",
            verticalalignment="center",
            transform=self._axes.transAxes,
            fontsize=12,
            color="#738096",
        )
        self._axes.axis("off")
        self._hover_annotation = self._axes.annotate(
            "",
            xy=(0, 0),
            xytext=(10, 10),
            textcoords="offset points",
            bbox={
                "boxstyle": "round,pad=0.3",
                "fc": "white",
                "ec": "#738096",
                "alpha": 0.95,
            },
            fontsize=9,
            color="#252b47",
            visible=False,
        )
        self.draw()

    @staticmethod
    def _jitter(value: float, key: str, scale: float = 0.22) -> float:
        h = hash(key) & 0xFFFFFFFF
        return value + scale * (h / 0xFFFFFFFF * 2 - 1)

    def _point_size(self, density: float, quadrant: str) -> float:
        size = max(30.0, (max(density, 0.1) ** 0.5) * 80.0)
        if quadrant == "一次性/僵尸信源":
            size *= 0.6
        return size

    def _draw(self, log_scale: bool, x_log_scale: bool) -> None:
        self._axes.clear()
        self._plot_points = []
        self._hover_annotation = self._axes.annotate(
            "",
            xy=(0, 0),
            xytext=(10, 10),
            textcoords="offset points",
            bbox={
                "boxstyle": "round,pad=0.3",
                "fc": "white",
                "ec": "#738096",
                "alpha": 0.95,
            },
            fontsize=9,
            color="#252b47",
            visible=False,
        )
        if not self._data or not self._data.get("platforms"):
            self._draw_empty()
            return

        platforms = self._data["platforms"]
        params = self._data["params"]
        x_values: list[float] = []
        y_values: list[float] = []
        sizes: list[float] = []
        colors: list[str] = []
        for point in platforms:
            breadth = point["breadth"]
            freq = point["freq"]
            x_plot = self._jitter(float(breadth), point["platform"] + "x")
            y_plot = self._jitter(float(freq), point["platform"] + "y")
            x_plot = max(0.3, x_plot) if x_log_scale else max(0.0, x_plot)
            y_plot = max(0.3, y_plot) if log_scale else max(0.0, y_plot)
            self._plot_points.append(
                (
                    point["platform"],
                    breadth,
                    freq,
                    x_plot,
                    y_plot,
                    point["density"],
                    point["quadrant"],
                )
            )
            x_values.append(x_plot)
            y_values.append(y_plot)
            sizes.append(self._point_size(point["density"], point["quadrant"]))
            colors.append(self.QUADRANT_COLORS.get(point["quadrant"], "#333333"))

        self._axes.scatter(
            x_values,
            y_values,
            s=sizes,
            c=colors,
            alpha=0.75,
            edgecolors="white",
            linewidths=0.8,
        )

        if params["split_mode"] == "median":
            x_line = params.get("medians", {}).get("breadth", 0)
            y_line = params.get("medians", {}).get("freq", 0)
        else:
            x_line = params["breadth_threshold"] - 0.5
            y_line = params["freq_threshold"]
        if x_line > 0:
            self._axes.axvline(x=x_line, color="#e74c3c", linestyle="--", linewidth=1)
        if y_line > 0:
            self._axes.axhline(y=y_line, color="#e74c3c", linestyle="--", linewidth=1)

        if x_values and y_values:
            x_min = max(0.5, min(x_values) * 0.8) if x_log_scale else 0.0
            x_max = max(x_values) * 1.15
            y_min = 0.5 if log_scale else 1.0
            y_max = max(y_values) * 1.15
            self._axes.set_xlim(x_min, x_max)
            self._axes.set_ylim(y_min, y_max)

        self._axes.set_xlabel("关键词覆盖广度", fontsize=10)
        self._axes.set_ylabel("引用频次", fontsize=10)
        self._axes.set_title(
            "气泡大小 = 平均引用密度，悬停查看平台名称", fontsize=11, color="#252b47"
        )
        if x_log_scale:
            self._axes.set_xscale("log")
        if log_scale:
            self._axes.set_yscale("log")
        self._axes.grid(True, linestyle=":", alpha=0.5)
        self._axes.set_axisbelow(True)
        self._figure.tight_layout()
        self.draw()

    def _on_hover(self, event: Any) -> None:
        if not self._plot_points or event.inaxes != self._axes:
            self._hover_annotation.set_visible(False)
            self.draw_idle()
            return

        nearest: tuple[str, int, int, float, float, float, str] | None = None
        min_distance = float("inf")
        threshold = 20.0
        for platform, breadth, freq, x_plot, y_plot, density, quadrant in self._plot_points:
            pixel_x, pixel_y = self._axes.transData.transform((x_plot, y_plot))
            distance = ((pixel_x - event.x) ** 2 + (pixel_y - event.y) ** 2) ** 0.5
            if distance < min_distance:
                min_distance = distance
                nearest = (platform, breadth, freq, x_plot, y_plot, density, quadrant)

        if nearest is None or min_distance > threshold:
            self._hover_annotation.set_visible(False)
        else:
            platform, breadth, freq, x_plot, y_plot, density, quadrant = nearest
            self._hover_annotation.set_text(
                f"{platform}\n广度 {breadth} · 频次 {freq} · 密度 {density}\n{quadrant}"
            )
            self._hover_annotation.xy = (x_plot, y_plot)
            self._hover_annotation.set_visible(True)
        self.draw_idle()

    def wheelEvent(self, event: Any) -> None:
        event.ignore()


class DesktopBackend:
    def __init__(self, bridge: Any, runtime: RuntimeConfig) -> None:
        self.bridge = bridge
        self.runtime = runtime
        self.settings_store = SettingsStore()
        self.research_store = ResearchStore(self.settings_store.data_root / "research.sqlite3")

        def client_factory(
            user_data_dir: Path,
            account_id: str,
            _: RuntimeConfig,
        ) -> EmbeddedBrowserClient:
            return EmbeddedBrowserClient(bridge, user_data_dir, account_id)

        self.account_pool = BrowserAccountPool(
            self.settings_store,
            runtime,
            client_factory=client_factory,
            runtime_store=self.research_store,
        )
        self.scheduler = ResearchScheduler(self.research_store, self.account_pool)
        self.loop = asyncio.new_event_loop()
        self.ready = threading.Event()
        self.start_error: BaseException | None = None
        self.thread = threading.Thread(
            target=self._run_loop,
            name="doubao-task-engine",
            daemon=True,
        )
        self.thread.start()
        if not self.ready.wait(10):
            raise RuntimeError("采集引擎启动超时")
        if self.start_error is not None:
            raise RuntimeError(f"采集引擎启动失败：{self.start_error}") from self.start_error

    def _run_loop(self) -> None:
        asyncio.set_event_loop(self.loop)
        try:
            self.loop.run_until_complete(self.scheduler.start())
        except BaseException as exc:
            self.start_error = exc
            self.ready.set()
            return
        self.ready.set()
        self.loop.run_forever()

    def submit(self, coroutine: Coroutine[Any, Any, Any]) -> concurrent.futures.Future[Any]:
        return asyncio.run_coroutine_threadsafe(coroutine, self.loop)

    def call(self, function: Callable[[], Any]) -> concurrent.futures.Future[Any]:
        future: concurrent.futures.Future[Any] = concurrent.futures.Future()

        def invoke() -> None:
            try:
                future.set_result(function())
            except Exception as exc:
                future.set_exception(exc)

        self.loop.call_soon_threadsafe(invoke)
        return future

    def shutdown(self) -> concurrent.futures.Future[Any]:
        async def stop() -> None:
            await self.scheduler.stop()
            await self.account_pool.stop_all()

        future = self.submit(stop())
        future.add_done_callback(lambda _: self.loop.call_soon_threadsafe(self.loop.stop))
        return future


class NativeDashboard(QWidget):
    captcha_detected = Signal(str)
    download_progress = Signal(int, int)

    def __init__(self, backend: DesktopBackend, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.setObjectName("dashboard")
        self.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)
        self.backend = backend
        self.captcha_detected.connect(self._on_captcha_detected)
        self.download_progress.connect(self._on_download_progress)
        backend.scheduler.on_captcha_callback = self.captcha_detected.emit

        self.pending: list[PendingOperation] = []
        self.refreshing_accounts = False
        self.refreshing_jobs = False
        self.refreshing_history = False
        self.refreshing_results = False
        self.refreshing_comparison = False
        self.result_rows: list[dict[str, Any]] = []
        self.platform_distribution_rows: list[dict[str, Any]] = []
        self.platform_distribution_total = 0
        self.result_signature: tuple[Any, ...] | None = None
        self.result_filter_signature: tuple[Any, ...] | None = None
        self.comparison_signature: tuple[Any, ...] | None = None
        self.long_tail_data: dict[str, Any] | None = None
        self.analyzing_long_tail = False
        self.results_focus_mode = False
        self._window_was_maximized = False
        self._window_was_fullscreen = False
        self.editing_template_id: str | None = None
        self.refreshing_schedules = False
        self._current_update_info: UpdateInfo | None = None
        self._downloaded_paths: dict[str, Path] = {}
        self._downloading_variant: str | None = None
        self._update_download_dir = Path(self.backend.settings_store.data_root) / "updates"
        self._build_ui()
        self._apply_style()
        self.escape_shortcut = QShortcut(QKeySequence("Esc"), self)
        self.escape_shortcut.activated.connect(self._exit_results_focus)

        self.future_timer = QTimer(self)
        self.future_timer.timeout.connect(self._poll_futures)
        self.future_timer.start(100)
        self.refresh_timer = QTimer(self)
        self.refresh_timer.timeout.connect(self.refresh_all)
        self.refresh_timer.start(3000)
        self.result_timer = QTimer(self)
        self.result_timer.timeout.connect(self._refresh_visible_results)
        self.result_timer.start(3000)
        QTimer.singleShot(100, self.refresh_all)
        QTimer.singleShot(250, self.restore_account_sessions)
        QTimer.singleShot(500, self._load_cached_update_info)
        QTimer.singleShot(3000, self._auto_check_update_if_enabled)

    def restore_account_sessions(self) -> None:
        account_ids = self.backend.account_pool.discover_account_ids()
        marker = getattr(self.backend.bridge, "mark_background_open", None)
        if callable(marker):
            marker(account_ids)

        async def restore() -> list[str]:
            failed: list[str] = []
            for account_id in account_ids:
                try:
                    await self.backend.account_pool.start_account(account_id)
                except Exception:
                    failed.append(account_id)
            self.backend.scheduler.wake()
            return failed

        def applied(failed: list[str]) -> None:
            self.ready_accounts_value.setText(
                str(len(self.backend.account_pool.discover_account_ids()))
            )
            self.refresh_jobs()
            if failed:
                self.engine_badge.setToolTip("以下账号未能自动恢复：" + "、".join(failed))

        self._watch(
            self.backend.submit(restore()),
            applied,
            timeout_seconds=90,
            silent=True,
            label="自动恢复账号页面",
        )

    def _refresh_visible_results(self) -> None:
        if self.sections.currentWidget() is self.results_page:
            self.refresh_results()
        elif self.sections.currentWidget() is self.comparison_page:
            self.refresh_source_comparison()

    def _build_ui(self) -> None:
        root = QVBoxLayout(self)
        root.setContentsMargins(22, 14, 22, 16)
        root.setSpacing(10)
        self.header_bar = QWidget()
        title_row = QHBoxLayout(self.header_bar)
        title_row.setContentsMargins(0, 0, 0, 0)
        title_row.setSpacing(12)
        logo = QLabel("豆")
        logo.setObjectName("logo")
        logo.setAlignment(Qt.AlignmentFlag.AlignCenter)
        logo.setFixedSize(46, 46)
        title_row.addWidget(logo)
        title = QLabel("豆包关键词资料采集器")
        title.setObjectName("title")
        subtitle = QLabel("多账号并行调度 · 参考资料实时保存 · 全程本地运行")
        subtitle.setObjectName("muted")
        title_box = QVBoxLayout()
        title_box.setSpacing(2)
        title_box.addWidget(title)
        title_box.addWidget(subtitle)
        title_row.addLayout(title_box)
        title_row.addStretch()
        self.engine_badge = QLabel("● 采集引擎运行中")
        self.engine_badge.setObjectName("statusPill")
        title_row.addWidget(self.engine_badge)
        local_badge = QLabel("本地数据")
        local_badge.setObjectName("localPill")
        local_badge.setToolTip("账号登录状态、任务和采集结果均保存在本机")
        title_row.addWidget(local_badge)
        exit_button = QPushButton("退出软件")
        exit_button.setObjectName("dangerButton")
        exit_button.clicked.connect(QApplication.instance().quit)
        title_row.addWidget(exit_button)
        root.addWidget(self.header_bar)

        self.sections = QTabWidget()
        self.sections.setDocumentMode(True)
        self.tasks_page = self._build_tasks_page()
        self.accounts_page = self._build_accounts_page()
        self.history_page = self._build_history_page()
        self.results_page = self._build_results_page()
        self.long_tail_page = self._build_long_tail_page()
        self.comparison_page = self._build_comparison_page()
        self.schedules_page = self._build_schedules_page()
        self.platforms_page = self._build_platforms_page()
        self.update_page = self._build_update_page()
        self.sections.addTab(self.tasks_page, "新建采集")
        self.sections.addTab(self.accounts_page, "账号环境")
        self.sections.addTab(self.history_page, "历史任务")
        self.sections.addTab(self.results_page, "采集结果")
        self.sections.addTab(self.long_tail_page, "长尾信源")
        self.sections.addTab(self.comparison_page, "信源对比")
        self.sections.addTab(self.schedules_page, "定时任务")
        self.sections.addTab(self.platforms_page, "平台信息")
        self.sections.addTab(self.update_page, "检查更新")
        self.sections.currentChanged.connect(self._on_section_changed)
        root.addWidget(self.sections, 1)

    def _metric_card(self, title: str, accent: str) -> tuple[QFrame, QLabel]:
        card = QFrame()
        card.setObjectName("metricCard")
        card.setProperty("accent", accent)
        card.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        layout = QVBoxLayout(card)
        layout.setContentsMargins(14, 9, 14, 9)
        layout.setSpacing(2)
        value = QLabel("0")
        value.setObjectName("metricValue")
        label = QLabel(title)
        label.setObjectName("muted")
        layout.addWidget(value)
        layout.addWidget(label)
        return card, value

    def _build_tasks_page(self) -> QWidget:
        page = QWidget()
        outer = QVBoxLayout(page)
        outer.setContentsMargins(0, 0, 0, 0)
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        content = QWidget()
        content.setMinimumWidth(900)
        layout = QVBoxLayout(content)
        layout.setContentsMargins(4, 8, 4, 4)
        layout.setSpacing(12)
        scroll.setWidget(content)
        outer.addWidget(scroll)

        workflow = QLabel(
            "① 登录并打开账号    →    ② 输入或导入关键词    →    ③ 设置间隔并开始采集"
        )
        workflow.setObjectName("workflowBanner")
        workflow.setToolTip("采集过程中可切换到账号标签页查看豆包页面，不影响任务状态保存")
        layout.addWidget(workflow)

        metrics = QGridLayout()
        metrics.setHorizontalSpacing(12)
        metrics.setVerticalSpacing(8)
        ready_card, self.ready_accounts_value = self._metric_card("可用账号", "green")
        active_card, self.active_jobs_value = self._metric_card("进行中任务", "blue")
        worker_card, self.active_workers_value = self._metric_card("并行采集中", "purple")
        result_card, self.total_results_value = self._metric_card("已保存资料", "amber")
        for column, card in enumerate((ready_card, active_card, worker_card, result_card)):
            metrics.addWidget(card, 0, column)
        layout.addLayout(metrics)

        form_group = QGroupBox("新建采集任务")
        form_group.setMinimumHeight(405)
        form = QVBoxLayout(form_group)
        form.setSpacing(12)

        name_row = QHBoxLayout()
        name_row.addWidget(QLabel("任务名称"))
        self.job_name = QLineEdit()
        self.job_name.setPlaceholderText("例如：7 月品牌词调研")
        name_row.addWidget(self.job_name, 1)
        form.addLayout(name_row)

        keyword_header = QHBoxLayout()
        keyword_header.addWidget(QLabel("关键词（一行一个）"))
        self.keyword_count = QLabel("0 个")
        self.keyword_count.setObjectName("countPill")
        keyword_header.addWidget(self.keyword_count)
        keyword_header.addStretch()
        import_button = QPushButton("导入 Excel / CSV")
        import_button.setObjectName("secondaryButton")
        import_button.clicked.connect(self.import_keywords)
        keyword_header.addWidget(import_button)
        form.addLayout(keyword_header)

        self.keywords = QTextEdit()
        self.keywords.setPlaceholderText("直接输入关键词，每行一个；也可以导入 Excel / CSV")
        self.keywords.setMinimumHeight(140)
        self.keywords.setMaximumHeight(200)
        self.keywords.textChanged.connect(self._update_keyword_count)
        form.addWidget(self.keywords)

        timing_row = QHBoxLayout()
        self.interval_seconds = QSpinBox()
        self.interval_seconds.setRange(1, 86400)
        self.interval_seconds.setValue(10)
        self.interval_seconds.setSuffix(" 秒")
        self.max_attempts = QSpinBox()
        self.max_attempts.setRange(1, 3)
        self.max_attempts.setValue(2)
        timing_row.addWidget(QLabel("关键词启动间隔"))
        timing_row.addWidget(self.interval_seconds, 1)
        timing_row.addSpacing(12)
        timing_row.addWidget(QLabel("最多尝试"))
        timing_row.addWidget(self.max_attempts, 1)
        timing_row.addStretch(2)
        form.addLayout(timing_row)

        self.prompt_template = QLineEdit(DEFAULT_PROMPT)
        self.prompt_template.setPlaceholderText("{keyword}")
        template_row = QHBoxLayout()
        template_row.addWidget(QLabel("提问内容"))
        template_row.addWidget(self.prompt_template, 1)
        template_hint = QLabel("默认直接发送关键词；{keyword} 代表当前关键词")
        template_hint.setObjectName("muted")
        template_row.addWidget(template_hint)
        form.addLayout(template_row)

        action_row = QHBoxLayout()
        immediate_hint = QLabel("任务异常会自动释放账号并按设置重试；已识别资料会即时落库")
        immediate_hint.setObjectName("muted")
        action_row.addWidget(immediate_hint)
        action_row.addStretch()
        create_button = QPushButton("开始采集")
        create_button.setObjectName("primaryButton")
        create_button.setMinimumWidth(140)
        create_button.clicked.connect(self.create_job)
        action_row.addWidget(create_button)
        form.addLayout(action_row)
        layout.addWidget(form_group)

        jobs_group = QGroupBox("实时任务进度")
        jobs_group.setMinimumHeight(285)
        jobs_layout = QVBoxLayout(jobs_group)
        self.jobs_summary = QLabel("暂无运行中的任务")
        self.jobs_summary.setObjectName("muted")
        jobs_layout.addWidget(self.jobs_summary)
        self.jobs_scroll = QScrollArea()
        self.jobs_scroll.setWidgetResizable(True)
        self.jobs_scroll.setFrameShape(QFrame.Shape.NoFrame)
        self.jobs_scroll.setMinimumHeight(220)
        self.jobs_container = QWidget()
        self.jobs_cards_layout = QVBoxLayout(self.jobs_container)
        self.jobs_cards_layout.setContentsMargins(0, 0, 0, 0)
        self.jobs_cards_layout.setSpacing(10)
        self.jobs_scroll.setWidget(self.jobs_container)
        jobs_layout.addWidget(self.jobs_scroll)
        layout.addWidget(jobs_group)
        layout.addStretch()
        return page

    def _update_keyword_count(self) -> None:
        count = len(normalize_keywords(self.keywords.toPlainText().splitlines()))
        self.keyword_count.setText(f"{count} 个")

    def _build_accounts_page(self) -> QWidget:
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(4, 8, 4, 4)
        layout.setSpacing(12)
        guide = QLabel(
            "每个账号拥有独立 Cookie 与缓存。先打开账号标签并完成登录，"
            "状态变为“可采集”后即可加入任务。"
        )
        guide.setObjectName("workflowBanner")
        guide.setWordWrap(True)
        layout.addWidget(guide)
        controls = QGroupBox("软件内账号浏览器")
        controls_layout = QHBoxLayout(controls)
        self.new_account = QLineEdit()
        self.new_account.setPlaceholderText("账号名称，例如：账号01")
        create_button = QPushButton("创建并在软件内打开")
        create_button.setObjectName("primaryButton")
        create_button.clicked.connect(self.create_account)
        refresh_button = QPushButton("刷新状态")
        refresh_button.clicked.connect(self.refresh_accounts)
        controls_layout.addWidget(self.new_account, 1)
        controls_layout.addWidget(create_button)
        controls_layout.addWidget(refresh_button)
        layout.addWidget(controls)
        self.accounts_summary = QLabel("正在读取账号状态…")
        self.accounts_summary.setObjectName("muted")
        layout.addWidget(self.accounts_summary)
        self.accounts_scroll = QScrollArea()
        self.accounts_scroll.setWidgetResizable(True)
        self.accounts_scroll.setFrameShape(QFrame.Shape.NoFrame)
        self.accounts_container = QWidget()
        self.accounts_cards_layout = QVBoxLayout(self.accounts_container)
        self.accounts_cards_layout.setContentsMargins(0, 0, 0, 0)
        self.accounts_cards_layout.setSpacing(10)
        self.accounts_scroll.setWidget(self.accounts_container)
        layout.addWidget(self.accounts_scroll, 1)
        return page

    def _build_history_page(self) -> QWidget:
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(4, 8, 4, 4)
        layout.setSpacing(12)
        header = QHBoxLayout()
        title = QLabel("已完成、失败和已取消的任务")
        title.setObjectName("muted")
        header.addWidget(title)
        header.addStretch()
        refresh_button = QPushButton("刷新")
        refresh_button.clicked.connect(self.refresh_history)
        sync_button = QPushButton("同步平台信息")
        sync_button.setObjectName("secondaryButton")
        sync_button.setToolTip("按最新导入的平台规则回填历史记录的平台类型")
        sync_button.clicked.connect(self.sync_platform_info)
        header.addWidget(refresh_button)
        header.addWidget(sync_button)
        layout.addLayout(header)
        self.history_table = QTableWidget(0, 7)
        self.history_table.setHorizontalHeaderLabels(
            ["任务", "状态", "关键词", "成功", "失败", "创建时间", "操作"]
        )
        self._configure_table(self.history_table)
        self.history_table.horizontalHeader().setSectionResizeMode(
            0, QHeaderView.ResizeMode.Stretch
        )
        self.history_table.horizontalHeader().setSectionResizeMode(6, QHeaderView.ResizeMode.Fixed)
        self.history_table.horizontalHeader().resizeSection(6, 320)
        layout.addWidget(self.history_table, 1)
        return page

    def _build_results_page(self) -> QWidget:
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(4, 8, 4, 4)
        layout.setSpacing(10)

        self.results_filters = QGroupBox("筛选与导出")
        filter_layout = QVBoxLayout(self.results_filters)
        first_row = QHBoxLayout()
        self.result_job = QComboBox()
        self.result_job.addItem("全部任务", "")
        self.result_keyword = MultiSelectFilter("全部关键词")
        self.result_platform = QComboBox()
        self.result_platform.addItem("全部平台", "")
        self.result_account = QComboBox()
        self.result_account.addItem("全部账号", "")
        first_row.addWidget(QLabel("任务"))
        first_row.addWidget(self.result_job, 2)
        first_row.addWidget(QLabel("关键词"))
        first_row.addWidget(self.result_keyword, 2)
        first_row.addWidget(QLabel("平台"))
        first_row.addWidget(self.result_platform, 1)
        first_row.addWidget(QLabel("账号"))
        first_row.addWidget(self.result_account, 1)
        filter_layout.addLayout(first_row)

        second_row = QHBoxLayout()
        self.result_date_enabled = QCheckBox("启用日期筛选")
        self.result_date_from = QDateEdit(QDate.currentDate().addDays(-30))
        self.result_date_to = QDateEdit(QDate.currentDate())
        for editor in (self.result_date_from, self.result_date_to):
            editor.setCalendarPopup(True)
            editor.setDisplayFormat("yyyy-MM-dd")
            editor.setMinimumWidth(135)
        query_button = QPushButton("查询")
        query_button.setObjectName("secondaryButton")
        query_button.clicked.connect(self.refresh_results)
        clear_button = QPushButton("重置筛选")
        clear_button.clicked.connect(self.clear_result_filters)
        export_button = QPushButton("导出 Excel")
        export_button.setObjectName("primaryButton")
        export_button.clicked.connect(self.export_results)
        second_row.addWidget(self.result_date_enabled)
        second_row.addWidget(self.result_date_from)
        second_row.addWidget(QLabel("至"))
        second_row.addWidget(self.result_date_to)
        second_row.addStretch()
        second_row.addWidget(clear_button)
        second_row.addWidget(query_button)
        second_row.addWidget(export_button)
        filter_layout.addLayout(second_row)
        layout.addWidget(self.results_filters)

        self.results_metrics = QWidget()
        metrics = QGridLayout(self.results_metrics)
        metrics.setContentsMargins(0, 0, 0, 0)
        metrics.setHorizontalSpacing(12)
        total_card, self.result_total_value = self._metric_card("筛选后资料", "blue")
        task_card, self.result_task_value = self._metric_card("关联任务", "purple")
        keyword_card, self.result_keyword_value = self._metric_card("关键词", "green")
        platform_card, self.result_platform_value = self._metric_card("来源平台", "amber")
        for column, card in enumerate((total_card, task_card, keyword_card, platform_card)):
            metrics.addWidget(card, 0, column)
        layout.addWidget(self.results_metrics)

        self.results_distribution_group = QGroupBox("筛选结果 · 信源分布与占比")
        distribution_layout = QVBoxLayout(self.results_distribution_group)
        distribution_header = QHBoxLayout()
        self.distribution_hint = QLabel("显示主要平台及其在当前筛选结果中的占比")
        self.distribution_hint.setObjectName("muted")
        distribution_header.addWidget(self.distribution_hint)
        distribution_header.addStretch()
        details_button = QPushButton("查看全部平台")
        details_button.setObjectName("linkButton")
        details_button.clicked.connect(self.show_platform_distribution)
        distribution_header.addWidget(details_button)
        distribution_layout.addLayout(distribution_header)
        self.source_distribution_chart = SourceDistributionChart()
        self.source_distribution_chart.setMaximumHeight(320)
        distribution_layout.addWidget(self.source_distribution_chart)
        layout.addWidget(self.results_distribution_group)

        result_header = QHBoxLayout()
        self.results_summary = QLabel("暂无结果")
        self.results_summary.setObjectName("muted")
        result_header.addWidget(self.results_summary)
        result_header.addStretch()
        self.focus_table_button = QPushButton("沉浸查看表格")
        self.focus_table_button.setObjectName("secondaryButton")
        self.focus_table_button.clicked.connect(self.toggle_results_focus)
        copy_button = QPushButton("复制链接")
        copy_button.clicked.connect(self.copy_selected_link)
        open_button = QPushButton("打开链接")
        open_button.clicked.connect(self.open_selected_link)
        result_header.addWidget(self.focus_table_button)
        result_header.addWidget(copy_button)
        result_header.addWidget(open_button)
        layout.addLayout(result_header)
        self.results_table = QTableWidget(0, 6)
        self.results_table.setMinimumHeight(260)
        self.results_table.setHorizontalHeaderLabels(
            ["任务", "日期", "提问关键词", "资料名称", "检索资料链接", "检索资料平台"]
        )
        self._configure_table(self.results_table)
        self.results_table.horizontalHeader().setSectionResizeMode(
            4, QHeaderView.ResizeMode.Stretch
        )
        self.results_table.cellDoubleClicked.connect(self._open_result_cell)
        layout.addWidget(self.results_table, 1)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        scroll.setWidget(page)
        return scroll

    def _build_long_tail_page(self) -> QWidget:
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(4, 8, 4, 4)
        layout.setSpacing(12)

        intro = QLabel(
            "识别跨多个关键词反复出现、但总量小的优质信源平台（气泡大小 = 平均引用密度）"
        )
        intro.setObjectName("workflowBanner")
        layout.addWidget(intro)

        scope = QGroupBox("分析范围")
        scope_layout = QGridLayout(scope)
        scope_layout.setHorizontalSpacing(12)
        scope_layout.setVerticalSpacing(8)

        self.long_tail_job = QComboBox()
        self.long_tail_job.addItem("全部任务", "")
        scope_layout.addWidget(QLabel("任务"), 0, 0)
        scope_layout.addWidget(self.long_tail_job, 0, 1)

        self.long_tail_platform = QComboBox()
        self.long_tail_platform.addItem("全部平台", "")
        scope_layout.addWidget(QLabel("平台"), 0, 2)
        scope_layout.addWidget(self.long_tail_platform, 0, 3)

        self.long_tail_account = QComboBox()
        self.long_tail_account.addItem("全部账号", "")
        scope_layout.addWidget(QLabel("账号"), 0, 4)
        scope_layout.addWidget(self.long_tail_account, 0, 5)

        self.long_tail_date_enabled = QCheckBox("启用日期筛选")
        self.long_tail_date_from = QDateEdit(QDate.currentDate().addDays(-30))
        self.long_tail_date_to = QDateEdit(QDate.currentDate())
        for editor in (self.long_tail_date_from, self.long_tail_date_to):
            editor.setCalendarPopup(True)
            editor.setDisplayFormat("yyyy-MM-dd")
            editor.setMinimumWidth(120)
        scope_layout.addWidget(self.long_tail_date_enabled, 1, 0)
        scope_layout.addWidget(self.long_tail_date_from, 1, 1)
        scope_layout.addWidget(QLabel("至"), 1, 2)
        scope_layout.addWidget(self.long_tail_date_to, 1, 3)
        scope_layout.setColumnStretch(5, 1)
        layout.addWidget(scope)

        params_group = QGroupBox("分类阈值")
        params_layout = QGridLayout(params_group)
        params_layout.setHorizontalSpacing(12)
        params_layout.setVerticalSpacing(8)

        self.long_tail_split_mode = QComboBox()
        self.long_tail_split_mode.addItem("业务阈值", "threshold")
        self.long_tail_split_mode.addItem("中位数", "median")
        params_layout.addWidget(QLabel("分割"), 0, 0)
        params_layout.addWidget(self.long_tail_split_mode, 0, 1)

        self.long_tail_breadth_threshold = QSpinBox()
        self.long_tail_breadth_threshold.setRange(1, 9999)
        self.long_tail_breadth_threshold.setValue(3)
        self.long_tail_breadth_threshold.setToolTip("高广度的判定阈值")
        params_layout.addWidget(QLabel("广度≥"), 0, 2)
        params_layout.addWidget(self.long_tail_breadth_threshold, 0, 3)

        self.long_tail_freq_threshold = QSpinBox()
        self.long_tail_freq_threshold.setRange(1, 99999)
        self.long_tail_freq_threshold.setValue(20)
        self.long_tail_freq_threshold.setToolTip("低频次的判定阈值")
        params_layout.addWidget(QLabel("频次≤"), 0, 4)
        params_layout.addWidget(self.long_tail_freq_threshold, 0, 5)

        self.long_tail_density_threshold = QDoubleSpinBox()
        self.long_tail_density_threshold.setRange(0.1, 999.0)
        self.long_tail_density_threshold.setValue(5.0)
        self.long_tail_density_threshold.setDecimals(1)
        self.long_tail_density_threshold.setSingleStep(0.5)
        self.long_tail_density_threshold.setToolTip("目标长尾的最大密度")
        params_layout.addWidget(QLabel("密度≤"), 1, 0)
        params_layout.addWidget(self.long_tail_density_threshold, 1, 1)

        self.long_tail_noise_density = QDoubleSpinBox()
        self.long_tail_noise_density.setRange(1.0, 999.0)
        self.long_tail_noise_density.setValue(20.0)
        self.long_tail_noise_density.setDecimals(1)
        self.long_tail_noise_density.setSingleStep(0.5)
        self.long_tail_noise_density.setToolTip("虚假长尾的最小密度")
        params_layout.addWidget(QLabel("虚假密度≥"), 1, 2)
        params_layout.addWidget(self.long_tail_noise_density, 1, 3)

        self.long_tail_log_scale = QCheckBox("Y 轴对数")
        self.long_tail_log_scale.setChecked(True)
        params_layout.addWidget(self.long_tail_log_scale, 1, 4)

        self.long_tail_x_log_scale = QCheckBox("X 轴对数")
        params_layout.addWidget(self.long_tail_x_log_scale, 1, 5)

        analyze_button = QPushButton("分析长尾信源")
        analyze_button.setObjectName("primaryButton")
        analyze_button.clicked.connect(self.analyze_long_tail)
        params_layout.addWidget(analyze_button, 1, 6)
        params_layout.setColumnStretch(6, 1)
        layout.addWidget(params_group)

        actions = QHBoxLayout()
        actions.setSpacing(10)
        export_button = QPushButton("导出优质长尾 Excel")
        export_button.setObjectName("secondaryButton")
        export_button.clicked.connect(self.export_long_tail_excel)
        copy_button = QPushButton("复制信源限定词")
        copy_button.setObjectName("linkButton")
        copy_button.clicked.connect(self.copy_long_tail_keywords)
        actions.addWidget(export_button)
        actions.addWidget(copy_button)
        actions.addStretch()
        layout.addLayout(actions)

        self.long_tail_summary = QLabel("设置分析范围后，点击“分析长尾信源”查看四象限图与推荐名单")
        self.long_tail_summary.setObjectName("muted")
        self.long_tail_summary.setWordWrap(True)
        layout.addWidget(self.long_tail_summary)

        self.long_tail_tabs = QTabWidget()
        self.long_tail_target_table = QTableWidget(0, 8)
        self.long_tail_target_table.setHorizontalHeaderLabels(
            ["平台", "域名", "代表性链接", "频次", "广度", "密度", "类型", "覆盖关键词"]
        )
        self._configure_table(self.long_tail_target_table)
        self.long_tail_target_table.horizontalHeader().setSectionResizeMode(
            1, QHeaderView.ResizeMode.Stretch
        )
        self.long_tail_target_table.setSelectionBehavior(
            QAbstractItemView.SelectionBehavior.SelectRows
        )
        self.long_tail_target_table.cellDoubleClicked.connect(self._open_long_tail_link)

        self.long_tail_quadrant_table = QTableWidget(0, 6)
        self.long_tail_quadrant_table.setHorizontalHeaderLabels(
            ["象限", "平台", "频次", "广度", "密度", "类型"]
        )
        self._configure_table(self.long_tail_quadrant_table)
        self.long_tail_quadrant_table.horizontalHeader().setSectionResizeMode(
            1, QHeaderView.ResizeMode.Stretch
        )
        self.long_tail_quadrant_table.setSelectionBehavior(
            QAbstractItemView.SelectionBehavior.SelectRows
        )

        self.long_tail_tabs.addTab(self.long_tail_target_table, "长尾推荐")
        self.long_tail_tabs.addTab(self.long_tail_quadrant_table, "象限清单")
        self.long_tail_tabs.setMaximumHeight(420)
        self.long_tail_target_table.setMinimumHeight(260)
        self.long_tail_quadrant_table.setMinimumHeight(260)

        body = QVBoxLayout()
        body.setSpacing(12)
        self.long_tail_chart = LongTailChart()
        body.addWidget(self.long_tail_chart, 3)
        body.addWidget(self.long_tail_tabs, 1)
        layout.addLayout(body, 1)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        scroll.setWidget(page)
        return scroll

    def _build_comparison_page(self) -> QWidget:
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        page = QWidget()
        page.setMinimumWidth(900)
        layout = QVBoxLayout(page)
        layout.setContentsMargins(4, 8, 4, 4)
        layout.setSpacing(10)

        intro = QLabel(
            "对比两个任务群的 AI 引用信源：A 群与 B 群各自勾选若干历史任务，分析平台来源的变化"
        )
        intro.setObjectName("workflowBanner")
        layout.addWidget(intro)

        global_filter = QGroupBox("全局筛选")
        global_filter_layout = QHBoxLayout(global_filter)
        self.compare_keyword = MultiSelectFilter("全部关键词")
        self.compare_account = QComboBox()
        self.compare_account.addItem("全部账号", "")
        global_filter_layout.addWidget(QLabel("关键词"))
        global_filter_layout.addWidget(self.compare_keyword, 2)
        global_filter_layout.addWidget(QLabel("账号"))
        global_filter_layout.addWidget(self.compare_account, 1)
        layout.addWidget(global_filter)

        groups = QHBoxLayout()
        groups.setSpacing(10)
        group_a = QFrame()
        group_a.setObjectName("compareCard")
        group_a_layout = QHBoxLayout(group_a)
        group_a_layout.addWidget(QLabel("A 任务群"))
        self.compare_jobs_a = MultiSelectFilter("选择任务")
        group_a_layout.addWidget(self.compare_jobs_a, 1)
        groups.addWidget(group_a, 1)

        group_b = QFrame()
        group_b.setObjectName("compareCard")
        group_b_layout = QHBoxLayout(group_b)
        group_b_layout.addWidget(QLabel("B 任务群"))
        self.compare_jobs_b = MultiSelectFilter("选择任务")
        group_b_layout.addWidget(self.compare_jobs_b, 1)
        groups.addWidget(group_b, 1)
        compare_button = QPushButton("分析信源变化")
        compare_button.setObjectName("primaryButton")
        compare_button.clicked.connect(self._on_compare_button_clicked)
        groups.addWidget(compare_button)
        layout.addLayout(groups)

        comparison_metrics = QGridLayout()
        comparison_metrics.setHorizontalSpacing(10)
        a_card, self.compare_a_value = self._metric_card("A 任务群信源平台", "blue")
        b_card, self.compare_b_value = self._metric_card("B 任务群信源平台", "purple")
        added_card, self.compare_added_value = self._metric_card("新增平台", "green")
        removed_card, self.compare_removed_value = self._metric_card("掉出平台", "amber")
        for column, card in enumerate((a_card, b_card, added_card, removed_card)):
            comparison_metrics.addWidget(card, 0, column)
        layout.addLayout(comparison_metrics)

        changes = QHBoxLayout()
        changes.setSpacing(10)
        self._compare_source_rows: dict[str, list[dict[str, Any]]] = {}

        def make_source_card(
            title: str, tone: str, status: str
        ) -> tuple[QGroupBox, QLabel, QPushButton]:
            group = QGroupBox(title)
            group.setProperty("tone", tone)
            group.setMaximumHeight(260)
            card_layout = QVBoxLayout(group)
            card_layout.setSpacing(6)
            label = QLabel("暂无")
            label.setObjectName("sourceList")
            label.setWordWrap(True)
            card_layout.addWidget(label, 1)
            button = QPushButton("查看全部")
            button.setObjectName("secondaryButton")
            button.setVisible(False)
            button.clicked.connect(lambda _, s=status: self._show_comparison_source_list(s))
            card_layout.addWidget(button)
            changes.addWidget(group, 1)
            return group, label, button

        _, self.compare_added_sources, self.compare_added_more = make_source_card(
            "＋ 新增信源（B 有、A 无）", "positive", "added"
        )
        _, self.compare_removed_sources, self.compare_removed_more = make_source_card(
            "－ 掉出信源（A 有、B 无）", "negative", "removed"
        )
        _, self.compare_continued_sources, self.compare_continued_more = make_source_card(
            "＝ 持续出现（A、B 均有）", "neutral", "continued"
        )
        layout.addLayout(changes)

        table_group = QGroupBox("信源变化明细")
        table_layout = QVBoxLayout(table_group)
        self.comparison_summary = QLabel("为 A、B 任务群各选择至少一个任务后，点击“分析信源变化”")
        self.comparison_summary.setObjectName("muted")
        table_layout.addWidget(self.comparison_summary)
        self.comparison_table = QTableWidget(0, 6)
        self.comparison_table.setHorizontalHeaderLabels(
            ["信源平台", "变化状态", "A 出现次数", "A 占比", "B 出现次数", "B 占比"]
        )
        self._configure_table(self.comparison_table)
        self.comparison_table.horizontalHeader().setSectionResizeMode(
            0, QHeaderView.ResizeMode.Stretch
        )
        table_layout.addWidget(self.comparison_table)
        layout.addWidget(table_group, 1)
        scroll.setWidget(page)
        return scroll

    def _show_comparison_source_list(self, status: str) -> None:
        rows = self._compare_source_rows.get(status, [])
        titles = {
            "added": "新增信源（B 有、A 无）",
            "removed": "掉出信源（A 有、B 无）",
            "continued": "持续出现信源（A、B 均有）",
        }
        dialog = QDialog(self)
        dialog.setWindowTitle(titles.get(status, "信源列表"))
        dialog.setMinimumSize(700, 500)
        dialog_layout = QVBoxLayout(dialog)
        table = QTableWidget(len(rows), 5)
        table.setHorizontalHeaderLabels(
            ["信源平台", "A 出现次数", "B 出现次数", "变化量", "变化率"]
        )
        self._configure_table(table)
        table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        for index, row in enumerate(rows):
            change_rate = row.get("change_rate")
            values = [
                row["platform"],
                str(row["a_count"]),
                str(row["b_count"]),
                f"{int(row['delta']):+d}",
                f"{change_rate}%" if change_rate is not None else "-",
            ]
            for column, value in enumerate(values):
                item = QTableWidgetItem(value)
                if column >= 1:
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                table.setItem(index, column, item)
        dialog_layout.addWidget(table)
        close_button = QPushButton("关闭")
        close_button.clicked.connect(dialog.close)
        dialog_layout.addWidget(close_button)
        dialog.exec()

    def _on_compare_button_clicked(self) -> None:
        if not self.compare_jobs_a.selected_values() or not self.compare_jobs_b.selected_values():
            QMessageBox.information(
                self,
                "提示",
                "请为 A 任务群和 B 任务群各选择至少一个任务",
            )
            return
        self.refresh_source_comparison()

    def _build_platforms_page(self) -> QWidget:
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(4, 8, 4, 4)
        layout.setSpacing(10)

        toolbar = QHBoxLayout()
        self.platform_info_summary = QLabel("加载中...")
        self.platform_info_summary.setObjectName("muted")
        toolbar.addWidget(self.platform_info_summary)
        toolbar.addStretch()
        import_button = QPushButton("导入")
        import_button.setObjectName("primaryButton")
        import_button.setToolTip("导入 Excel：URL / 平台名 / 平台类型")
        import_button.clicked.connect(self.import_platforms)
        toolbar.addWidget(import_button)
        layout.addLayout(toolbar)

        self.platform_info_table = QTableWidget(0, 3)
        self.platform_info_table.setHorizontalHeaderLabels(["URL（域名）", "平台名", "平台类型"])
        self._configure_table(self.platform_info_table)
        self.platform_info_table.horizontalHeader().setSectionResizeMode(
            0, QHeaderView.ResizeMode.Stretch
        )
        self.platform_info_table.horizontalHeader().setSectionResizeMode(
            1, QHeaderView.ResizeMode.Fixed
        )
        self.platform_info_table.horizontalHeader().setSectionResizeMode(
            2, QHeaderView.ResizeMode.Fixed
        )
        self.platform_info_table.horizontalHeader().resizeSection(1, 160)
        self.platform_info_table.horizontalHeader().resizeSection(2, 160)
        self.platform_info_table.verticalHeader().setDefaultSectionSize(28)
        self.platform_info_table.setMinimumHeight(400)
        layout.addWidget(self.platform_info_table, 1)
        return page

    def _build_update_page(self) -> QWidget:
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(4, 8, 4, 4)
        layout.setSpacing(14)

        current_row = QHBoxLayout()
        variant_names = {
            "single": "单文件版",
            "portable": "便携版",
            "unknown": "开发环境",
        }
        current_variant = _detect_variant()
        variant_text = variant_names.get(current_variant, current_variant)
        self.update_current_version_label = QLabel(f"当前版本：v{__version__}（{variant_text}）")
        current_row.addWidget(self.update_current_version_label)
        current_row.addStretch()
        self.update_auto_check_checkbox = QCheckBox("启动时自动检查更新")
        self.update_auto_check_checkbox.setChecked(
            self.backend.settings_store.settings.auto_check_updates
        )
        self.update_auto_check_checkbox.stateChanged.connect(self._on_update_auto_check_changed)
        current_row.addWidget(self.update_auto_check_checkbox)
        layout.addLayout(current_row)

        check_row = QHBoxLayout()
        self.update_check_button = QPushButton("立即检查更新")
        self.update_check_button.setObjectName("primaryButton")
        self.update_check_button.clicked.connect(self._check_for_updates)
        check_row.addWidget(self.update_check_button)
        check_row.addStretch()
        layout.addLayout(check_row)

        self.update_status_label = QLabel("点击“立即检查更新”按钮开始检查")
        self.update_status_label.setObjectName("muted")
        layout.addWidget(self.update_status_label)

        info_group = QGroupBox("最新版本信息")
        info_layout = QVBoxLayout(info_group)
        info_layout.setSpacing(10)

        self.update_local_version_label = QLabel(f"本地版本：v{__version__}")
        self.update_local_version_label.setObjectName("muted")
        info_layout.addWidget(self.update_local_version_label)

        self.update_latest_version_label = QLabel("最新版本：—")
        self.update_published_at_label = QLabel("发布时间：—")
        info_layout.addWidget(self.update_latest_version_label)
        info_layout.addWidget(self.update_published_at_label)

        self.update_release_notes = QTextEdit()
        self.update_release_notes.setReadOnly(True)
        self.update_release_notes.setPlaceholderText("更新内容将显示在这里")
        self.update_release_notes.setMinimumHeight(240)
        info_layout.addWidget(self.update_release_notes, 1)

        self.update_progress_bar = QProgressBar()
        self.update_progress_bar.setRange(0, 100)
        self.update_progress_bar.setValue(0)
        self.update_progress_bar.setTextVisible(True)
        self.update_progress_bar.setVisible(False)
        info_layout.addWidget(self.update_progress_bar)

        self.update_downloaded_path_label = QLabel("")
        self.update_downloaded_path_label.setWordWrap(True)
        self.update_downloaded_path_label.setObjectName("muted")
        self.update_downloaded_path_label.setVisible(False)
        info_layout.addWidget(self.update_downloaded_path_label)

        button_row = QHBoxLayout()
        self.update_download_single_button = QPushButton("下载单文件版")
        self.update_download_single_button.setObjectName("primaryButton")
        self.update_download_single_button.setVisible(False)
        self.update_download_single_button.clicked.connect(
            lambda: self._start_update_download("single")
        )
        button_row.addWidget(self.update_download_single_button)

        self.update_download_portable_button = QPushButton("下载便携版")
        self.update_download_portable_button.setObjectName("primaryButton")
        self.update_download_portable_button.setVisible(False)
        self.update_download_portable_button.clicked.connect(
            lambda: self._start_update_download("portable")
        )
        button_row.addWidget(self.update_download_portable_button)

        self.update_install_button = QPushButton("立即安装")
        self.update_install_button.setObjectName("primaryButton")
        self.update_install_button.setEnabled(False)
        self.update_install_button.setVisible(False)
        self.update_install_button.clicked.connect(self._install_update)
        button_row.addWidget(self.update_install_button)

        button_row.addStretch()

        self.update_open_release_button = QPushButton("去下载页面")
        self.update_open_release_button.setEnabled(False)
        self.update_open_release_button.clicked.connect(self._open_update_download_page)
        button_row.addWidget(self.update_open_release_button)
        info_layout.addLayout(button_row)

        info_group.setEnabled(False)
        self.update_info_group = info_group
        layout.addWidget(info_group, 1)

        layout.addStretch()
        self._update_download_url = ""
        return page

    def _on_update_auto_check_changed(self, state: int) -> None:
        self.backend.settings_store.update({"auto_check_updates": bool(state)})

    def _auto_check_update_if_enabled(self) -> None:
        if self.backend.settings_store.settings.auto_check_updates:
            self._check_for_updates()

    def _check_for_updates(self) -> None:
        self.update_check_button.setEnabled(False)
        self.update_status_label.setText("正在检查...")
        self.update_status_label.setObjectName("muted")

        async def check() -> UpdateInfo | None:
            checker = UpdateChecker(current_version=__version__)
            info = await checker.fetch_latest_release()
            if info is not None:
                save_cached_update_info(self.backend.settings_store.data_root, info)
            return info

        def apply(info: UpdateInfo | None) -> None:
            self._apply_update_info(info)

        def on_error(exc: BaseException) -> None:
            self.update_status_label.setText(f"检查失败：{exc}")
            self.update_check_button.setEnabled(True)

        self._watch(
            self.backend.submit(check()),
            apply,
            error_callback=on_error,
            timeout_seconds=15,
            label="检查更新",
        )

    def _load_cached_update_info(self) -> None:
        """启动时读取本地缓存的更新信息并直接展示。"""
        info = load_cached_update_info(self.backend.settings_store.data_root)
        if info is not None:
            self._apply_update_info(info, from_cache=True)

    def _apply_update_info(self, info: UpdateInfo | None, *, from_cache: bool = False) -> None:
        self.update_check_button.setEnabled(True)
        self.update_info_group.setEnabled(True)
        self._current_update_info = info
        if info is None:
            self.update_status_label.setText("检查失败：无法获取版本信息")
            self.update_latest_version_label.setText("最新版本：—")
            self.update_published_at_label.setText("发布时间：—")
            self.update_release_notes.clear()
            self.update_open_release_button.setEnabled(False)
            self.update_downloaded_path_label.setVisible(False)
            self._update_download_url = ""
            self._refresh_update_buttons()
            self._refresh_downloaded_buttons(None)
            return

        self.update_latest_version_label.setText(f"最新版本：{info.tag_name}")
        self.update_published_at_label.setText(
            f"发布时间：{self._format_published_at(info.published_at)}"
        )
        if hasattr(self.update_release_notes, "setMarkdown"):
            self.update_release_notes.setMarkdown(info.release_notes or "作者未提供更新说明")
        else:
            self.update_release_notes.setPlainText(info.release_notes or "作者未提供更新说明")

        settings = self.backend.settings_store.settings
        ignored = (
            settings.last_ignored_version
            and _normalize_version(settings.last_ignored_version) == info.version
        )
        checker = UpdateChecker(current_version=__version__)
        has_update = checker.is_newer(info.version)
        cache_hint = "（来自本地缓存）" if from_cache else ""
        if ignored:
            self.update_status_label.setText(f"已忽略版本 {info.tag_name}{cache_hint}")
        elif has_update:
            self.update_status_label.setText(f"发现新版本 {info.tag_name}{cache_hint}")
        else:
            self.update_status_label.setText(f"已是最新版{cache_hint}")

        self._update_download_url = info.release_url
        self.update_open_release_button.setEnabled(bool(info.release_url))

        # 控制两个版本下载按钮的显示/可用状态
        self._refresh_update_buttons()

        # 刷新已下载文件状态
        self._refresh_downloaded_buttons(info)

    def _refresh_update_buttons(self) -> None:
        """根据当前更新信息刷新下载按钮的显示和可用状态。"""
        info = self._current_update_info
        if info is None:
            self.update_download_single_button.setVisible(False)
            self.update_download_portable_button.setVisible(False)
            return

        checker = UpdateChecker(current_version=__version__)
        has_update = checker.is_newer(info.version)
        single_asset = checker.asset_for_variant(info, "single")
        portable_asset = checker.asset_for_variant(info, "portable")
        downloading = self._downloading_variant is not None

        self.update_download_single_button.setVisible(bool(has_update and single_asset))
        self.update_download_single_button.setEnabled(
            bool(has_update and single_asset and not downloading)
        )
        self.update_download_portable_button.setVisible(bool(has_update and portable_asset))
        self.update_download_portable_button.setEnabled(
            bool(has_update and portable_asset and not downloading)
        )

    def _refresh_downloaded_buttons(self, info: UpdateInfo | None) -> None:
        """根据本地已下载文件刷新按钮和路径显示。"""
        self._downloaded_paths = {}
        if info is None:
            self.update_install_button.setVisible(False)
            self.update_install_button.setEnabled(False)
            self.update_downloaded_path_label.setVisible(False)
            return

        for variant in ("single", "portable"):
            asset = UpdateChecker(current_version=__version__).asset_for_variant(info, variant)
            if asset is None:
                continue
            path = self._update_download_dir / asset.name
            if path.exists() and path.stat().st_size > 0:
                self._downloaded_paths[variant] = path

        any_downloaded = bool(self._downloaded_paths)
        self.update_install_button.setVisible(any_downloaded)
        self.update_install_button.setEnabled(any_downloaded)

        if any_downloaded:
            lines = [f"已下载：{path}" for variant, path in self._downloaded_paths.items()]
            self.update_downloaded_path_label.setText("\n".join(lines))
            self.update_downloaded_path_label.setVisible(True)
        else:
            self.update_downloaded_path_label.setVisible(False)

    def _format_published_at(self, published_at: str) -> str:
        if not published_at:
            return "—"
        try:
            from datetime import datetime, timezone

            dt = datetime.fromisoformat(published_at.replace("Z", "+00:00"))
            dt = dt.astimezone(timezone.utc)
            return dt.strftime("%Y-%m-%d %H:%M:%S UTC")
        except ValueError:
            return published_at

    def _open_update_download_page(self) -> None:
        url = self._update_download_url
        if url:
            QDesktopServices.openUrl(QUrl(url))

    def _start_update_download(self, variant: str) -> None:
        info = self._current_update_info
        if info is None:
            return
        checker = UpdateChecker(current_version=__version__)
        asset = checker.asset_for_variant(info, variant)
        if asset is None:
            self.update_status_label.setText(f"未找到 {variant} 版本的下载链接")
            return

        self._downloading_variant = variant
        self._refresh_update_buttons()
        self.update_install_button.setEnabled(False)
        self.update_progress_bar.setVisible(True)
        self.update_progress_bar.setValue(0)
        self.update_status_label.setText(f"正在下载 {asset.name}...")

        async def download() -> DownloadResult:
            return await checker.download_asset(
                info,
                asset,
                self._update_download_dir,
                variant=variant,
                progress_callback=lambda d, t: self.download_progress.emit(d, t),
            )

        def apply(result: DownloadResult) -> None:
            self._on_download_finished(result, variant)

        def on_error(exc: BaseException) -> None:
            self._on_download_error(exc)

        self._watch(
            self.backend.submit(download()),
            apply,
            error_callback=on_error,
            timeout_seconds=600,
            label=f"下载 {variant}",
        )

    def _on_download_progress(self, downloaded: int, total: int) -> None:
        percent = int(min(100, downloaded * 100 / total)) if total > 0 else 0
        self.update_progress_bar.setValue(percent)
        size_mb = downloaded / (1024 * 1024)
        if total > 0:
            total_mb = total / (1024 * 1024)
            self.update_status_label.setText(
                f"已下载 {size_mb:.2f} / {total_mb:.2f} MB ({percent}%)"
            )
        else:
            self.update_status_label.setText(f"已下载 {size_mb:.2f} MB")

    def _on_download_finished(self, result: DownloadResult, variant: str) -> None:
        self._downloading_variant = None
        self._refresh_update_buttons()
        self.update_progress_bar.setVisible(False)
        self.update_status_label.setText(
            f"{result.asset.name} 下载完成（{result.verification_method} 校验通过）"
        )
        self._downloaded_paths[variant] = result.path
        self._refresh_downloaded_buttons(self._current_update_info)

    def _on_download_error(self, exc: BaseException) -> None:
        self._downloading_variant = None
        self._refresh_update_buttons()
        self.update_progress_bar.setVisible(False)
        self.update_progress_bar.setValue(0)
        self.update_status_label.setText(f"下载失败：{exc}")
        self._refresh_downloaded_buttons(self._current_update_info)

    def _install_update(self) -> None:
        """Task 5 占位：目前仅打开下载目录并提示用户。"""
        paths = list(self._downloaded_paths.values())
        if not paths:
            return
        directory = paths[0].parent
        if directory.exists():
            QDesktopServices.openUrl(QUrl.fromLocalFile(str(directory)))
        QMessageBox.information(
            self,
            "安装更新",
            "自动替换安装功能将在后续版本实现。\n已为你打开下载目录，可手动覆盖安装。",
        )

    def _build_schedules_page(self) -> QWidget:
        page = QWidget()
        outer = QVBoxLayout(page)
        outer.setContentsMargins(0, 0, 0, 0)
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        content = QWidget()
        content.setMinimumWidth(900)
        layout = QVBoxLayout(content)
        layout.setContentsMargins(4, 8, 4, 4)
        layout.setSpacing(14)
        scroll.setWidget(content)
        outer.addWidget(scroll)

        intro = QLabel(
            "先创建任务模板，再基于模板创建定时计划。计划触发时会按模板的最新配置自动生成采集任务。"
        )
        intro.setObjectName("workflowBanner")
        intro.setWordWrap(True)
        layout.addWidget(intro)

        # ------------------- 任务模板 -------------------
        templates_group = QGroupBox("任务模板")
        templates_layout = QVBoxLayout(templates_group)
        templates_layout.setSpacing(12)

        template_form = QGridLayout()
        template_form.setSpacing(10)
        template_form.addWidget(QLabel("模板名称"), 0, 0)
        self.template_name = QLineEdit()
        self.template_name.setPlaceholderText("例如：每日品牌词调研")
        template_form.addWidget(self.template_name, 0, 1, 1, 3)

        template_form.addWidget(QLabel("关键词"), 1, 0, Qt.AlignmentFlag.AlignTop)
        self.template_keywords = QTextEdit()
        self.template_keywords.setPlaceholderText("每行一个关键词")
        self.template_keywords.setMinimumHeight(80)
        self.template_keywords.setMaximumHeight(120)
        template_form.addWidget(self.template_keywords, 1, 1, 1, 3)

        template_form.addWidget(QLabel("提问模板"), 2, 0)
        self.template_prompt_template = QLineEdit(DEFAULT_PROMPT)
        self.template_prompt_template.setPlaceholderText("{keyword}")
        template_form.addWidget(self.template_prompt_template, 2, 1, 1, 3)

        template_form.addWidget(QLabel("关键词间隔"), 3, 0)
        self.template_interval_seconds = QSpinBox()
        self.template_interval_seconds.setRange(1, 86400)
        self.template_interval_seconds.setValue(10)
        self.template_interval_seconds.setSuffix(" 秒")
        template_form.addWidget(self.template_interval_seconds, 3, 1)

        template_form.addWidget(QLabel("账号冷却"), 3, 2)
        self.template_account_cooldown_seconds = QSpinBox()
        self.template_account_cooldown_seconds.setRange(0, 86400)
        self.template_account_cooldown_seconds.setValue(0)
        self.template_account_cooldown_seconds.setSuffix(" 秒")
        template_form.addWidget(self.template_account_cooldown_seconds, 3, 3)

        template_form.addWidget(QLabel("最多尝试"), 4, 0)
        self.template_max_attempts = QSpinBox()
        self.template_max_attempts.setRange(1, 3)
        self.template_max_attempts.setValue(2)
        template_form.addWidget(self.template_max_attempts, 4, 1)
        templates_layout.addLayout(template_form)

        template_buttons = QHBoxLayout()
        template_buttons.addStretch()
        self.template_cancel_button = QPushButton("取消")
        self.template_cancel_button.setObjectName("secondaryButton")
        self.template_cancel_button.clicked.connect(self._reset_template_form)
        self.template_save_button = QPushButton("保存模板")
        self.template_save_button.setObjectName("primaryButton")
        self.template_save_button.clicked.connect(self.save_job_template)
        template_buttons.addWidget(self.template_cancel_button)
        template_buttons.addWidget(self.template_save_button)
        templates_layout.addLayout(template_buttons)

        self.templates_table = QTableWidget(0, 5)
        self.templates_table.setHorizontalHeaderLabels(
            ["名称", "关键词数", "提问模板", "编辑", "删除"]
        )
        self._configure_table(self.templates_table)
        self.templates_table.horizontalHeader().setSectionResizeMode(
            0, QHeaderView.ResizeMode.Stretch
        )
        self.templates_table.horizontalHeader().setSectionResizeMode(
            2, QHeaderView.ResizeMode.Stretch
        )
        self.templates_table.verticalHeader().setDefaultSectionSize(40)
        templates_layout.addWidget(self.templates_table, 1)

        layout.addWidget(templates_group)

        # ------------------- 触发计划 -------------------
        schedules_group = QGroupBox("触发计划")
        schedules_layout = QVBoxLayout(schedules_group)
        schedules_layout.setSpacing(12)

        schedule_form = QGridLayout()
        schedule_form.setSpacing(10)
        schedule_form.addWidget(QLabel("计划名称"), 0, 0)
        self.schedule_name = QLineEdit()
        self.schedule_name.setPlaceholderText("例如：每天早上 9 点执行")
        schedule_form.addWidget(self.schedule_name, 0, 1, 1, 3)

        schedule_form.addWidget(QLabel("选择模板"), 1, 0)
        self.schedule_template_id = QComboBox()
        schedule_form.addWidget(self.schedule_template_id, 1, 1, 1, 3)

        schedule_form.addWidget(QLabel("触发类型"), 2, 0)
        self.schedule_type = QComboBox()
        self.schedule_type.addItem("按间隔", "interval")
        self.schedule_type.addItem("一次性", "once")
        self.schedule_type.addItem("每日定时", "daily")
        self.schedule_type.currentIndexChanged.connect(self._on_schedule_type_changed)
        schedule_form.addWidget(self.schedule_type, 2, 1)

        schedule_form.addWidget(QLabel("触发参数"), 2, 2)
        self.schedule_value = QLineEdit()
        self.schedule_value.setPlaceholderText("秒数，例如 3600")
        schedule_form.addWidget(self.schedule_value, 2, 3)
        schedules_layout.addLayout(schedule_form)

        schedule_buttons = QHBoxLayout()
        schedule_buttons.addStretch()
        create_schedule_button = QPushButton("创建计划")
        create_schedule_button.setObjectName("primaryButton")
        create_schedule_button.clicked.connect(self.create_schedule)
        schedule_buttons.addWidget(create_schedule_button)
        schedules_layout.addLayout(schedule_buttons)

        self.schedules_table = QTableWidget(0, 8)
        self.schedules_table.setHorizontalHeaderLabels(
            ["名称", "关联模板", "触发类型", "触发参数", "下次执行", "运行次数", "启用", "操作"]
        )
        self._configure_table(self.schedules_table)
        self.schedules_table.horizontalHeader().setSectionResizeMode(
            0, QHeaderView.ResizeMode.Stretch
        )
        self.schedules_table.horizontalHeader().setSectionResizeMode(
            1, QHeaderView.ResizeMode.Stretch
        )
        self.schedules_table.horizontalHeader().setSectionResizeMode(
            4, QHeaderView.ResizeMode.Stretch
        )
        self.schedules_table.verticalHeader().setDefaultSectionSize(40)
        schedules_layout.addWidget(self.schedules_table, 1)

        layout.addWidget(schedules_group)
        layout.addStretch()
        return page

    def _on_schedule_type_changed(self, index: int) -> None:
        schedule_type = self.schedule_type.itemData(index)
        placeholders = {
            "interval": "秒数，例如 3600",
            "once": "ISO 时间，例如 2026-08-10T09:00:00",
            "daily": "HH:MM，例如 09:00",
        }
        self.schedule_value.setPlaceholderText(placeholders.get(schedule_type, ""))

    def _reset_template_form(self) -> None:
        self.editing_template_id = None
        self.template_name.clear()
        self.template_keywords.clear()
        self.template_prompt_template.setText(DEFAULT_PROMPT)
        self.template_interval_seconds.setValue(10)
        self.template_account_cooldown_seconds.setValue(0)
        self.template_max_attempts.setValue(2)
        self.template_save_button.setText("保存模板")

    def _update_combo(
        self,
        combo: QComboBox,
        items: list[tuple[str, Any]],
        default_text: str,
    ) -> None:
        """Refresh a combo box only when its items actually changed."""

        current = [(combo.itemText(index), combo.itemData(index)) for index in range(combo.count())]
        if current == [(default_text, "")] + items:
            return
        previous = combo.currentData()
        combo.clear()
        combo.addItem(default_text, "")
        for text, data in items:
            combo.addItem(text, data)
        selected = combo.findData(previous)
        if selected >= 0:
            combo.setCurrentIndex(selected)

    def clear_result_filters(self) -> None:
        self.result_job.setCurrentIndex(0)
        self.result_keyword.clear_selection()
        self.result_platform.setCurrentIndex(0)
        self.result_account.setCurrentIndex(0)
        self.result_date_enabled.setChecked(False)
        self.refresh_results()

    def toggle_results_focus(self) -> None:
        entering = not self.results_focus_mode
        window = self.window()
        if entering:
            self._window_was_maximized = window.isMaximized()
            self._window_was_fullscreen = window.isFullScreen()
        self.results_focus_mode = entering
        for widget in (
            self.header_bar,
            self.results_filters,
            self.results_metrics,
            self.results_distribution_group,
        ):
            widget.setVisible(not entering)
        self.sections.tabBar().setVisible(not entering)
        self.focus_table_button.setText("退出沉浸（Esc）" if entering else "沉浸查看表格")
        if entering or self._window_was_fullscreen:
            window.showFullScreen()
        elif self._window_was_maximized:
            window.showMaximized()
        else:
            window.showNormal()

    def _exit_results_focus(self) -> None:
        if self.results_focus_mode:
            self.toggle_results_focus()

    def show_platform_distribution(self) -> None:
        if not self.platform_distribution_rows:
            QMessageBox.information(self, "提示", "当前筛选条件下没有信源平台")
            return
        dialog = QDialog(self)
        dialog.setWindowTitle("全部信源平台")
        dialog.resize(860, 620)
        layout = QVBoxLayout(dialog)
        layout.setSpacing(12)
        summary = QLabel(
            f"共 {len(self.platform_distribution_rows)} 个平台 · "
            f"{self.platform_distribution_total} 条资料"
        )
        summary.setObjectName("muted")
        layout.addWidget(summary)
        chart = SourceDistributionChart(dialog)
        chart.set_rows(
            self.platform_distribution_rows,
            self.platform_distribution_total,
            detail=True,
        )
        layout.addWidget(chart, 1)
        actions = QHBoxLayout()
        actions.addStretch()
        close = QPushButton("关闭")
        close.setObjectName("secondaryButton")
        close.clicked.connect(dialog.accept)
        actions.addWidget(close)
        layout.addLayout(actions)
        dialog.exec()

    def _configure_table(self, table: QTableWidget) -> None:
        table.setAlternatingRowColors(True)
        table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        table.setShowGrid(False)
        table.verticalHeader().setVisible(False)
        table.verticalHeader().setDefaultSectionSize(54)
        table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)

    def _status_item(self, text: str, status: str) -> QTableWidgetItem:
        item = QTableWidgetItem(text)
        colors = {
            "running": ("#087a55", "#e9f8f2"),
            "completed": ("#087a55", "#e9f8f2"),
            "ready": ("#087a55", "#e9f8f2"),
            "paused": ("#9a6212", "#fff5df"),
            "pending": ("#5e687e", "#f0f2f6"),
            "failed": ("#b42c48", "#fff0f3"),
            "cancelled": ("#b42c48", "#fff0f3"),
            "offline": ("#687386", "#f0f2f6"),
        }
        foreground, background = colors.get(status, colors["offline"])
        item.setForeground(QColor(foreground))
        item.setBackground(QColor(background))
        item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        return item

    def _clear_cards(self, layout: QVBoxLayout) -> None:
        while layout.count():
            item = layout.takeAt(0)
            widget = item.widget()
            if widget is not None:
                widget.deleteLater()

    def _status_badge(self, text: str, tone: str) -> QLabel:
        badge = QLabel(text)
        badge.setObjectName("statusBadge")
        badge.setProperty("tone", tone)
        badge.setAlignment(Qt.AlignmentFlag.AlignCenter)
        badge.setMinimumWidth(76)
        return badge

    def _account_card(self, row: dict[str, Any]) -> QFrame:
        card = QFrame()
        card.setObjectName("accountCard")
        layout = QHBoxLayout(card)
        layout.setContentsMargins(16, 13, 16, 13)
        layout.setSpacing(12)
        avatar = QLabel(row["account_id"][:1].upper())
        avatar.setObjectName("accountAvatar")
        avatar.setAlignment(Qt.AlignmentFlag.AlignCenter)
        avatar.setFixedSize(40, 40)
        layout.addWidget(avatar)
        text_box = QVBoxLayout()
        text_box.setSpacing(3)
        title = QLabel(row["account_id"])
        title.setObjectName("cardTitle")
        detail = QLabel(
            "标签页已打开 · 独立登录环境" if row["started"] else "标签页未打开 · 独立登录环境"
        )
        detail.setObjectName("muted")
        text_box.addWidget(title)
        text_box.addWidget(detail)
        layout.addLayout(text_box, 1)
        is_paused = bool(row.get("is_paused"))
        pause_reason = row.get("pause_reason") or ""
        status_text = (
            "状态超时"
            if row.get("snapshot_error")
            else "已暂停—需处理验证"
            if is_paused or row["needs_captcha"]
            else "可采集"
            if row["chat_ready"]
            else "已登录"
            if row["logged_in"]
            else "未登录"
        )
        tone = (
            "failed"
            if is_paused or row["needs_captcha"]
            else "ready"
            if row["chat_ready"]
            else "paused"
            if row["logged_in"]
            else "offline"
        )
        badge = self._status_badge(status_text, tone)
        if row.get("snapshot_error"):
            badge.setToolTip(row["snapshot_error"])
        elif is_paused and pause_reason:
            badge.setToolTip(pause_reason)
        layout.addWidget(badge)
        captcha_like = (
            is_paused
            or row.get("needs_captcha")
            or "验证码" in (row.get("snapshot_error") or "")
            or "无响应" in (row.get("snapshot_error") or "")
        )
        if captcha_like and row["started"]:
            clear = QPushButton("验证已完成")
            clear.setObjectName("secondaryButton")
            clear.setMinimumWidth(100)
            clear.clicked.connect(lambda _, aid=row["account_id"]: self.clear_account_captcha(aid))
            layout.addWidget(clear)
        if row["started"]:
            focus = QPushButton("切换标签")
            focus.setMinimumWidth(88)
            focus.clicked.connect(lambda _, aid=row["account_id"]: self.focus_account(aid))
            tab_hidden = bool(row.get("tab_hidden"))
            visibility = QPushButton("显示标签" if tab_hidden else "隐藏标签")
            visibility.setObjectName("secondaryButton")
            visibility.setMinimumWidth(88)
            visibility.clicked.connect(
                lambda _, aid=row["account_id"], h=tab_hidden: self.toggle_account_tab_hidden(
                    aid, not h
                )
            )
            close = QPushButton("关闭")
            close.clicked.connect(lambda _, aid=row["account_id"]: self.stop_account(aid))
            layout.addWidget(focus)
            layout.addWidget(visibility)
            layout.addWidget(close)
        else:
            open_button = QPushButton("打开并登录")
            open_button.setObjectName("secondaryButton")
            open_button.setMinimumWidth(100)
            open_button.clicked.connect(lambda _, aid=row["account_id"]: self.open_account(aid))
            layout.addWidget(open_button)
        if row["can_delete"]:
            rename = QPushButton("重命名")
            rename.clicked.connect(lambda _, aid=row["account_id"]: self.rename_account(aid))
            delete = QPushButton("删除")
            delete.setObjectName("dangerButton")
            delete.clicked.connect(lambda _, aid=row["account_id"]: self.delete_account(aid))
            layout.addWidget(rename)
            layout.addWidget(delete)
        return card

    def _job_card(self, row: dict[str, Any]) -> QFrame:
        card = QFrame()
        card.setObjectName("jobCard")
        layout = QVBoxLayout(card)
        layout.setContentsMargins(16, 14, 16, 14)
        layout.setSpacing(9)
        header = QHBoxLayout()
        title = QLabel(row["name"])
        title.setObjectName("cardTitle")
        header.addWidget(title, 1)
        header.addWidget(
            self._status_badge(
                STATUS_TEXT.get(row["status"], row["status"]),
                row["status"],
            )
        )
        layout.addLayout(header)
        completed = row["completed"]
        failed = row["failed"]
        pending = row["pending"]
        summary = QLabel(
            f"共 {row['total']} 个关键词 · 成功 {completed} · 失败 {failed}"
            f" · 进行中 {row['running_tasks']} · 待处理 {pending}"
        )
        summary.setObjectName("muted")
        layout.addWidget(summary)
        progress = QProgressBar()
        progress.setRange(0, 100)
        progress.setValue(row["progress_percent"])
        progress.setFormat(f"{completed + failed}/{row['total']} · %p%")
        progress.setMinimumHeight(22)
        layout.addWidget(progress)
        active = QLabel(
            f"当前执行：{row['active_details']}"
            if row["active_details"]
            else "当前执行：等待可用账号或下一个启动间隔"
        )
        active.setWordWrap(True)
        layout.addWidget(active)
        stage = QLabel(f"最新进展：{row['last_error'] or '等待调度'}")
        stage.setObjectName("muted")
        stage.setWordWrap(True)
        layout.addWidget(stage)
        actions = QHBoxLayout()
        actions.addStretch()
        if row["status"] == "running":
            pause = QPushButton("暂停任务")
            pause.clicked.connect(lambda _, jid=row["id"]: self.set_job_status(jid, "paused"))
            actions.addWidget(pause)
        elif row["status"] == "paused":
            resume = QPushButton("继续任务")
            resume.setObjectName("secondaryButton")
            resume.clicked.connect(lambda _, jid=row["id"]: self.set_job_status(jid, "running"))
            actions.addWidget(resume)
        cancel = QPushButton("取消任务")
        cancel.setObjectName("dangerButton")
        cancel.clicked.connect(lambda _, jid=row["id"]: self.set_job_status(jid, "cancelled"))
        actions.addWidget(cancel)
        layout.addLayout(actions)
        return card

    def _apply_style(self) -> None:
        self.setStyleSheet(
            """
            QWidget {
              font-family: "Microsoft YaHei UI";
              font-size: 13px;
              color: #172033;
            }
            QWidget#dashboard { background: #f7f7f8; }
            QLabel { background: transparent; }
            QLabel#logo {
              color: white;
              background: #272a33;
              border-radius: 11px;
              font-size: 20px;
              font-weight: 800;
            }
            QLabel#title { font-size: 21px; font-weight: 750; color: #17181c; }
            QLabel#muted { color: #728097; }
            QLabel#statusPill {
              color: #087a55;
              background: #e8f8f1;
              border: 1px solid #bce9d7;
              border-radius: 12px;
              padding: 6px 11px;
              font-weight: 650;
            }
            QLabel#localPill, QLabel#countPill {
              color: #5057a8;
              background: #eef0ff;
              border: 1px solid #d8dcff;
              border-radius: 11px;
              padding: 5px 10px;
              font-weight: 650;
            }
            QLabel#workflowBanner {
              color: #4c515d;
              background: #ffffff;
              border: 1px solid #e3e4e8;
              border-radius: 10px;
              padding: 11px 14px;
              font-weight: 650;
            }
            QFrame#metricCard {
              background: white;
              border: 1px solid #e2e3e7;
              border-radius: 10px;
            }
            QLabel#metricValue {
              color: #22242b;
              font-size: 22px;
              font-weight: 800;
            }
            QFrame#jobCard, QFrame#accountCard, QFrame#compareCard {
              background: #ffffff;
              border: 1px solid #e2e3e7;
              border-radius: 10px;
            }
            QLabel#cardTitle {
              color: #222a47;
              font-size: 15px;
              font-weight: 750;
            }
            QLabel#accountAvatar {
              color: #5153d3;
              background: #ededff;
              border: 1px solid #d8d9ff;
              border-radius: 12px;
              font-size: 17px;
              font-weight: 800;
            }
            QLabel#statusBadge {
              border-radius: 11px;
              padding: 5px 10px;
              font-weight: 700;
            }
            QLabel#statusBadge[tone="ready"], QLabel#statusBadge[tone="running"] {
              color: #087a55;
              background: #e8f8f1;
              border: 1px solid #bce9d7;
            }
            QLabel#statusBadge[tone="paused"] {
              color: #94600f;
              background: #fff4dc;
              border: 1px solid #f0d6a2;
            }
            QLabel#statusBadge[tone="failed"] {
              color: #b42c48;
              background: #fff0f3;
              border: 1px solid #f2c8d1;
            }
            QLabel#statusBadge[tone="offline"] {
              color: #657086;
              background: #f0f2f6;
              border: 1px solid #dfe3ea;
            }
            QScrollArea, QScrollArea > QWidget > QWidget {
              background: transparent;
              border: 0;
            }
            QGroupBox {
              background: white;
              border: 1px solid #e2e3e7;
              border-radius: 10px;
              margin-top: 10px;
              padding: 11px;
            }
            QGroupBox::title {
              subcontrol-origin: margin;
              left: 14px;
              padding: 0 7px;
              color: #29314f;
              font-weight: 750;
            }
            QLineEdit, QTextEdit, QSpinBox, QDateEdit, QComboBox {
              background: #ffffff;
              border: 1px solid #d8dfeb;
              border-radius: 8px;
              padding: 8px;
              min-height: 20px;
              selection-background-color: #6264df;
            }
            QLineEdit:focus, QTextEdit:focus, QSpinBox:focus,
            QDateEdit:focus, QComboBox:focus {
              border: 1px solid #7779eb;
            }
            QPushButton {
              background: white;
              border: 1px solid #d2d9e6;
              border-radius: 8px;
              padding: 8px 13px;
              min-height: 20px;
              font-weight: 650;
            }
            QPushButton:hover { color: #4e50ce; border-color: #8f91ec; background: #f7f7ff; }
            QPushButton:pressed { background: #ededff; }
            QPushButton:disabled { color: #9ca6b8; background: #f1f3f7; border-color: #e1e5ec; }
            QPushButton#primaryButton {
              color: white;
              background: #30333b;
              border-color: #30333b;
              padding: 9px 17px;
              min-height: 22px;
            }
            QPushButton#primaryButton:hover { background: #202228; border-color: #202228; }
            QPushButton#secondaryButton {
              color: #30333b;
              background: #f1f2f4;
              border-color: #dfe1e5;
            }
            QPushButton#linkButton {
              color: #5b5ce2;
              background: transparent;
              border: 0;
              padding: 3px 6px;
            }
            QPushButton#dangerButton { color: #b82f49; border-color: #efc3cc; background: #fffafb; }
            QToolButton#filterButton, QPushButton#rangePicker {
              color: #25304a;
              background: white;
              border: 1px solid #d8dfeb;
              border-radius: 8px;
              padding: 8px 26px 8px 10px;
              min-height: 20px;
              text-align: left;
            }
            QToolButton#filterButton:hover, QToolButton#filterButton:focus,
            QPushButton#rangePicker:hover, QPushButton#rangePicker:focus {
              border-color: #7779eb;
              background: #fbfbff;
            }
            QToolButton#filterButton::menu-indicator {
              subcontrol-origin: padding;
              subcontrol-position: center right;
              right: 9px;
            }
            QGroupBox[tone="positive"] {
              background: #f3fbf7;
              border-color: #c9eadb;
            }
            QGroupBox[tone="negative"] {
              background: #fff6f7;
              border-color: #f0d0d6;
            }
            QGroupBox[tone="neutral"] {
              background: #f7f8fb;
              border-color: #dfe4ec;
            }
            QLabel#sourceList {
              color: #39425c;
              font-size: 14px;
              padding: 7px;
              min-height: 42px;
            }
            QMenu#filterMenu {
              background: white;
              border: 1px solid #d8dfeb;
              border-radius: 9px;
              padding: 8px;
            }
            QListWidget {
              background: white;
              border: 1px solid #e2e6ee;
              border-radius: 7px;
              outline: 0;
              padding: 4px;
            }
            QListWidget::item { padding: 7px 5px; border-radius: 5px; }
            QListWidget::item:hover { background: #f1f2ff; }
            QTabWidget::pane {
              background: transparent;
              border: 0;
              top: -1px;
            }
            QTabBar { background: transparent; }
            QTabBar::tab {
              color: #667188;
              background: #eceef1;
              padding: 8px 19px;
              margin-right: 5px;
              border-radius: 9px;
              font-weight: 650;
            }
            QTabBar::tab:hover { color: #252831; background: #e4e6e9; }
            QTabBar::tab:selected {
              color: #202228;
              background: white;
              border: 1px solid #dfe1e5;
            }
            QHeaderView::section {
              color: #5b657b;
              background: #f2f3f5;
              padding: 9px;
              border: 0;
              border-bottom: 1px solid #e2e6ed;
              font-weight: 700;
            }
            QTableWidget {
              background: white;
              alternate-background-color: #fafbfe;
              border: 1px solid #dfe4ed;
              border-radius: 10px;
              outline: 0;
              selection-background-color: #e9eaed;
              selection-color: #242b48;
            }
            QTableWidget::item { padding: 6px; border-bottom: 1px solid #edf0f5; }
            QProgressBar {
              color: #3c4464;
              background: #eceff5;
              border: 0;
              border-radius: 7px;
              text-align: center;
              min-height: 18px;
            }
            QProgressBar::chunk {
              background: #6567df;
              border-radius: 7px;
            }
            QScrollBar:vertical {
              background: transparent;
              width: 10px;
              margin: 2px;
            }
            QScrollBar::handle:vertical {
              background: #cbd2df;
              border-radius: 5px;
              min-height: 28px;
            }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical { height: 0; }
            """
        )

    def _watch(
        self,
        future: concurrent.futures.Future[Any],
        callback: Callable[[Any], None] | None = None,
        *,
        error_callback: Callable[[BaseException], None] | None = None,
        timeout_seconds: float = 90,
        silent: bool = False,
        label: str = "操作",
    ) -> None:
        self.pending.append(
            PendingOperation(
                future=future,
                callback=callback,
                error_callback=error_callback,
                deadline=time.monotonic() + timeout_seconds,
                silent=silent,
                label=label,
            )
        )

    def _poll_futures(self) -> None:
        remaining: list[PendingOperation] = []
        now = time.monotonic()
        for operation in self.pending:
            if not operation.future.done() and now < operation.deadline:
                remaining.append(operation)
                continue
            try:
                if not operation.future.done():
                    operation.future.cancel()
                    raise TimeoutError(f"{operation.label}超时，已自动取消等待")
                result = operation.future.result()
                if operation.callback is not None:
                    operation.callback(result)
            except Exception as exc:
                if operation.error_callback is not None:
                    operation.error_callback(exc)
                if operation.silent:
                    self.engine_badge.setText("● 后台刷新暂不可用")
                    self.engine_badge.setToolTip(str(exc))
                else:
                    QMessageBox.critical(self, f"{operation.label}失败", str(exc))
        self.pending = remaining

    def refresh_all(self) -> None:
        current = self.sections.currentWidget()
        if current is self.tasks_page:
            self.ready_accounts_value.setText(
                str(len(self.backend.account_pool.discover_account_ids()))
            )
            self.refresh_jobs()
        elif current is self.accounts_page:
            self.refresh_accounts()
        elif current is self.history_page:
            self.refresh_history()
        elif current is self.schedules_page:
            self.refresh_schedules_page()
        elif current is self.platforms_page:
            self.refresh_platforms()

    def _on_section_changed(self, _: int) -> None:
        # 账号环境页不需要像任务页那样频繁刷新；未激活时降到 10 秒。
        if self.sections.currentWidget() is self.accounts_page:
            self.refresh_timer.setInterval(10000)
        else:
            self.refresh_timer.setInterval(3000)
        self.refresh_all()
        if self.sections.currentWidget() is self.long_tail_page:
            self.refresh_long_tail_options()

    def refresh_long_tail_options(self) -> None:
        def load() -> tuple[Any, ...]:
            return (
                self.backend.research_store.result_jobs(),
                self.backend.research_store.platforms(),
                self.backend.research_store.result_accounts(),
            )

        future = self.backend.call(load)

        def apply(payload: tuple[Any, ...]) -> None:
            jobs, platforms, accounts = payload
            self._update_combo(
                self.long_tail_job,
                [(f"{job['name']}（{job['result_count']}）", job["id"]) for job in jobs],
                "全部任务",
            )
            self._update_combo(
                self.long_tail_platform,
                [(value, value) for value in platforms],
                "全部平台",
            )
            self._update_combo(
                self.long_tail_account,
                [(value, value) for value in accounts],
                "全部账号",
            )

        self._watch(future, apply, label="刷新长尾选项", silent=True)

    def refresh_schedules_page(self) -> None:
        if self.refreshing_schedules:
            return
        self.refreshing_schedules = True

        def load() -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
            return (
                self.backend.research_store.list_job_templates(),
                self.backend.research_store.list_schedules(),
            )

        future = self.backend.call(load)

        def apply(payload: tuple[list[dict[str, Any]], list[dict[str, Any]]]) -> None:
            self.refreshing_schedules = False
            templates, schedules = payload
            self._refresh_templates_table(templates)
            self._refresh_schedules_table(schedules)
            self._update_combo(
                self.schedule_template_id,
                [(t["name"], t["id"]) for t in templates],
                "请选择模板",
            )

        def on_error(exc: BaseException) -> None:
            self.refreshing_schedules = False
            self.engine_badge.setText("● 定时任务刷新失败")
            self.engine_badge.setToolTip(str(exc))

        self._watch(future, apply, error_callback=on_error, label="刷新定时任务", silent=True)

    def _refresh_templates_table(self, templates: list[dict[str, Any]]) -> None:
        table = self.templates_table
        table.setRowCount(len(templates))
        for row, template in enumerate(templates):
            table.setItem(row, 0, QTableWidgetItem(str(template.get("name", ""))))
            keywords = template.get("keywords", [])
            count_item = QTableWidgetItem(str(len(keywords)))
            count_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            table.setItem(row, 1, count_item)
            table.setItem(row, 2, QTableWidgetItem(str(template.get("prompt_template", ""))))

            edit_button = QPushButton("编辑")
            edit_button.setObjectName("secondaryButton")
            edit_button.clicked.connect(lambda _, tid=template["id"]: self.edit_job_template(tid))
            table.setCellWidget(row, 3, edit_button)

            delete_button = QPushButton("删除")
            delete_button.setObjectName("dangerButton")
            delete_button.clicked.connect(
                lambda _, tid=template["id"]: self.delete_job_template(tid)
            )
            table.setCellWidget(row, 4, delete_button)

    def _refresh_schedules_table(self, schedules: list[dict[str, Any]]) -> None:
        table = self.schedules_table
        table.setRowCount(len(schedules))
        type_labels = {"interval": "按间隔", "once": "一次性", "daily": "每日定时"}
        for row, schedule in enumerate(schedules):
            table.setItem(row, 0, QTableWidgetItem(str(schedule.get("name", ""))))
            table.setItem(row, 1, QTableWidgetItem(str(schedule.get("template_name", ""))))
            type_text = type_labels.get(schedule.get("schedule_type", ""), "未知")
            table.setItem(row, 2, QTableWidgetItem(type_text))
            table.setItem(row, 3, QTableWidgetItem(str(schedule.get("schedule_value", ""))))
            table.setItem(row, 4, QTableWidgetItem(str(schedule.get("next_run_at", ""))))

            run_count_item = QTableWidgetItem(str(schedule.get("run_count", 0)))
            run_count_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            table.setItem(row, 5, run_count_item)

            enabled = bool(schedule.get("enabled", 0))
            toggle_button = QPushButton("禁用" if enabled else "启用")
            toggle_button.setObjectName("secondaryButton")
            toggle_button.clicked.connect(
                lambda _, sid=schedule["id"], en=enabled: self.toggle_schedule(sid, not en)
            )
            table.setCellWidget(row, 6, toggle_button)

            actions = QWidget()
            actions_layout = QHBoxLayout(actions)
            actions_layout.setContentsMargins(6, 2, 6, 2)
            actions_layout.setSpacing(8)
            run_now = QPushButton("立即执行")
            run_now.setObjectName("secondaryButton")
            run_now.clicked.connect(lambda _, sid=schedule["id"]: self.run_schedule_now(sid))
            delete = QPushButton("删除")
            delete.setObjectName("dangerButton")
            delete.clicked.connect(lambda _, sid=schedule["id"]: self.delete_schedule(sid))
            actions_layout.addWidget(run_now)
            actions_layout.addWidget(delete)
            actions_layout.addStretch()
            table.setCellWidget(row, 7, actions)

    def save_job_template(self) -> None:
        name = self.template_name.text().strip()
        keywords_text = self.template_keywords.toPlainText().strip()
        keywords = normalize_keywords(keywords_text)
        prompt_template = self.template_prompt_template.text().strip() or DEFAULT_PROMPT
        interval_seconds = self.template_interval_seconds.value()
        account_cooldown_seconds = self.template_account_cooldown_seconds.value()
        max_attempts = self.template_max_attempts.value()

        def save() -> dict[str, Any]:
            if self.editing_template_id:
                return self.backend.research_store.update_job_template(
                    self.editing_template_id,
                    name=name,
                    keywords=keywords,
                    prompt_template=prompt_template,
                    interval_seconds=interval_seconds,
                    account_cooldown_seconds=account_cooldown_seconds,
                    max_attempts=max_attempts,
                )
            return self.backend.research_store.create_job_template(
                name=name,
                keywords=keywords,
                prompt_template=prompt_template,
                interval_seconds=interval_seconds,
                account_cooldown_seconds=account_cooldown_seconds,
                max_attempts=max_attempts,
            )

        future = self.backend.call(save)

        def apply(_result: dict[str, Any]) -> None:
            self._reset_template_form()
            self.refresh_schedules_page()

        self._watch(future, apply, label="保存任务模板")

    def edit_job_template(self, template_id: str) -> None:
        def load() -> dict[str, Any]:
            return self.backend.research_store.get_job_template(template_id)

        future = self.backend.call(load)

        def apply(template: dict[str, Any]) -> None:
            self.editing_template_id = template_id
            self.template_name.setText(str(template.get("name", "")))
            self.template_keywords.setPlainText("\n".join(template.get("keywords", [])))
            self.template_prompt_template.setText(
                str(template.get("prompt_template", DEFAULT_PROMPT))
            )
            self.template_interval_seconds.setValue(int(template.get("interval_seconds", 10)))
            self.template_account_cooldown_seconds.setValue(
                int(template.get("account_cooldown_seconds", 0))
            )
            self.template_max_attempts.setValue(int(template.get("max_attempts", 2)))
            self.template_save_button.setText("更新模板")

        self._watch(future, apply, label="读取任务模板")

    def delete_job_template(self, template_id: str) -> None:
        reply = QMessageBox.question(
            self,
            "确认删除",
            "删除模板将同时删除引用它的所有定时计划，是否继续？",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if reply != QMessageBox.StandardButton.Yes:
            return

        def delete() -> None:
            self.backend.research_store.delete_job_template(template_id)

        future = self.backend.call(delete)

        def apply(_: None) -> None:
            if self.editing_template_id == template_id:
                self._reset_template_form()
            self.refresh_schedules_page()

        self._watch(future, apply, label="删除任务模板")

    def create_schedule(self) -> None:
        name = self.schedule_name.text().strip()
        template_id = self.schedule_template_id.currentData()
        schedule_type = self.schedule_type.currentData()
        schedule_value = self.schedule_value.text().strip()

        if not template_id:
            QMessageBox.warning(self, "提示", "请选择一个任务模板")
            return

        def save() -> dict[str, Any]:
            return self.backend.research_store.create_schedule(
                name=name,
                template_id=template_id,
                schedule_type=schedule_type,
                schedule_value=schedule_value,
            )

        future = self.backend.call(save)

        def apply(_result: dict[str, Any]) -> None:
            self.schedule_name.clear()
            self.schedule_value.clear()
            self.schedule_template_id.setCurrentIndex(0)
            self.refresh_schedules_page()

        self._watch(future, apply, label="创建触发计划")

    def toggle_schedule(self, schedule_id: str, enabled: bool) -> None:
        def toggle() -> dict[str, Any]:
            return self.backend.research_store.toggle_schedule(schedule_id, enabled)

        future = self.backend.call(toggle)
        self._watch(future, lambda _result: self.refresh_schedules_page(), label="切换计划状态")

    def delete_schedule(self, schedule_id: str) -> None:
        reply = QMessageBox.question(
            self,
            "确认删除",
            "确定删除该触发计划吗？",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if reply != QMessageBox.StandardButton.Yes:
            return

        def delete() -> None:
            self.backend.research_store.delete_schedule(schedule_id)

        future = self.backend.call(delete)
        self._watch(future, lambda _: self.refresh_schedules_page(), label="删除触发计划")

    def run_schedule_now(self, schedule_id: str) -> None:
        def run() -> dict[str, Any]:
            return self.backend.research_store.create_job_from_schedule(schedule_id)

        future = self.backend.call(run)

        def apply(job: dict[str, Any]) -> None:
            task_count = len(job.get("tasks", []))
            QMessageBox.information(
                self,
                "已触发",
                f"已按模板最新配置生成采集任务：{job.get('name', '')}\n共 {task_count} 个关键词",
            )
            self.backend.scheduler.wake()

        self._watch(future, apply, label="立即执行计划")

    def refresh_accounts(self) -> None:
        if self.refreshing_accounts:
            return
        self.refreshing_accounts = True
        future = self.backend.submit(self.backend.account_pool.snapshots())

        def apply(rows: list[dict[str, Any]]) -> None:
            self.refreshing_accounts = False
            self.engine_badge.setText("● 采集引擎运行中")
            self.engine_badge.setToolTip("")
            ready_count = sum(1 for row in rows if row["chat_ready"])
            opened_count = sum(1 for row in rows if row["started"])
            warning_count = sum(1 for row in rows if row.get("snapshot_error"))
            self.ready_accounts_value.setText(str(ready_count))
            self.accounts_summary.setText(
                f"共 {len(rows)} 个账号 · 已打开 {opened_count} · 可采集 {ready_count}"
                + (f" · {warning_count} 个状态读取超时" if warning_count else "")
            )
            signature = tuple(
                (
                    row["account_id"],
                    row["started"],
                    row["logged_in"],
                    row["chat_ready"],
                    row["needs_captcha"],
                    row.get("is_paused"),
                    row.get("tab_hidden"),
                    row.get("snapshot_error"),
                )
                for row in rows
            )
            if getattr(self, "_accounts_signature", None) == signature:
                return
            self._accounts_signature = signature
            self._clear_cards(self.accounts_cards_layout)
            for row in rows:
                self.accounts_cards_layout.addWidget(self._account_card(row))
            if not rows:
                empty = QLabel("暂无账号，请在上方创建账号并完成登录")
                empty.setObjectName("muted")
                empty.setAlignment(Qt.AlignmentFlag.AlignCenter)
                self.accounts_cards_layout.addWidget(empty)
            self.accounts_cards_layout.addStretch()

        def failed(_: BaseException) -> None:
            self.refreshing_accounts = False

        self._watch(
            future,
            apply,
            error_callback=failed,
            timeout_seconds=15,
            silent=True,
            label="刷新账号状态",
        )

    def create_account(self) -> None:
        value = self.new_account.text().strip()
        if not value:
            QMessageBox.information(self, "提示", "请填写账号名称")
            return
        try:
            account_id = normalize_account_id(value)
        except ValueError as exc:
            QMessageBox.warning(self, "账号名称无效", str(exc))
            return
        future = self.backend.submit(self.backend.account_pool.start_account(account_id))

        def completed(_: Any) -> None:
            self.new_account.clear()
            self.refresh_accounts()

        self._watch(future, completed)

    def open_account(self, account_id: str) -> None:
        self._watch(
            self.backend.submit(self.backend.account_pool.start_account(account_id)),
            lambda _: self.refresh_accounts(),
        )

    def focus_account(self, account_id: str) -> None:
        async def focus() -> None:
            self.backend.account_pool.set_tab_hidden(account_id, False)
            managed = self.backend.account_pool.get_if_started(account_id)
            if managed is not None:
                await managed.client.bring_to_front()

        self._watch(
            self.backend.submit(focus()),
            lambda _: self.refresh_accounts(),
        )

    def _on_captcha_detected(self, account_id: str) -> None:
        """Bring the offending account tab to the front so the user can solve the challenge."""
        self.focus_account(account_id)

    def toggle_account_tab_hidden(self, account_id: str, hidden: bool) -> None:
        async def toggle() -> None:
            self.backend.account_pool.set_tab_hidden(account_id, hidden)
            managed = self.backend.account_pool.get_if_started(account_id)
            if managed is not None:
                await managed.client.set_tab_visible(not hidden)

        self._watch(
            self.backend.submit(toggle()),
            lambda _: self.refresh_accounts(),
        )

    def stop_account(self, account_id: str) -> None:
        self._watch(
            self.backend.submit(self.backend.account_pool.stop_account(account_id)),
            lambda _: self.refresh_accounts(),
        )

    def clear_account_captcha(self, account_id: str) -> None:
        async def clear() -> None:
            managed = self.backend.account_pool.get_if_started(account_id)
            if managed is not None:
                await managed.client.reset_captcha()
            self.backend.research_store.resume_account(account_id)

        def applied(_: Any) -> None:
            self.refresh_accounts()
            self.backend.scheduler.wake()

        self._watch(
            self.backend.submit(clear()),
            applied,
            label="恢复账号验证状态",
        )

    def rename_account(self, account_id: str) -> None:
        value, accepted = QInputDialog.getText(
            self,
            "重命名账号",
            "新的账号名称：",
            text=account_id,
        )
        if not accepted or not value.strip() or value.strip() == account_id:
            return
        try:
            new_account_id = normalize_account_id(value)
        except ValueError as exc:
            QMessageBox.warning(self, "账号名称无效", str(exc))
            return

        async def rename() -> None:
            await self.backend.account_pool.rename_account(account_id, new_account_id)
            self.backend.research_store.rename_account_references(
                account_id,
                new_account_id,
            )

        self._watch(
            self.backend.submit(rename()),
            lambda _: self.refresh_accounts(),
        )

    def delete_account(self, account_id: str) -> None:
        answer = QMessageBox.question(
            self,
            "删除账号",
            f"确定删除账号“{account_id}”及其本地登录环境吗？\n此操作不会删除已经采集的历史结果。",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )
        if answer != QMessageBox.StandardButton.Yes:
            return

        async def delete() -> None:
            await self.backend.account_pool.delete_account(account_id)
            self.backend.research_store.remove_account_references(account_id)

        self._watch(
            self.backend.submit(delete()),
            lambda _: self.refresh_accounts(),
        )

    def import_keywords(self) -> None:
        filename, _ = QFileDialog.getOpenFileName(
            self,
            "导入关键词",
            "",
            "关键词文件 (*.xlsx *.csv *.tsv)",
        )
        if not filename:
            return
        try:
            values = parse_keyword_file(filename, Path(filename).read_bytes())
        except Exception as exc:
            QMessageBox.critical(self, "导入失败", str(exc))
            return
        current = self.keywords.toPlainText().strip()
        self.keywords.setPlainText("\n".join(([current] if current else []) + values))
        QMessageBox.information(self, "导入完成", f"已导入 {len(values)} 个关键词")

    def create_job(self) -> None:
        keywords = normalize_keywords(self.keywords.toPlainText().splitlines())
        if not keywords:
            QMessageBox.information(self, "提示", "请填写或导入关键词")
            return
        template = self.prompt_template.text().strip()
        if "{keyword}" not in template:
            QMessageBox.warning(self, "模板无效", "提问模板必须包含 {keyword}")
            return

        def create_and_wake() -> dict[str, Any]:
            job = self.backend.research_store.create_job(
                name=self.job_name.text(),
                keywords=keywords,
                account_ids=[],
                prompt_template=template,
                scheduled_at=None,
                interval_seconds=self.interval_seconds.value(),
                account_cooldown_seconds=0,
                max_attempts=self.max_attempts.value(),
            )
            self.backend.scheduler.wake()
            return job

        future = self.backend.call(create_and_wake)

        def completed(_: Any) -> None:
            self.keywords.clear()
            self.job_name.clear()
            self.refresh_jobs()
            QMessageBox.information(
                self,
                "采集任务",
                "任务已启动，系统会自动调度全部可用账号。",
            )

        self._watch(future, completed)

    def refresh_jobs(self) -> None:
        if self.refreshing_jobs:
            return
        self.refreshing_jobs = True

        def load() -> tuple[list[dict[str, Any]], dict[str, Any], int]:
            return (
                self.backend.research_store.list_jobs(),
                self.backend.scheduler.snapshot(),
                self.backend.research_store.result_count(),
            )

        future = self.backend.call(load)

        def apply(result: tuple[list[dict[str, Any]], dict[str, Any], int]) -> None:
            self.refreshing_jobs = False
            rows, scheduler, result_count = result
            rows = [row for row in rows if row["status"] in {"running", "paused"}]
            active_count = sum(1 for row in rows if row["status"] == "running")
            self.active_jobs_value.setText(str(active_count))
            self.active_workers_value.setText(str(scheduler["active_workers"]))
            self.total_results_value.setText(str(result_count))
            self.jobs_summary.setText(
                f"{active_count} 个任务运行中 · {scheduler['active_workers']} 个账号正在采集"
                if rows
                else "暂无运行中的任务，可以在上方创建新的采集任务"
            )
            if scheduler["last_error"]:
                self.engine_badge.setText("● 调度器已自动恢复")
                self.engine_badge.setToolTip(scheduler["last_error"])
            else:
                self.engine_badge.setText("● 采集引擎运行中")
                self.engine_badge.setToolTip("多账号任务引擎运行正常")
            self._clear_cards(self.jobs_cards_layout)
            for row in rows:
                self.jobs_cards_layout.addWidget(self._job_card(row))
            if not rows:
                empty = QLabel("暂无运行中的任务，可以在上方创建新的采集任务")
                empty.setObjectName("muted")
                empty.setAlignment(Qt.AlignmentFlag.AlignCenter)
                self.jobs_cards_layout.addWidget(empty)
            self.jobs_cards_layout.addStretch()

        def failed(_: BaseException) -> None:
            self.refreshing_jobs = False

        self._watch(
            future,
            apply,
            error_callback=failed,
            timeout_seconds=12,
            silent=True,
            label="刷新任务进度",
        )

    def set_job_status(self, job_id: str, status: str) -> None:
        def update() -> dict[str, Any]:
            job = self.backend.research_store.set_job_status(job_id, status)
            if status == "cancelled":
                self.backend.scheduler.cancel_job(job_id)
            self.backend.scheduler.wake()
            return job

        self._watch(
            self.backend.call(update),
            lambda _: self.refresh_jobs(),
            label="更新任务状态",
        )

    def refresh_history(self) -> None:
        if self.refreshing_history:
            return
        self.refreshing_history = True
        future = self.backend.call(self.backend.research_store.list_jobs)

        def apply(rows: list[dict[str, Any]]) -> None:
            self.refreshing_history = False
            rows = [row for row in rows if row["status"] in {"completed", "failed", "cancelled"}]
            self.history_table.setUpdatesEnabled(False)
            try:
                self.history_table.setRowCount(len(rows))
                for index, row in enumerate(rows):
                    values = [
                        row["name"],
                        STATUS_TEXT.get(row["status"], row["status"]),
                        str(row["total"]),
                        str(row["completed"]),
                        str(row["failed"]),
                        row["created_at"].replace("T", " ")[:19],
                    ]
                    for column, value in enumerate(values):
                        item = (
                            self._status_item(value, row["status"])
                            if column == 1
                            else QTableWidgetItem(value)
                        )
                        self.history_table.setItem(index, column, item)
                    box = QWidget()
                    actions = QHBoxLayout(box)
                    actions.setContentsMargins(2, 2, 2, 2)
                    view = QPushButton("查看结果")
                    view.clicked.connect(lambda _, jid=row["id"]: self.show_job_results(jid))
                    export = QPushButton("导出")
                    export.setObjectName("secondaryButton")
                    export.clicked.connect(
                        lambda _, jid=row["id"], name=row["name"]: self.export_job_results(
                            jid, name
                        )
                    )
                    rename = QPushButton("重命名")
                    rename.setObjectName("secondaryButton")
                    rename.clicked.connect(
                        lambda _, jid=row["id"], name=row["name"]: self.rename_job(jid, name)
                    )
                    delete = QPushButton("删除")
                    delete.setObjectName("dangerButton")
                    delete.clicked.connect(
                        lambda _, jid=row["id"], name=row["name"]: self.delete_job(jid, name)
                    )
                    actions.addWidget(view)
                    actions.addWidget(export)
                    actions.addWidget(rename)
                    actions.addWidget(delete)
                    self.history_table.setCellWidget(index, 6, box)
            finally:
                self.history_table.setUpdatesEnabled(True)

        def failed(_: BaseException) -> None:
            self.refreshing_history = False

        self._watch(
            future,
            apply,
            error_callback=failed,
            timeout_seconds=12,
            silent=True,
            label="刷新历史任务",
        )

    def delete_job(self, job_id: str, name: str) -> None:
        answer = QMessageBox.question(
            self,
            "删除任务",
            f"确定删除任务“{name}”吗？\n该任务已经采集的结果也会一并删除。",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )
        if answer != QMessageBox.StandardButton.Yes:
            return
        self._watch(
            self.backend.call(lambda: self.backend.research_store.delete_job(job_id)),
            lambda _: (self.refresh_jobs(), self.refresh_history(), self.refresh_results()),
        )

    def rename_job(self, job_id: str, current_name: str) -> None:
        text, ok = QInputDialog.getText(
            self,
            "重命名任务",
            "请输入新的任务名称：",
            text=current_name or "",
        )
        if not ok:
            return
        new_name = text.strip()
        if not new_name:
            QMessageBox.warning(self, "名称无效", "任务名称不能为空")
            return
        if new_name == (current_name or "").strip():
            return
        self._watch(
            self.backend.call(lambda: self.backend.research_store.rename_job(job_id, new_name)),
            lambda _: (self.refresh_jobs(), self.refresh_history()),
            label="重命名任务",
        )

    def sync_platform_info(self) -> None:
        self._watch(
            self.backend.call(self.backend.research_store.sync_platform_info),
            lambda count: (
                QMessageBox.information(
                    self,
                    "同步完成",
                    f"已更新 {count} 条记录的平台信息。",
                ),
                self.refresh_results(),
            ),
            label="同步平台信息",
        )

    def show_job_results(self, job_id: str) -> None:
        index = self.result_job.findData(job_id)
        if index < 0:
            self.result_job.addItem("所选任务", job_id)
            index = self.result_job.count() - 1
        self.result_job.setCurrentIndex(index)
        self.sections.setCurrentWidget(self.results_page)
        self.refresh_results()

    def refresh_results(self) -> None:
        if self.refreshing_results:
            return
        self.refreshing_results = True
        job_id = str(self.result_job.currentData() or "")
        keywords = self.result_keyword.selected_values()
        platform = str(self.result_platform.currentData() or "")
        account_id = str(self.result_account.currentData() or "")
        date_from = ""
        date_to = ""
        if self.result_date_enabled.isChecked():
            date_from = self.result_date_from.date().toString("yyyy-MM-dd")
            date_to = self.result_date_to.date().toString("yyyy-MM-dd")
            if date_from > date_to:
                self.refreshing_results = False
                QMessageBox.warning(self, "日期范围无效", "开始日期不能晚于结束日期")
                return

        def load() -> tuple[Any, ...]:
            filters = {
                "job_id": job_id,
                "keyword": keywords,
                "platform": platform,
                "account_id": account_id,
                "date_from": date_from,
                "date_to": date_to,
            }
            return (
                self.backend.research_store.list_results(
                    **filters,
                    limit=501,
                ),
                self.backend.research_store.result_dashboard(**filters),
                self.backend.research_store.platforms(),
                self.backend.research_store.result_jobs(),
                self.backend.research_store.result_accounts(),
                self.backend.research_store.result_keywords(),
            )

        future = self.backend.call(load)

        def apply(result: tuple[Any, ...]) -> None:
            self.refreshing_results = False
            rows, dashboard, platforms, jobs, accounts, keyword_options = result
            has_more = len(rows) > 500
            rows = rows[:500]
            summary = dashboard["summary"]
            filter_signature = (
                job_id,
                tuple(keywords),
                platform,
                account_id,
                date_from,
                date_to,
            )
            signature = (
                *filter_signature,
                tuple(row.get("id") for row in rows),
                tuple(summary.items()),
                tuple((row["platform"], row["count"]) for row in dashboard["platforms"]),
                tuple(platforms),
                tuple((job["id"], job["result_count"]) for job in jobs),
                tuple(accounts),
                tuple(keyword_options),
            )
            if signature == self.result_signature:
                return
            self.result_signature = signature

            old_rows = self.result_rows
            prepend_count = 0
            prepend_only = False
            if filter_signature == self.result_filter_signature and old_rows:
                for offset, row in enumerate(rows):
                    if row.get("id") == old_rows[0].get("id"):
                        remaining = len(rows) - offset
                        if remaining >= len(old_rows) and all(
                            rows[offset + index].get("id") == old_rows[index].get("id")
                            for index in range(len(old_rows))
                        ):
                            prepend_count = offset
                            prepend_only = True
                        break
            self.result_filter_signature = filter_signature
            self.result_rows = rows

            self._update_combo(
                self.result_job,
                [(f"{job['name']}（{job['result_count']}）", job["id"]) for job in jobs],
                "全部任务",
            )
            self._update_combo(
                self.result_platform,
                [(value, value) for value in platforms],
                "全部平台",
            )
            self._update_combo(
                self.result_account,
                [(value, value) for value in accounts],
                "全部账号",
            )
            self.result_keyword.set_options(keyword_options)
            self.result_total_value.setText(str(summary["total"]))
            self.result_task_value.setText(str(summary["jobs"]))
            self.result_keyword_value.setText(str(summary["keywords"]))
            self.result_platform_value.setText(str(summary["platforms"]))
            self.platform_distribution_rows = list(dashboard["platforms"])
            self.platform_distribution_total = int(summary["total"])
            self.source_distribution_chart.set_rows(dashboard["platforms"], int(summary["total"]))
            platform_count = len(dashboard["platforms"])
            self.distribution_hint.setText(
                f"共识别 {platform_count} 个平台；默认展示前 {SourceDistributionChart.TOP_N} 名，"
                f"其余 {max(0, platform_count - SourceDistributionChart.TOP_N)} 个平台合并为“其他”"
            )
            self.results_summary.setText(
                f"筛选结果 {summary['total']} 条 · 当前展示 {len(rows)} 条"
                + ("（仅展示前 500 条；导出包含全部筛选结果）" if has_more else "")
                + " · 双击链接可直接打开"
            )
            self.results_table.blockSignals(True)
            self.results_table.setUpdatesEnabled(False)
            try:
                self.results_table.horizontalHeader().setSectionResizeMode(
                    QHeaderView.ResizeMode.Fixed
                )
                if prepend_only:
                    for _ in range(prepend_count):
                        self.results_table.insertRow(0)
                    displayed = rows[:prepend_count]
                else:
                    self.results_table.setRowCount(len(rows))
                    displayed = rows
                for row_index, row in enumerate(displayed):
                    values = [
                        row["job_name"],
                        row["collected_date"],
                        row["keyword"],
                        row["title"][:200],
                        row["link"][:200],
                        row["platform"],
                    ]
                    for column, value in enumerate(values):
                        item = QTableWidgetItem(value)
                        if column in {3, 4}:
                            item.setToolTip(row["title"] if column == 3 else row["link"])
                        if column == 4:
                            item.setForeground(QColor("#4f51c8"))
                        self.results_table.setItem(row_index, column, item)
            finally:
                self.results_table.horizontalHeader().setSectionResizeMode(
                    QHeaderView.ResizeMode.ResizeToContents
                )
                self.results_table.blockSignals(False)
                self.results_table.setUpdatesEnabled(True)

        def failed(_: BaseException) -> None:
            self.refreshing_results = False

        self._watch(
            future,
            apply,
            error_callback=failed,
            timeout_seconds=15,
            silent=True,
            label="刷新采集结果",
        )

    def refresh_platforms(self) -> None:
        entries = all_entries()
        self.platform_info_summary.setText(f"共 {len(entries)} 条规则")
        self.platform_info_table.setUpdatesEnabled(False)
        try:
            self.platform_info_table.setRowCount(len(entries))
            for row_index, entry in enumerate(entries):
                values = [entry["domain"], entry["name"], entry["category"]]
                for column, value in enumerate(values):
                    item = QTableWidgetItem(value)
                    if column == 0:
                        item.setToolTip(value)
                    self.platform_info_table.setItem(row_index, column, item)
        finally:
            self.platform_info_table.setUpdatesEnabled(True)

    def import_platforms(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self,
            "导入平台信息",
            "",
            "Excel 文件 (*.xlsx *.xls)",
        )
        if not path:
            return

        try:
            workbook = load_workbook(path, read_only=True)
            sheet = workbook.active
            if sheet is None:
                raise ValueError("Excel 中没有工作表")
            rows = list(sheet.values)
            workbook.close()
        except Exception as exc:
            QMessageBox.critical(self, "导入失败", f"无法读取 Excel：\n{exc}")
            return

        if not rows:
            QMessageBox.warning(self, "导入失败", "Excel 为空")
            return

        header = [str(cell).strip() if cell is not None else "" for cell in rows[0]]

        def column_index(*candidates: str) -> int:
            for candidate in candidates:
                normalized = candidate.strip().lower()
                for index, value in enumerate(header):
                    if value.strip().lower() == normalized:
                        return index
            raise ValueError(f"缺少必要的列：{', '.join(candidates)}")

        try:
            url_index = column_index("url", "域名")
            name_index = column_index("平台名", "平台名称")
            category_index = column_index("平台类型", "类型")
        except ValueError as exc:
            QMessageBox.critical(self, "导入失败", str(exc))
            return

        data: list[dict[str, str]] = []
        for row in rows[1:]:
            if not row or all(cell is None or str(cell).strip() == "" for cell in row):
                continue
            data.append(
                {
                    "url": str(row[url_index]) if url_index < len(row) else "",
                    "平台名": str(row[name_index]) if name_index < len(row) else "",
                    "平台类型": str(row[category_index]) if category_index < len(row) else "",
                }
            )

        if not data:
            QMessageBox.warning(self, "导入失败", "未找到有效数据行")
            return

        try:
            result = add_entries(data)
        except Exception as exc:
            QMessageBox.critical(self, "导入失败", f"写入规则库时出错：\n{exc}")
            return

        self.refresh_platforms()
        QMessageBox.information(
            self,
            "导入完成",
            f"新增 {result['added']} 条平台规则，忽略 {result['ignored']} 条（已存在或格式错误）。",
        )

    def refresh_source_comparison(self) -> None:
        if self.refreshing_comparison:
            return
        self.refreshing_comparison = True
        job_ids_a = self.compare_jobs_a.selected_values()
        job_ids_b = self.compare_jobs_b.selected_values()
        keywords = self.compare_keyword.selected_values()
        account_id = str(self.compare_account.currentData() or "")
        has_groups = bool(job_ids_a) and bool(job_ids_b)

        def load() -> tuple[Any, ...]:
            jobs = self.backend.research_store.result_jobs()
            accounts = self.backend.research_store.result_accounts()
            keyword_options = self.backend.research_store.result_keywords()
            if not has_groups:
                return (None, jobs, accounts, keyword_options)
            result = self.backend.research_store.source_comparison(
                job_ids_a=job_ids_a,
                job_ids_b=job_ids_b,
                keyword=keywords,
                account_id=account_id,
            )
            return (result, jobs, accounts, keyword_options)

        future = self.backend.call(load)

        def apply(payload: tuple[Any, ...]) -> None:
            self.refreshing_comparison = False
            result, jobs, accounts, keyword_options = payload

            job_options = [(f"{job['name']}（{job['result_count']}）", job["id"]) for job in jobs]
            self.compare_jobs_a.set_options(job_options)
            self.compare_jobs_b.set_options(job_options)

            previous_account = self.compare_account.currentData()
            self.compare_account.clear()
            self.compare_account.addItem("全部账号", "")
            for value in accounts:
                self.compare_account.addItem(value, value)
            selected_account = self.compare_account.findData(previous_account)
            if selected_account >= 0:
                self.compare_account.setCurrentIndex(selected_account)
            self.compare_keyword.set_options(keyword_options)

            if result is None:
                return

            summary = result["summary"]
            rows = result["rows"]
            signature = (
                tuple(job_ids_a),
                tuple(job_ids_b),
                tuple(keywords),
                account_id,
                tuple(
                    (
                        row["platform"],
                        row["status"],
                        row["a_count"],
                        row["b_count"],
                    )
                    for row in rows
                ),
                tuple((job["id"], job["result_count"]) for job in jobs),
                tuple(accounts),
                tuple(keyword_options),
            )
            if signature == self.comparison_signature:
                return
            self.comparison_signature = signature

            self.compare_a_value.setText(str(summary["a_sources"]))
            self.compare_b_value.setText(str(summary["b_sources"]))
            self.compare_added_value.setText(str(summary["added_platforms"]))
            self.compare_removed_value.setText(str(summary["removed_platforms"]))

            self._compare_source_rows = {
                "added": [row for row in rows if row["status"] == "added"],
                "removed": [row for row in rows if row["status"] == "removed"],
                "continued": [row for row in rows if row["status"] == "continued"],
            }
            top_n = 10
            for status, label, button, empty_text in (
                (
                    "added",
                    self.compare_added_sources,
                    self.compare_added_more,
                    "暂无新增平台",
                ),
                (
                    "removed",
                    self.compare_removed_sources,
                    self.compare_removed_more,
                    "暂无掉出平台",
                ),
                (
                    "continued",
                    self.compare_continued_sources,
                    self.compare_continued_more,
                    "暂无持续平台",
                ),
            ):
                full = self._compare_source_rows[status]
                shown = full[:top_n]
                label.setText("  ·  ".join(str(row["platform"]) for row in shown) or empty_text)
                button.setVisible(bool(full))
                button.setText(f"查看全部（{len(full)}）" if full else "查看全部")

            self.comparison_summary.setText(
                f"A 任务群 {summary['a_sources']} 个平台 / {summary['a_total']} 次引用；"
                f"B 任务群 {summary['b_sources']} 个平台 / {summary['b_total']} 次引用；"
                f"新增 {summary['added_platforms']}，掉出 {summary['removed_platforms']}，"
                f"持续 {summary['continued_platforms']}"
            )
            self.comparison_table.setUpdatesEnabled(False)
            try:
                self.comparison_table.setRowCount(len(rows))
                for index, row in enumerate(rows):
                    status_text = {
                        "added": "新增",
                        "removed": "掉出",
                        "continued": "持续",
                    }[str(row["status"])]
                    values = [
                        row["platform"],
                        status_text,
                        str(row["a_count"]),
                        f"{float(row['a_share']):.1f}%",
                        str(row["b_count"]),
                        f"{float(row['b_share']):.1f}%",
                    ]
                    for column, value in enumerate(values):
                        item = QTableWidgetItem(value)
                        if column >= 1:
                            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                        if column == 1:
                            tone = {
                                "added": "#087a55",
                                "removed": "#b42c48",
                                "continued": "#59647a",
                            }[str(row["status"])]
                            item.setForeground(QColor(tone))
                        self.comparison_table.setItem(index, column, item)
            finally:
                self.comparison_table.setUpdatesEnabled(True)

        def failed(_: BaseException) -> None:
            self.refreshing_comparison = False

        self._watch(
            future,
            apply,
            error_callback=failed,
            timeout_seconds=15,
            silent=True,
            label="刷新信源对比",
        )

    def _selected_result_url(self) -> str:
        row = self.results_table.currentRow()
        if row < 0:
            return ""
        item = self.results_table.item(row, 4)
        return item.text().strip() if item else ""

    def copy_selected_link(self) -> None:
        link = self._selected_result_url()
        if not link:
            QMessageBox.information(self, "提示", "请先在结果表中选择一条资料")
            return
        QApplication.clipboard().setText(link)
        self.results_summary.setText("链接已复制到剪贴板")

    def open_selected_link(self) -> None:
        link = self._selected_result_url()
        if not link:
            QMessageBox.information(self, "提示", "请先在结果表中选择一条资料")
            return
        QDesktopServices.openUrl(QUrl(link))

    def _open_result_cell(self, row: int, column: int) -> None:
        if column != 4:
            return
        self.results_table.setCurrentCell(row, column)
        self.open_selected_link()

    def export_results(self) -> None:
        if not self.result_rows:
            QMessageBox.information(self, "提示", "当前没有可导出的结果")
            return
        filename, _ = QFileDialog.getSaveFileName(
            self,
            "导出采集结果",
            "豆包思考过程资料链接.xlsx",
            "Excel 工作簿 (*.xlsx)",
        )
        if not filename:
            return
        if not filename.casefold().endswith(".xlsx"):
            filename += ".xlsx"
        job_id = str(self.result_job.currentData() or "")
        keywords = self.result_keyword.selected_values()
        platform = str(self.result_platform.currentData() or "")
        account_id = str(self.result_account.currentData() or "")
        date_from = ""
        date_to = ""
        if self.result_date_enabled.isChecked():
            date_from = self.result_date_from.date().toString("yyyy-MM-dd")
            date_to = self.result_date_to.date().toString("yyyy-MM-dd")

        def export() -> int:
            rows = self.backend.research_store.list_results(
                job_id=job_id,
                keyword=keywords,
                platform=platform,
                account_id=account_id,
                date_from=date_from,
                date_to=date_to,
                limit=100000,
            )
            Path(filename).write_bytes(build_results_workbook(rows))
            return len(rows)

        future = self.backend.call(export)
        self._watch(
            future,
            lambda count: QMessageBox.information(
                self,
                "导出完成",
                f"已导出 {count} 条结果：\n{filename}",
            ),
            label="导出 Excel",
        )

    def export_job_results(self, job_id: str, name: str) -> None:
        default_name = f"{name or job_id}_结果_{QDate.currentDate().toString('yyyyMMdd')}.xlsx"
        filename, _ = QFileDialog.getSaveFileName(
            self,
            "导出任务结果",
            default_name,
            "Excel 工作簿 (*.xlsx)",
        )
        if not filename:
            return
        if not filename.casefold().endswith(".xlsx"):
            filename += ".xlsx"

        def export() -> int:
            rows = self.backend.research_store.list_results(
                job_id=job_id,
                keyword="",
                platform="",
                account_id="",
                date_from="",
                date_to="",
                limit=100000,
            )
            Path(filename).write_bytes(build_results_workbook(rows))
            return len(rows)

        future = self.backend.call(export)
        self._watch(
            future,
            lambda count: QMessageBox.information(
                self,
                "导出完成",
                f"已导出 {count} 条结果：\n{filename}",
            ),
            label="导出 Excel",
        )

    def analyze_long_tail(self) -> None:
        if self.analyzing_long_tail:
            return
        self.analyzing_long_tail = True
        job_id = str(self.long_tail_job.currentData() or "")
        platform = str(self.long_tail_platform.currentData() or "")
        account_id = str(self.long_tail_account.currentData() or "")
        date_from = ""
        date_to = ""
        if self.long_tail_date_enabled.isChecked():
            date_from = self.long_tail_date_from.date().toString("yyyy-MM-dd")
            date_to = self.long_tail_date_to.date().toString("yyyy-MM-dd")
        split_mode = str(self.long_tail_split_mode.currentData() or "threshold")
        breadth_threshold = self.long_tail_breadth_threshold.value()
        freq_threshold = self.long_tail_freq_threshold.value()
        density_threshold = self.long_tail_density_threshold.value()
        noise_density = self.long_tail_noise_density.value()
        log_scale = self.long_tail_log_scale.isChecked()
        x_log_scale = self.long_tail_x_log_scale.isChecked()

        def load() -> dict[str, Any]:
            return self.backend.research_store.long_tail_analysis(
                job_id=job_id,
                keyword="",
                platform=platform,
                account_id=account_id,
                date_from=date_from,
                date_to=date_to,
                split_mode=split_mode,
                breadth_threshold=breadth_threshold,
                freq_threshold=freq_threshold,
                density_threshold=density_threshold,
                noise_density_threshold=noise_density,
            )

        future = self.backend.call(load)

        def apply(data: dict[str, Any]) -> None:
            self.analyzing_long_tail = False
            self.long_tail_data = data
            self.long_tail_chart.set_data(data, log_scale=log_scale, x_log_scale=x_log_scale)
            self._fill_long_tail_tables(data)
            summary = data["summary"]
            params = data["params"]
            self.long_tail_summary.setText(
                f"总记录 {summary['total_records']} 条 · "
                f"识别 {summary['platform_count']} 个平台 · "
                f"目标长尾 {summary['target_count']} 个 · "
                f"虚假噪声 {summary['noise_count']} 个\n"
                f"当前阈值：广度≥{params['breadth_threshold']} · "
                f"频次≤{params['freq_threshold']} · "
                f"密度≤{params['density_threshold']} · "
                f"虚假密度≥{params['noise_density_threshold']}"
            )

        def failed(_: BaseException) -> None:
            self.analyzing_long_tail = False

        self._watch(
            future,
            apply,
            error_callback=failed,
            label="长尾信源分析",
        )

    def _fill_long_tail_tables(self, data: dict[str, Any]) -> None:
        targets = data["target_long_tail"]
        self.long_tail_target_table.setRowCount(len(targets))
        for row_index, row in enumerate(targets):
            values = [
                row["platform"],
                row["domain"],
                row["representative_link"],
                str(row["freq"]),
                str(row["breadth"]),
                str(row["density"]),
                row["type"] or "未分类",
                "、".join(row["keywords_sample"]),
            ]
            for column, value in enumerate(values):
                item = QTableWidgetItem(value)
                if column == 2 and value:
                    item.setToolTip(value)
                    item.setForeground(QColor("#4f51c8"))
                self.long_tail_target_table.setItem(row_index, column, item)

        all_rows: list[dict[str, Any]] = []
        quadrant_order = [
            "垂直长尾宝藏",
            "虚假长尾(噪声)",
            "头部主流媒体",
            "特定品类垂直站",
            "普通垂直信源",
            "一次性/僵尸信源",
        ]
        order_index = {name: index for index, name in enumerate(quadrant_order)}
        for quadrant, items in data["quadrants"].items():
            for item in items:
                all_rows.append({**item, "quadrant": quadrant})
        all_rows.sort(key=lambda row: (order_index.get(row["quadrant"], 99), -row["freq"]))

        self.long_tail_quadrant_table.setRowCount(len(all_rows))
        for row_index, row in enumerate(all_rows):
            values = [
                row["quadrant"],
                row["platform"],
                str(row["freq"]),
                str(row["breadth"]),
                str(row["density"]),
                row["type"] or "未分类",
            ]
            for column, value in enumerate(values):
                self.long_tail_quadrant_table.setItem(row_index, column, QTableWidgetItem(value))

    def export_long_tail_excel(self) -> None:
        if not self.long_tail_data:
            QMessageBox.information(self, "提示", "请先点击“分析长尾信源”")
            return
        targets = self.long_tail_data["target_long_tail"]
        if not targets:
            QMessageBox.information(self, "提示", "当前没有可导出的优质长尾信源")
            return
        filename, _ = QFileDialog.getSaveFileName(
            self,
            "导出优质长尾信源",
            f"优质长尾信源_{QDate.currentDate().toString('yyyyMMdd')}.xlsx",
            "Excel 工作簿 (*.xlsx)",
        )
        if not filename:
            return
        if not filename.casefold().endswith(".xlsx"):
            filename += ".xlsx"

        data = self.long_tail_data

        def export() -> int:
            params = data["params"]
            summary = data["summary"]
            workbook = Workbook()
            sheet = workbook.active
            if sheet is None:
                workbook.create_sheet()
                sheet = workbook.active
            assert sheet is not None
            sheet.append(
                [
                    "总参数",
                    f"总记录数：{summary['total_records']}",
                    f"目标长尾数：{summary['target_count']}",
                    (
                        f"阈值：广度≥{params['breadth_threshold']}，"
                        f"频次≤{params['freq_threshold']}，"
                        f"密度≤{params['density_threshold']}，"
                        f"虚假密度≥{params['noise_density_threshold']}"
                    ),
                ]
            )
            sheet.append(
                [
                    "URL（域名）",
                    "平台名",
                    "代表性链接",
                    "频次",
                    "广度",
                    "密度",
                    "平台类型",
                    "覆盖关键词",
                ]
            )
            for row in targets:
                sheet.append(
                    [
                        row["domain"],
                        row["platform"],
                        row["representative_link"],
                        row["freq"],
                        row["breadth"],
                        row["density"],
                        row["type"] or "未分类",
                        "、".join(row["keywords_sample"]),
                    ]
                )
            workbook.save(filename)
            return len(targets)

        future = self.backend.call(export)
        self._watch(
            future,
            lambda count: QMessageBox.information(
                self,
                "导出完成",
                f"已导出 {count} 个优质长尾信源：\n{filename}",
            ),
            label="导出优质长尾 Excel",
        )

    def copy_long_tail_keywords(self) -> None:
        if not self.long_tail_data:
            QMessageBox.information(self, "提示", "请先点击“分析长尾信源”")
            return
        names = [p["platform"] for p in self.long_tail_data["target_long_tail"]]
        if not names:
            QMessageBox.information(self, "提示", "当前没有识别到目标长尾信源")
            return
        text = "、".join(names)
        QApplication.clipboard().setText(text)
        QMessageBox.information(
            self,
            "已复制",
            f"已复制 {len(names)} 个平台限定词到剪贴板",
        )

    def _open_long_tail_link(self, row: int, _: int) -> None:
        item = self.long_tail_target_table.item(row, 2)
        if item is None:
            return
        link = item.text()
        if link:
            QDesktopServices.openUrl(QUrl(link))
