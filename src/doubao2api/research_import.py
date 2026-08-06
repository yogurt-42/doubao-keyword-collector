from __future__ import annotations

import csv
import io
from pathlib import Path

from openpyxl import load_workbook

KEYWORD_HEADERS = {"关键词", "关键字", "提问关键词", "keyword", "keywords", "query"}


def normalize_keywords(values: list[object]) -> list[str]:
    output: list[str] = []
    seen: set[str] = set()
    for value in values:
        text = str(value or "").strip()
        if not text:
            continue
        key = text.casefold()
        if key in KEYWORD_HEADERS or key in seen:
            continue
        seen.add(key)
        output.append(text)
    return output


def _xlsx_keywords(data: bytes) -> list[str]:
    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        column = 0
        start_row = 0
        for row_index, row in enumerate(rows[:20]):
            for column_index, value in enumerate(row):
                if str(value or "").strip().casefold() in KEYWORD_HEADERS:
                    column = column_index
                    start_row = row_index + 1
                    break
            else:
                continue
            break
        return normalize_keywords([row[column] for row in rows[start_row:] if len(row) > column])
    finally:
        workbook.close()


def _delimited_keywords(data: bytes, suffix: str) -> list[str]:
    try:
        text = data.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = data.decode("gb18030")
    delimiter = "\t" if suffix == ".tsv" else ","
    rows = list(csv.reader(io.StringIO(text), delimiter=delimiter))
    if not rows:
        return []
    column = 0
    start_row = 0
    for row_index, row in enumerate(rows[:20]):
        for column_index, value in enumerate(row):
            if value.strip().casefold() in KEYWORD_HEADERS:
                column = column_index
                start_row = row_index + 1
                break
        else:
            continue
        break
    return normalize_keywords([row[column] for row in rows[start_row:] if len(row) > column])


def parse_keyword_file(filename: str, data: bytes) -> list[str]:
    suffix = Path(filename).suffix.casefold()
    if suffix == ".xlsx":
        return _xlsx_keywords(data)
    if suffix in {".csv", ".tsv"}:
        return _delimited_keywords(data, suffix)
    raise ValueError("仅支持 .xlsx、.csv 和 .tsv 关键词文件")


def preview_keyword_file(filename: str, data: bytes, max_rows: int = 20) -> list[list[str]]:
    """Return the first few raw rows of a keyword file for UI preview."""

    suffix = Path(filename).suffix.casefold()
    if suffix == ".xlsx":
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        try:
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
            return [
                [str(cell) if cell is not None else "" for cell in row]
                for row in rows[:max_rows]
            ]
        finally:
            workbook.close()
    if suffix in {".csv", ".tsv"}:
        try:
            text = data.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = data.decode("gb18030")
        delimiter = "\t" if suffix == ".tsv" else ","
        rows = list(csv.reader(io.StringIO(text), delimiter=delimiter))
        return rows[:max_rows]
    raise ValueError("仅支持 .xlsx、.csv 和 .tsv 关键词文件")
